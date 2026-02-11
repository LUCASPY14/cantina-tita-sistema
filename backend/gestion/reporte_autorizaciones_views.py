"""
Vistas de reportes para autorizaciones de saldo negativo

Este módulo contiene vistas para generar reportes y análisis
de las autorizaciones de saldo negativo.

Autor: CantiTita
Fecha: 2026-01-12
"""

from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum, Avg, Q, F
from django.db.models.functions import TruncDate
from django.utils import timezone
from datetime import timedelta, datetime
from decimal import Decimal
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from gestion.models import AutorizacionSaldoNegativo, Empleado
from gestion.permisos import solo_gerente_o_superior
from django.core.paginator import Paginator


@login_required
@solo_gerente_o_superior
def reporte_autorizaciones_saldo_negativo(request):
    """
    Dashboard completo con métricas y gráficos de autorizaciones
    """
    # Filtros
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    regularizado = request.GET.get('regularizado')
    supervisor_id = request.GET.get('supervisor')
    
    # Query base
    query = AutorizacionSaldoNegativo.objects.select_related(
        'id_venta',
        'nro_tarjeta',
        'id_empleado_autoriza'
    ).all()
    
    # Aplicar filtros
    filtros_activos = {}
    
    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            query = query.filter(fecha_autorizacion__gte=fecha_desde_dt)
            filtros_activos['fecha_desde'] = fecha_desde
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            fecha_hasta_dt = fecha_hasta_dt.replace(hour=23, minute=59, second=59)
            query = query.filter(fecha_autorizacion__lte=fecha_hasta_dt)
            filtros_activos['fecha_hasta'] = fecha_hasta
        except ValueError:
            pass
    
    if regularizado in ['0', '1']:
        query = query.filter(regularizado=(regularizado == '1'))
        filtros_activos['regularizado'] = regularizado
    
    if supervisor_id:
        query = query.filter(id_empleado_autoriza_id=supervisor_id)
        filtros_activos['supervisor'] = supervisor_id
    
    # Verificar exportación Excel
    if request.GET.get('export') == 'excel':
        return exportar_autorizaciones_excel(query, filtros_activos)
    
    # Calcular estadísticas
    stats = calcular_estadisticas(query)
    
    # Datos para gráficos
    chart_data = generar_datos_graficos(query)
    
    # Agregar días de deuda a cada autorización
    hoy = timezone.now()
    autorizaciones_list = list(query.order_by('-fecha_autorizacion'))
    for auth in autorizaciones_list:
        if not auth.regularizado:
            auth.dias_deuda = (hoy - auth.fecha_autorizacion).days
        else:
            auth.dias_deuda = 0
    
    # Paginación
    paginator = Paginator(autorizaciones_list, 20)
    page_number = request.GET.get('page', 1)
    autorizaciones = paginator.get_page(page_number)
    
    # Query string para paginación
    filtros_query = '&'.join([f'{k}={v}' for k, v in filtros_activos.items()])
    if filtros_query:
        filtros_query = '&' + filtros_query
    
    # Supervisores para filtro
    supervisores = Empleado.objects.filter(
        activo=True,
        id_rol__nombre_rol__in=['ADMINISTRADOR', 'GERENTE']
    ).order_by('nombre', 'apellido')
    
    context = {
        'autorizaciones': autorizaciones,
        'stats': stats,
        'chart_data': json.dumps(chart_data),
        'filtros': filtros_activos,
        'filtros_query': filtros_query,
        'supervisores': supervisores,
    }
    
    return render(request, 'pos/reportes/autorizaciones_saldo_negativo.html', context)


def calcular_estadisticas(query):
    """Calcular todas las estadísticas del reporte"""
    hoy = timezone.now()
    
    # Totales
    total_autorizaciones = query.count()
    pendientes = query.filter(regularizado=False).count()
    regularizados = query.filter(regularizado=True).count()
    
    # Montos
    monto_total_deuda = query.filter(regularizado=False).aggregate(
        total=Sum('monto_autorizado')
    )['total'] or Decimal('0')
    
    monto_total_regularizado = query.filter(regularizado=True).aggregate(
        total=Sum('monto_autorizado')
    )['total'] or Decimal('0')
    
    promedio_deuda = query.aggregate(
        avg=Avg('monto_autorizado')
    )['avg'] or Decimal('0')
    
    # Tiempo promedio de regularización
    regularizadas = query.filter(regularizado=True, fecha_regularizacion__isnull=False)
    if regularizadas.exists():
        suma_dias = 0
        count = 0
        for auth in regularizadas:
            dias = (auth.fecha_regularizacion - auth.fecha_autorizacion).days
            suma_dias += dias
            count += 1
        promedio_dias_regularizacion = suma_dias / count if count > 0 else 0
    else:
        promedio_dias_regularizacion = 0
    
    # Tasa de regularización
    tasa_regularizacion = (regularizados / total_autorizaciones * 100) if total_autorizaciones > 0 else 0
    
    # Autorizaciones hoy
    autorizaciones_hoy = query.filter(
        fecha_autorizacion__date=hoy.date()
    ).count()
    
    return {
        'total_autorizaciones': total_autorizaciones,
        'pendientes': pendientes,
        'regularizados': regularizados,
        'monto_total_deuda': monto_total_deuda,
        'monto_total_regularizado': monto_total_regularizado,
        'promedio_deuda': promedio_deuda,
        'promedio_dias_regularizacion': promedio_dias_regularizacion,
        'tasa_regularizacion': tasa_regularizacion,
        'autorizaciones_hoy': autorizaciones_hoy,
    }


def generar_datos_graficos(query):
    """Generar datos para todos los gráficos"""
    
    # Top 10 Supervisores
    top_supervisores = query.values(
        'id_empleado_autoriza__nombre',
        'id_empleado_autoriza__apellido'
    ).annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    supervisores_labels = [
        f"{s['id_empleado_autoriza__nombre']} {s['id_empleado_autoriza__apellido']}"
        for s in top_supervisores
    ]
    supervisores_data = [s['total'] for s in top_supervisores]
    
    # Estados
    regularizados = query.filter(regularizado=True).count()
    pendientes = query.filter(regularizado=False).count()
    
    # Tendencia últimos 30 días
    fecha_inicio = timezone.now() - timedelta(days=30)
    tendencia = query.filter(
        fecha_autorizacion__gte=fecha_inicio
    ).annotate(
        fecha=TruncDate('fecha_autorizacion')
    ).values('fecha').annotate(
        total=Count('id')
    ).order_by('fecha')
    
    # Crear array con todos los días (llenar huecos con 0)
    tendencia_dict = {item['fecha']: item['total'] for item in tendencia}
    tendencia_labels = []
    tendencia_data = []
    
    for i in range(30):
        fecha = (timezone.now() - timedelta(days=29-i)).date()
        tendencia_labels.append(fecha.strftime('%d/%m'))
        tendencia_data.append(tendencia_dict.get(fecha, 0))
    
    # Top 10 Estudiantes
    top_estudiantes = query.values(
        'estudiante_nombre'
    ).annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    estudiantes_labels = [s['estudiante_nombre'] for s in top_estudiantes]
    estudiantes_data = [s['total'] for s in top_estudiantes]
    
    # Tiempo de regularización (rangos)
    regularizadas = query.filter(regularizado=True, fecha_regularizacion__isnull=False)
    tiempo_rangos = [0, 0, 0, 0, 0]  # 0-3, 4-7, 8-15, 16-30, 30+
    
    for auth in regularizadas:
        dias = (auth.fecha_regularizacion - auth.fecha_autorizacion).days
        if dias <= 3:
            tiempo_rangos[0] += 1
        elif dias <= 7:
            tiempo_rangos[1] += 1
        elif dias <= 15:
            tiempo_rangos[2] += 1
        elif dias <= 30:
            tiempo_rangos[3] += 1
        else:
            tiempo_rangos[4] += 1
    
    return {
        'top_supervisores': {
            'labels': supervisores_labels,
            'data': supervisores_data
        },
        'estados': {
            'regularizados': regularizados,
            'pendientes': pendientes
        },
        'tendencia': {
            'labels': tendencia_labels,
            'data': tendencia_data
        },
        'top_estudiantes': {
            'labels': estudiantes_labels,
            'data': estudiantes_data
        },
        'tiempo_regularizacion': tiempo_rangos
    }


def exportar_autorizaciones_excel(query, filtros):
    """Exportar autorizaciones a Excel"""
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Autorizaciones"
    
    # Estilos
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Headers
    headers = [
        'ID', 'Fecha Autorización', 'Tarjeta', 'Estudiante', 
        'ID Venta', 'Monto Autorizado', 'Saldo Anterior', 'Saldo Nuevo',
        'Supervisor', 'Motivo', 'Estado', 'Fecha Regularización', 
        'ID Recarga', 'Días Deuda'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    hoy = timezone.now()
    for row_num, auth in enumerate(query.order_by('-fecha_autorizacion'), 2):
        ws.cell(row=row_num, column=1, value=auth.id)
        ws.cell(row=row_num, column=2, value=auth.fecha_autorizacion.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row_num, column=3, value=auth.nro_tarjeta.nro_tarjeta)
        ws.cell(row=row_num, column=4, value=auth.estudiante_nombre)
        ws.cell(row=row_num, column=5, value=auth.id_venta.id_venta if auth.id_venta else '')
        ws.cell(row=row_num, column=6, value=float(auth.monto_autorizado))
        ws.cell(row=row_num, column=7, value=float(auth.saldo_anterior))
        ws.cell(row=row_num, column=8, value=float(auth.saldo_nuevo))
        ws.cell(row=row_num, column=9, value=f"{auth.id_empleado_autoriza.nombre} {auth.id_empleado_autoriza.apellido}")
        ws.cell(row=row_num, column=10, value=auth.motivo_autorizacion)
        ws.cell(row=row_num, column=11, value='Regularizado' if auth.regularizado else 'Pendiente')
        ws.cell(row=row_num, column=12, value=auth.fecha_regularizacion.strftime('%d/%m/%Y') if auth.fecha_regularizacion else '')
        ws.cell(row=row_num, column=13, value=auth.id_carga_regularizacion.id_carga if auth.id_carga_regularizacion else '')
        
        dias_deuda = (hoy - auth.fecha_autorizacion).days if not auth.regularizado else 0
        ws.cell(row=row_num, column=14, value=dias_deuda)
    
    # Ajustar anchos de columna
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=autorizaciones_saldo_negativo_{timezone.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(response)
    return response
