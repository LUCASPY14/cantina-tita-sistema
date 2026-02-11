# Funciones auxiliares para exportación de reportes
from django.http import HttpResponse
from datetime import datetime
import csv
import json


def exportar_reporte_formato(datos, columnas, titulo, stats, formato, fecha_desde, fecha_hasta, tipo_reporte):
    """Exporta reporte en el formato solicitado"""
    
    if formato == 'csv':
        return exportar_csv(datos, columnas, titulo)
    elif formato == 'excel':
        return exportar_excel_simple(datos, columnas, titulo, fecha_desde, fecha_hasta)
    elif formato == 'pdf':
        return exportar_pdf_simple(datos, columnas, titulo, fecha_desde, fecha_hasta, stats)
    
    return HttpResponse('Formato no válido', status=400)


def exportar_csv(datos, columnas, titulo):
    """Exporta datos a CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="reporte_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # Agregar BOM para Excel
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # Título
    writer.writerow([titulo])
    writer.writerow([])
    
    # Encabezados
    writer.writerow(columnas)
    
    # Datos
    for fila in datos:
        writer.writerow(fila)
    
    return response


def exportar_excel_simple(datos, columnas, titulo, fecha_desde, fecha_hasta):
    """Exporta datos a Excel usando openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import io
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        # Título
        ws.merge_cells('A1:' + get_column_letter(len(columnas)) + '1')
        titulo_cell = ws['A1']
        titulo_cell.value = titulo
        titulo_cell.font = Font(size=16, bold=True, color="FFFFFF")
        titulo_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Período
        ws['A2'] = f"Período: {fecha_desde} - {fecha_hasta}"
        ws['A2'].font = Font(italic=True)
        
        # Encabezados
        for col_num, columna in enumerate(columnas, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = columna
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        for row_num, fila in enumerate(datos, 5):
            for col_num, valor in enumerate(fila, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = valor
                
                # Alternar colores de fila
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # Ajustar anchos de columna
        for col_num in range(1, len(columnas) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Crear respuesta
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
        
    except ImportError:
        # Si openpyxl no está disponible, usar CSV
        return exportar_csv(datos, columnas, titulo)


def exportar_pdf_simple(datos, columnas, titulo, fecha_desde, fecha_hasta, stats):
    """Exporta datos a PDF usando reportlab"""
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        import io
        
        # Crear buffer
        buffer = io.BytesIO()
        
        # Crear documento (landscape para más columnas)
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.5*inch
        )
        
        # Elementos del documento
        elements = []
        styles = getSampleStyleSheet()
        
        # Estilo personalizado para título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        # Título
        elements.append(Paragraph(titulo, title_style))
        elements.append(Paragraph(f"Período: {fecha_desde} - {fecha_hasta}", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Estadísticas si existen
        if stats and stats.get('total'):
            stats_text = f"<b>Total registros:</b> {stats.get('total', 0)}"
            if stats.get('monto_total'):
                stats_text += f" | <b>Monto Total:</b> {int(stats['monto_total']):,} Gs."
            elements.append(Paragraph(stats_text, styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))
        
        # Preparar datos para tabla
        tabla_datos = [columnas] + datos
        
        # Limitar número de filas si es muy grande
        if len(tabla_datos) > 51:  # 1 header + 50 filas
            tabla_datos = tabla_datos[:51]
            elements.append(Paragraph("<i>Mostrando primeros 50 registros</i>", styles['Italic']))
            elements.append(Spacer(1, 0.1*inch))
        
        # Crear tabla
        tabla = Table(tabla_datos)
        
        # Estilo de tabla
        tabla.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Cuerpo
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            
            # Bordes
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#1f2937')),
        ]))
        
        elements.append(tabla)
        
        # Pie de página con fecha
        elements.append(Spacer(1, 0.3*inch))
        fecha_generacion = datetime.now().strftime("%d/%m/%Y %H:%M")
        elements.append(Paragraph(
            f"<i>Generado el {fecha_generacion} - MetrePay</i>",
            ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_RIGHT)
        ))
        
        # Construir PDF
        doc.build(elements)
        
        # Obtener valor del buffer
        pdf = buffer.getvalue()
        buffer.close()
        
        # Crear respuesta
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        response.write(pdf)
        
        return response
        
    except ImportError:
        # Si reportlab no está disponible, usar CSV
        return exportar_csv(datos, columnas, titulo)
