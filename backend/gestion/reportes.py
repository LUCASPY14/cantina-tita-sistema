"""
Sistema de Generación de Reportes para Cantina Tita
Soporta PDF y Excel con filtros por fecha y periodo
Incluye gráficos visuales con matplotlib
"""

from django.http import HttpResponse
from django.db.models import Sum, Count, Q, Avg
from datetime import datetime, date, timedelta
from decimal import Decimal
import io
import os
import tempfile

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

import matplotlib
matplotlib.use('Agg')  # Backend sin GUI
import matplotlib.pyplot as plt
from matplotlib.patches import Wedge
import matplotlib.patches as mpatches

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from pos.models import Venta as Ventas, DetalleVenta, PagoVenta as PagosVenta
from gestion.models import (
    Producto, Cliente, Tarjeta,
    ConsumoTarjeta, CargasSaldo, StockUnico, VistaStockAlerta,
    VistaSaldoClientes, Proveedor,
    Empleado, Categoria, Compras,
    PagosProveedores, AplicacionPagosVentas, AplicacionPagosCompras,
    NotasCreditoCliente, NotasCreditoProveedor
)


class ReportesPDF:
    """Generador de reportes en PDF"""
    
    @staticmethod
    def _crear_header(story, titulo, subtitulo=None):
        """Crea el encabezado del reporte"""
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        story.append(Paragraph(titulo, title_style))
        
        if subtitulo:
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Normal'],
                fontSize=12,
                textColor=colors.HexColor('#7f8c8d'),
                spaceAfter=20,
                alignment=TA_CENTER
            )
            story.append(Paragraph(subtitulo, subtitle_style))
        
        story.append(Spacer(1, 0.3 * inch))
    
    @staticmethod
    def _aplicar_estilo_tabla(tabla, header_color='#3498db'):
        """Aplica estilos a una tabla"""
        style = TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(header_color)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Body
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ])
        tabla.setStyle(style)
        return tabla
    
    @staticmethod
    def _generar_grafico_barras(datos, labels, titulo, color='#3498db', width=6, height=4):
        """Genera un gráfico de barras y retorna la imagen"""
        fig, ax = plt.subplots(figsize=(width, height))
        ax.bar(labels, datos, color=color, alpha=0.7, edgecolor='black')
        ax.set_title(titulo, fontsize=14, fontweight='bold', pad=20)
        ax.set_ylabel('Monto (Gs.)', fontsize=10)
        ax.grid(axis='y', alpha=0.3, linestyle='--')
        
        # Rotar labels si son muchos
        if len(labels) > 5:
            plt.xticks(rotation=45, ha='right')
        
        plt.tight_layout()
        
        # Guardar en buffer
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
        buf.seek(0)
        plt.close()
        
        return buf
    
    @staticmethod
    def _generar_grafico_linea(datos, labels, titulo, color='#2ecc71', width=6, height=4):
        """Genera un gráfico de línea y retorna la imagen"""
        fig, ax = plt.subplots(figsize=(width, height))
        ax.plot(labels, datos, color=color, linewidth=2, marker='o', markersize=6, alpha=0.8)
        ax.fill_between(range(len(datos)), datos, alpha=0.2, color=color)
        ax.set_title(titulo, fontsize=14, fontweight='bold', pad=20)
        ax.set_ylabel('Monto (Gs.)', fontsize=10)
        ax.grid(True, alpha=0.3, linestyle='--')
        
        # Rotar labels si son muchos
        if len(labels) > 5:
            plt.xticks(rotation=45, ha='right')
        
        plt.tight_layout()
        
        # Guardar en buffer
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
        buf.seek(0)
        plt.close()
        
        return buf
    
    @staticmethod
    def _generar_grafico_torta(datos, labels, titulo, width=6, height=6):
        """Genera un gráfico de torta y retorna la imagen"""
        fig, ax = plt.subplots(figsize=(width, height))
        
        colors_list = ['#3498db', '#2ecc71', '#e74c3c', '#f39c12', '#9b59b6', 
                      '#1abc9c', '#34495e', '#e67e22', '#95a5a6', '#16a085']
        
        wedges, texts, autotexts = ax.pie(
            datos, 
            labels=labels, 
            autopct='%1.1f%%',
            startangle=90,
            colors=colors_list[:len(datos)],
            textprops={'fontsize': 9}
        )
        
        # Mejorar legibilidad
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontweight('bold')
        
        ax.set_title(titulo, fontsize=14, fontweight='bold', pad=20)
        plt.tight_layout()
        
        # Guardar en buffer
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
        buf.seek(0)
        plt.close()
        
        return buf
    
    @staticmethod
    def reporte_ventas(fecha_inicio=None, fecha_fin=None, id_cliente=None, id_cajero=None, 
                       estado=None, id_tipo_pago=None, monto_minimo=None, monto_maximo=None):
        """Genera reporte de ventas en PDF con filtros avanzados"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_ventas_{date.today()}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter)
        story = []
        
        # Fechas
        if not fecha_inicio:
            fecha_inicio = date.today()
        if not fecha_fin:
            fecha_fin = date.today()
        
        # Header
        titulo = "📊 Reporte de Ventas - Cantina Tita"
        subtitulo = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        
        # Agregar info de filtros al subtítulo
        if id_cliente:
            try:
                cliente = Cliente.objects.get(pk=id_cliente)
                subtitulo += f" | Cliente: {cliente.nombre_completo}"
            except Cliente.DoesNotExist:
                pass
        if id_cajero:
            try:
                cajero = Empleado.objects.get(pk=id_cajero)
                subtitulo += f" | Cajero: {cajero.nombre} {cajero.apellido}"
            except Empleado.DoesNotExist:
                pass
        if estado:
            subtitulo += f" | Estado: {estado}"
        
        ReportesPDF._crear_header(story, titulo, subtitulo)
        
        # Filtrar ventas
        ventas = Ventas.objects.filter(
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin
        ).select_related('id_cliente', 'id_empleado_cajero', 'id_tipo_pago')
        
        # Aplicar filtros adicionales
        if id_cliente:
            ventas = ventas.filter(id_cliente_id=id_cliente)
        if id_cajero:
            ventas = ventas.filter(id_empleado_cajero_id=id_cajero)
        if estado:
            ventas = ventas.filter(estado=estado)
        if id_tipo_pago:
            ventas = ventas.filter(id_tipo_pago_id=id_tipo_pago)
        if monto_minimo:
            ventas = ventas.filter(monto_total__gte=int(monto_minimo))
        if monto_maximo:
            ventas = ventas.filter(monto_total__lte=int(monto_maximo))
        
        ventas = ventas.order_by('-fecha')[:50]
        
        # Resumen
        total_ventas = ventas.aggregate(
            total=Sum('monto_total'),
            cantidad=Count('id_venta')
        )
        
        data_resumen = [
            ['Métrica', 'Valor'],
            ['Total Ventas', f"Gs. {total_ventas['total'] or 0:,.0f}"],
            ['Cantidad Transacciones', f"{total_ventas['cantidad'] or 0}"],
            ['Promedio por Venta', f"Gs. {(total_ventas['total'] or 0) / (total_ventas['cantidad'] or 1):,.0f}"],
        ]
        
        tabla_resumen = Table(data_resumen, colWidths=[3*inch, 3*inch])
        ReportesPDF._aplicar_estilo_tabla(tabla_resumen, '#27ae60')
        story.append(tabla_resumen)
        story.append(Spacer(1, 0.3 * inch))
        
        # GRÁFICO: Ventas por día
        if total_ventas['cantidad'] > 0:
            from django.db.models.functions import TruncDate
            
            ventas_por_dia = ventas.annotate(
                dia=TruncDate('fecha')
            ).values('dia').annotate(
                total=Sum('monto_total')
            ).order_by('dia')[:10]
            
            if len(ventas_por_dia) > 1:
                story.append(Paragraph("<b>Evolución de Ventas Diarias</b>", getSampleStyleSheet()['Heading2']))
                story.append(Spacer(1, 0.2 * inch))
                
                labels = [item['dia'].strftime('%d/%m') for item in ventas_por_dia]
                datos = [float(item['total']) for item in ventas_por_dia]
                
                img_buf = ReportesPDF._generar_grafico_linea(
                    datos, labels, 
                    'Ventas por Día (Guaraníes)',
                    color='#2ecc71'
                )
                img = Image(img_buf, width=5*inch, height=3*inch)
                story.append(img)
                story.append(Spacer(1, 0.3 * inch))
        
        # Detalle de ventas
        story.append(Paragraph("<b>Detalle de Ventas</b>", getSampleStyleSheet()['Heading2']))
        story.append(Spacer(1, 0.2 * inch))
        
        data_detalle = [['Fecha', 'Cliente', 'Monto', 'Cajero']]
        for venta in ventas[:50]:  # Limitar a 50 para no saturar
            data_detalle.append([
                venta.fecha.strftime('%d/%m/%Y %H:%M'),
                f"{venta.id_cliente.nombres} {venta.id_cliente.apellidos}",
                f"Gs. {venta.monto_total:,.0f}",
                f"{venta.id_empleado_cajero.nombre if venta.id_empleado_cajero else 'N/A'}"
            ])
        
        if len(data_detalle) > 1:
            tabla_detalle = Table(data_detalle, colWidths=[1.5*inch, 2*inch, 1.5*inch, 1.5*inch])
            ReportesPDF._aplicar_estilo_tabla(tabla_detalle)
            story.append(tabla_detalle)
        else:
            story.append(Paragraph("No hay ventas en el período seleccionado", getSampleStyleSheet()['Normal']))
        
        doc.build(story)
        return response
    
    @staticmethod
    def reporte_productos_vendidos(fecha_inicio=None, fecha_fin=None, id_categoria=None):
        """Genera reporte de productos más vendidos con filtro de categoría"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_productos_{date.today()}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter)
        story = []
        
        # Fechas
        if not fecha_inicio:
            fecha_inicio = date.today()
        if not fecha_fin:
            fecha_fin = date.today()
        
        # Header
        titulo = "🏆 Reporte de Productos Más Vendidos"
        subtitulo = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        
        if id_categoria:
            try:
                categoria = Categoria.objects.get(pk=id_categoria)
                subtitulo += f" | Categoría: {categoria.nombre}"
            except Categoria.DoesNotExist:
                pass
        
        ReportesPDF._crear_header(story, titulo, subtitulo)
        
        # Consultar productos vendidos
        productos_query = DetalleVenta.objects.filter(
            id_venta__fecha__date__gte=fecha_inicio,
            id_venta__fecha__date__lte=fecha_fin,
            id_venta__estado='Completada'
        )
        
        # Filtro de categoría
        if id_categoria:
            productos_query = productos_query.filter(id_producto__id_categoria_id=id_categoria)
        
        productos = productos_query.values(
            'id_producto__codigo',
            'id_producto__descripcion'
        ).annotate(
            cantidad_total=Sum('cantidad'),
            monto_total=Sum('subtotal_total'),
            num_ventas=Count('id_detalle')
        ).order_by('-cantidad_total')[:30]
        
        # Resumen
        total_productos = len(productos)
        total_cantidad = sum(p['cantidad_total'] for p in productos)
        total_monto = sum(p['monto_total'] for p in productos)
        
        data_resumen = [
            ['Métrica', 'Valor'],
            ['Productos Vendidos', f"{total_productos}"],
            ['Unidades Totales', f"{total_cantidad:.0f}"],
            ['Monto Total', f"Gs. {total_monto:,.0f}"],
        ]
        
        tabla_resumen = Table(data_resumen, colWidths=[3*inch, 3*inch])
        ReportesPDF._aplicar_estilo_tabla(tabla_resumen, '#27ae60')
        story.append(tabla_resumen)
        story.append(Spacer(1, 0.3 * inch))
        
        # GRÁFICO: Top 10 productos
        if len(productos) > 0:
            story.append(Paragraph("<b>Top 10 Productos Más Vendidos</b>", getSampleStyleSheet()['Heading2']))
            story.append(Spacer(1, 0.2 * inch))
            
            top_productos = list(productos)[:10]
            labels = [p['id_producto__descripcion'][:15] + '...' if len(p['id_producto__descripcion']) > 15 
                     else p['id_producto__descripcion'] for p in top_productos]
            datos = [float(p['cantidad_total']) for p in top_productos]
            
            img_buf = ReportesPDF._generar_grafico_barras(
                datos, labels,
                'Cantidad de Unidades Vendidas',
                color='#e74c3c'
            )
            img = Image(img_buf, width=5*inch, height=3*inch)
            story.append(img)
            story.append(Spacer(1, 0.3 * inch))
        
        # Tabla
        story.append(Paragraph("<b>Detalle de Productos</b>", getSampleStyleSheet()['Heading2']))
        story.append(Spacer(1, 0.2 * inch))
        
        data = [['#', 'Código', 'Producto', 'Cantidad', 'Total Vendido', 'Ventas']]
        for idx, producto in enumerate(productos, 1):
            data.append([
                str(idx),
                producto['id_producto__codigo'] or 'N/A',
                producto['id_producto__descripcion'],
                f"{producto['cantidad_total']:.0f}",
                f"Gs. {producto['monto_total']:,.0f}",
                str(producto['num_ventas'])
            ])
        
        if len(data) > 1:
            tabla = Table(data, colWidths=[0.4*inch, 1*inch, 2.5*inch, 1*inch, 1.5*inch, 0.8*inch])
            ReportesPDF._aplicar_estilo_tabla(tabla, '#e74c3c')
            story.append(tabla)
        else:
            story.append(Paragraph("No hay productos vendidos en el período", getSampleStyleSheet()['Normal']))
        
        doc.build(story)
        return response
    
    @staticmethod
    def reporte_inventario():
        """Genera reporte de inventario actual"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_inventario_{date.today()}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter)
        story = []
        
        # Header
        titulo = "📦 Reporte de Inventario - Cantina Tita"
        subtitulo = f"Fecha: {date.today().strftime('%d/%m/%Y')}"
        ReportesPDF._crear_header(story, titulo, subtitulo)
        
        # Alertas de stock
        alertas = VistaStockAlerta.objects.all().order_by('-nivel_alerta', 'stock_actual')[:30]
        
        # GRÁFICO: Distribución de alertas
        if alertas:
            criticos = sum(1 for a in alertas if a.stock_actual == 0)
            bajos = sum(1 for a in alertas if a.stock_actual > 0 and a.stock_actual < a.stock_minimo)
            ok = len(alertas) - criticos - bajos
            
            if criticos + bajos > 0:
                story.append(Paragraph("<b>Distribución de Alertas de Stock</b>", getSampleStyleSheet()['Heading2']))
                story.append(Spacer(1, 0.2 * inch))
                
                labels = []
                datos = []
                if criticos > 0:
                    labels.append(f'Crítico ({criticos})')
                    datos.append(criticos)
                if bajos > 0:
                    labels.append(f'Bajo ({bajos})')
                    datos.append(bajos)
                if ok > 0:
                    labels.append(f'Normal ({ok})')
                    datos.append(ok)
                
                img_buf = ReportesPDF._generar_grafico_torta(
                    datos, labels,
                    'Estado del Inventario'
                )
                img = Image(img_buf, width=4*inch, height=4*inch)
                story.append(img)
                story.append(Spacer(1, 0.3 * inch))
        
        if alertas:
            story.append(Paragraph("<b>⚠️ Alertas de Stock Crítico</b>", getSampleStyleSheet()['Heading2']))
            story.append(Spacer(1, 0.2 * inch))
            
            data = [['Producto', 'Stock Actual', 'Stock Mín.', 'Faltante', 'Estado']]
            for alerta in alertas:
                estado = '🔴 CRÍTICO' if alerta.stock_actual == 0 else '🟠 BAJO' if alerta.stock_actual < alerta.stock_minimo else '🟢 OK'
                data.append([
                    alerta.producto,
                    f"{alerta.stock_actual:.0f}",
                    f"{alerta.stock_minimo:.0f}",
                    f"{alerta.cantidad_faltante:.0f}",
                    estado
                ])
            
            tabla = Table(data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch, 1*inch])
            ReportesPDF._aplicar_estilo_tabla(tabla, '#e67e22')
            story.append(tabla)
        else:
            story.append(Paragraph("✅ No hay alertas de stock", getSampleStyleSheet()['Normal']))
        
        doc.build(story)
        return response
    
    @staticmethod
    def reporte_consumos_tarjeta(fecha_inicio=None, fecha_fin=None):
        """Genera reporte de consumos por tarjeta"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_consumos_{date.today()}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter)
        story = []
        
        # Fechas
        if not fecha_inicio:
            fecha_inicio = date.today()
        if not fecha_fin:
            fecha_fin = date.today()
        
        # Header
        titulo = "💳 Reporte de Consumos por Tarjeta"
        subtitulo = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        ReportesPDF._crear_header(story, titulo, subtitulo)
        
        # Consultar consumos
        consumos = ConsumoTarjeta.objects.filter(
            fecha_consumo__date__gte=fecha_inicio,
            fecha_consumo__date__lte=fecha_fin
        ).select_related('nro_tarjeta').order_by('-fecha_consumo')[:50]
        
        # Resumen
        total_consumos = consumos.aggregate(
            total=Sum('monto_consumido'),
            cantidad=Count('id_consumo')
        )
        
        data_resumen = [
            ['Métrica', 'Valor'],
            ['Total Consumido', f"Gs. {total_consumos['total'] or 0:,.0f}"],
            ['Cantidad Consumos', f"{total_consumos['cantidad'] or 0}"],
        ]
        
        tabla_resumen = Table(data_resumen, colWidths=[3*inch, 3*inch])
        ReportesPDF._aplicar_estilo_tabla(tabla_resumen, '#9b59b6')
        story.append(tabla_resumen)
        story.append(Spacer(1, 0.3 * inch))
        
        # GRÁFICO: Consumos por día
        if total_consumos['cantidad'] > 0:
            from django.db.models.functions import TruncDate
            
            consumos_por_dia = consumos.annotate(
                dia=TruncDate('fecha_consumo')
            ).values('dia').annotate(
                total=Sum('monto_consumido')
            ).order_by('dia')[:10]
            
            if len(consumos_por_dia) > 1:
                story.append(Paragraph("<b>Consumos Diarios</b>", getSampleStyleSheet()['Heading2']))
                story.append(Spacer(1, 0.2 * inch))
                
                labels = [item['dia'].strftime('%d/%m') for item in consumos_por_dia]
                datos = [float(item['total']) for item in consumos_por_dia]
                
                img_buf = ReportesPDF._generar_grafico_barras(
                    datos, labels,
                    'Consumo por Día (Guaraníes)',
                    color='#9b59b6'
                )
                img = Image(img_buf, width=5*inch, height=3*inch)
                story.append(img)
                story.append(Spacer(1, 0.3 * inch))
        
        # Detalle
        story.append(Paragraph("<b>Detalle de Consumos</b>", getSampleStyleSheet()['Heading2']))
        story.append(Spacer(1, 0.2 * inch))
        
        data = [['Fecha', 'Tarjeta', 'Monto', 'Saldo Anterior', 'Saldo Nuevo']]
        for consumo in consumos:
            data.append([
                consumo.fecha_consumo.strftime('%d/%m/%Y %H:%M'),
                consumo.nro_tarjeta.nro_tarjeta,
                f"Gs. {consumo.monto_consumido:,.0f}",
                f"Gs. {consumo.saldo_anterior:,.0f}",
                f"Gs. {consumo.saldo_posterior:,.0f}"
            ])
        
        if len(data) > 1:
            tabla = Table(data, colWidths=[1.5*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
            ReportesPDF._aplicar_estilo_tabla(tabla)
            story.append(tabla)
        else:
            story.append(Paragraph("No hay consumos en el período", getSampleStyleSheet()['Normal']))
        
        doc.build(story)
        return response
    
    @staticmethod
    def reporte_clientes():
        """Genera reporte de clientes con saldos"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_clientes_{date.today()}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter)
        story = []
        
        # Header
        titulo = "👥 Reporte de Clientes - Cantina Tita"
        subtitulo = f"Fecha: {date.today().strftime('%d/%m/%Y')}"
        ReportesPDF._crear_header(story, titulo, subtitulo)
        
        # Clientes con saldo
        queryset_completo = VistaSaldoClientes.objects.all().order_by('-saldo_actual')
        
        # Resumen (calculamos antes del slice)
        total_clientes = Cliente.objects.filter(activo=True).count()
        con_saldo = queryset_completo.filter(saldo_actual__gt=0).count()
        
        # Ahora sí tomamos los primeros 40 para el detalle
        clientes = queryset_completo[:40]
        
        data_resumen = [
            ['Métrica', 'Valor'],
            ['Total Clientes Activos', str(total_clientes)],
            ['Clientes con Saldo a Favor', str(con_saldo)],
        ]
        
        tabla_resumen = Table(data_resumen, colWidths=[3*inch, 3*inch])
        ReportesPDF._aplicar_estilo_tabla(tabla_resumen, '#1abc9c')
        story.append(tabla_resumen)
        story.append(Spacer(1, 0.3 * inch))
        
        # GRÁFICO: Top 10 clientes con mayor saldo
        if con_saldo > 0:
            story.append(Paragraph("<b>Top 10 Clientes con Mayor Saldo</b>", getSampleStyleSheet()['Heading2']))
            story.append(Spacer(1, 0.2 * inch))
            
            top_clientes = list(queryset_completo.filter(saldo_actual__gt=0)[:10])
            labels = [c.nombre_completo[:15] + '...' if len(c.nombre_completo) > 15 
                     else c.nombre_completo for c in top_clientes]
            datos = [float(c.saldo_actual) for c in top_clientes]
            
            img_buf = ReportesPDF._generar_grafico_barras(
                datos, labels,
                'Saldo Actual (Guaraníes)',
                color='#1abc9c'
            )
            img = Image(img_buf, width=5*inch, height=3*inch)
            story.append(img)
            story.append(Spacer(1, 0.3 * inch))
        
        # Detalle
        story.append(Paragraph("<b>Clientes con Saldo</b>", getSampleStyleSheet()['Heading2']))
        story.append(Spacer(1, 0.2 * inch))
        
        data = [['Cliente', 'RUC/CI', 'Tipo', 'Saldo Actual']]
        for cliente in clientes:
            if cliente.saldo_actual != 0:
                data.append([
                    cliente.nombre_completo,
                    cliente.ruc_ci,
                    cliente.tipo_cliente,
                    f"Gs. {cliente.saldo_actual:,.0f}"
                ])
        
        if len(data) > 1:
            tabla = Table(data, colWidths=[2.5*inch, 1.5*inch, 1.2*inch, 1.3*inch])
            ReportesPDF._aplicar_estilo_tabla(tabla)
            story.append(tabla)
        else:
            story.append(Paragraph("No hay clientes con saldo", getSampleStyleSheet()['Normal']))
        
        doc.build(story)
        return response
    
    @staticmethod
    def reporte_cta_corriente_cliente(id_cliente=None, fecha_inicio=None, fecha_fin=None):
        """Genera reporte de cuenta corriente de cliente en PDF usando el nuevo sistema"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="cta_corriente_cliente_{date.today()}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter)
        story = []
        
        # Header
        cliente_nombre = "Todos los Clientes"
        if id_cliente:
            try:
                cliente = Cliente.objects.get(pk=id_cliente)
                cliente_nombre = cliente.nombre_completo
            except Cliente.DoesNotExist:
                cliente_nombre = "Cliente no encontrado"
        
        titulo = f"Cuenta Corriente - {cliente_nombre}"
        subtitulo = f"Fecha: {date.today().strftime('%d/%m/%Y')}"
        ReportesPDF._crear_header(story, titulo, subtitulo)
        
        # Query ventas pendientes
        ventas = Ventas.objects.filter(
            estado_pago__in=['PENDIENTE', 'PARCIAL']
        ).select_related('id_cliente', 'id_empleado_cajero')
        
        if id_cliente:
            ventas = ventas.filter(id_cliente_id=id_cliente)
        
        if fecha_inicio and fecha_fin:
            ventas = ventas.filter(fecha__date__gte=fecha_inicio, fecha__date__lte=fecha_fin)
        
        ventas = ventas.order_by('id_cliente', 'fecha')[:200]
        
        # Tabla
        data = [['Fecha', 'Cliente', 'Venta #', 'Total', 'Saldo Pend.', 'Estado']]
        total_pendiente = Decimal('0')
        
        for venta in ventas:
            data.append([
                venta.fecha.strftime('%d/%m/%Y'),
                venta.id_cliente.nombre_completo[:30],
                str(venta.id_venta),
                f"Gs. {venta.monto_total:,.0f}",
                f"Gs. {venta.saldo_pendiente:,.0f}",
                venta.estado_pago
            ])
            total_pendiente += venta.saldo_pendiente
        
        # Total
        data.append(['', '', '', 'TOTAL PENDIENTE:', f"Gs. {total_pendiente:,.0f}", ''])
        
        tabla = Table(data, colWidths=[1.2*inch, 1.8*inch, 0.8*inch, 1.2*inch, 1.2*inch, 1*inch])
        ReportesPDF._aplicar_estilo_tabla(tabla)
        story.append(tabla)
        
        doc.build(story)
        return response
    
    @staticmethod
    def reporte_cta_corriente_proveedor(id_proveedor=None, fecha_inicio=None, fecha_fin=None):
        """Genera reporte de cuenta corriente de proveedor en PDF usando el nuevo sistema"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="cta_corriente_proveedor_{date.today()}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter)
        story = []
        
        # Header
        proveedor_nombre = "Todos los Proveedores"
        if id_proveedor:
            try:
                proveedor = Proveedor.objects.get(pk=id_proveedor)
                proveedor_nombre = proveedor.razon_social
            except Proveedor.DoesNotExist:
                proveedor_nombre = "Proveedor no encontrado"
        
        titulo = f"Cuenta Corriente - {proveedor_nombre}"
        subtitulo = f"Fecha: {date.today().strftime('%d/%m/%Y')}"
        ReportesPDF._crear_header(story, titulo, subtitulo)
        
        # Query compras pendientes
        compras = Compras.objects.filter(
            estado_pago__in=['PENDIENTE', 'PARCIAL']
        ).select_related('id_proveedor')
        
        if id_proveedor:
            compras = compras.filter(id_proveedor_id=id_proveedor)
        
        if fecha_inicio and fecha_fin:
            compras = compras.filter(fecha__date__gte=fecha_inicio, fecha__date__lte=fecha_fin)
        
        compras = compras.order_by('id_proveedor', 'fecha')[:200]
        
        # Tabla
        data = [['Fecha', 'Proveedor', 'Compra #', 'Total', 'Saldo Pend.', 'Estado']]
        total_pendiente = Decimal('0')
        
        for compra in compras:
            data.append([
                compra.fecha.strftime('%d/%m/%Y'),
                compra.id_proveedor.razon_social[:30],
                str(compra.id_compra),
                f"Gs. {compra.monto_total:,.0f}",
                f"Gs. {compra.saldo_pendiente:,.0f}",
                compra.estado_pago
            ])
            total_pendiente += compra.saldo_pendiente
        
        # Total
        data.append(['', '', '', 'TOTAL PENDIENTE:', f"Gs. {total_pendiente:,.0f}", ''])
        
        tabla = Table(data, colWidths=[1.2*inch, 1.8*inch, 0.8*inch, 1.2*inch, 1.2*inch, 1*inch])
        ReportesPDF._aplicar_estilo_tabla(tabla)
        story.append(tabla)
        
        doc.build(story)
        return response


class ReportesExcel:
    """Generador de reportes en Excel"""
    
    @staticmethod
    def _aplicar_estilo_header(ws, row=1):
        """Aplica estilos al header de Excel"""
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    @staticmethod
    def _ajustar_columnas(ws):
        """Ajusta el ancho de las columnas automáticamente"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def reporte_ventas(fecha_inicio=None, fecha_fin=None):
        """Genera reporte de ventas en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"
        
        # Fechas
        if not fecha_inicio:
            fecha_inicio = date.today()
        if not fecha_fin:
            fecha_fin = date.today()
        
        # Título
        ws['A1'] = 'Reporte de Ventas - Cantina Tita'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        
        # Header
        headers = ['Fecha', 'Cliente', 'RUC/CI', 'Monto Total', 'Estado', 'Tipo Venta', 'Cajero']
        ws.append([])
        ws.append(headers)
        ReportesExcel._aplicar_estilo_header(ws, row=4)
        
        # Datos
        ventas = Ventas.objects.filter(
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin,
            estado='Completada'
        ).select_related('id_cliente', 'id_empleado_cajero')
        
        for venta in ventas:
            ws.append([
                venta.fecha.strftime('%d/%m/%Y %H:%M'),
                f"{venta.id_cliente.nombres} {venta.id_cliente.apellidos}",
                venta.id_cliente.ruc_ci,
                float(venta.monto_total),
                venta.estado,
                venta.tipo_venta,
                venta.id_empleado_cajero.nombre if venta.id_empleado_cajero else 'N/A'
            ])
        
        # Formato números
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = '#,##0'
        
        # Totales
        ws.append([])
        total_row = ws.max_row + 1
        ws[f'A{total_row}'] = 'TOTAL'
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'D{total_row}'] = f'=SUM(D5:D{ws.max_row-1})'
        ws[f'D{total_row}'].font = Font(bold=True)
        ws[f'D{total_row}'].number_format = '#,##0'
        
        ReportesExcel._ajustar_columnas(ws)
        
        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_ventas_{date.today()}.xlsx"'
        wb.save(response)
        return response
    
    @staticmethod
    def reporte_productos_vendidos(fecha_inicio=None, fecha_fin=None):
        """Genera reporte de productos en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Productos Vendidos"
        
        # Fechas
        if not fecha_inicio:
            fecha_inicio = date.today()
        if not fecha_fin:
            fecha_fin = date.today()
        
        # Título
        ws['A1'] = 'Reporte de Productos Más Vendidos'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        
        # Header
        headers = ['#', 'Código', 'Producto', 'Cantidad Vendida', 'Total Vendido', 'Núm. Ventas']
        ws.append([])
        ws.append(headers)
        ReportesExcel._aplicar_estilo_header(ws, row=4)
        
        # Datos
        productos = DetalleVenta.objects.filter(
            id_venta__fecha__date__gte=fecha_inicio,
            id_venta__fecha__date__lte=fecha_fin,
            id_venta__estado='Completada'
        ).values(
            'id_producto__codigo',
            'id_producto__descripcion'
        ).annotate(
            cantidad_total=Sum('cantidad'),
            monto_total=Sum('subtotal_total'),
            num_ventas=Count('id_detalle')
        ).order_by('-cantidad_total')
        
        for idx, producto in enumerate(productos, 1):
            ws.append([
                idx,
                producto['id_producto__codigo'] or 'N/A',
                producto['id_producto__descripcion'],
                float(producto['cantidad_total']),
                float(producto['monto_total']),
                producto['num_ventas']
            ])
        
        # Formato números
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = '#,##0'
        
        ReportesExcel._ajustar_columnas(ws)
        
        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_productos_{date.today()}.xlsx"'
        wb.save(response)
        return response
    
    @staticmethod
    def reporte_inventario():
        """Genera reporte de inventario en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"
        
        # Título
        ws['A1'] = 'Reporte de Inventario - Cantina Tita'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Fecha: {date.today().strftime('%d/%m/%Y')}"
        
        # Header
        headers = ['Producto', 'Código', 'Categoría', 'Stock Actual', 'Stock Mínimo', 'Faltante', 'Nivel Alerta']
        ws.append([])
        ws.append(headers)
        ReportesExcel._aplicar_estilo_header(ws, row=4)
        
        # Datos
        productos = VistaStockAlerta.objects.all().order_by('-nivel_alerta', 'stock_actual')
        
        for producto in productos:
            ws.append([
                producto.producto,
                producto.codigo or 'N/A',
                producto.categoria,
                float(producto.stock_actual),
                float(producto.stock_minimo),
                float(producto.cantidad_faltante),
                producto.nivel_alerta
            ])
        
        ReportesExcel._ajustar_columnas(ws)
        
        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_inventario_{date.today()}.xlsx"'
        wb.save(response)
        return response
    
    @staticmethod
    def reporte_consumos_tarjeta(fecha_inicio=None, fecha_fin=None):
        """Genera reporte de consumos en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Consumos Tarjeta"
        
        # Fechas
        if not fecha_inicio:
            fecha_inicio = date.today()
        if not fecha_fin:
            fecha_fin = date.today()
        
        # Título
        ws['A1'] = 'Reporte de Consumos por Tarjeta'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        
        # Header
        headers = ['Fecha', 'Nro. Tarjeta', 'Monto Consumido', 'Saldo Anterior', 'Saldo Posterior', 'Detalle']
        ws.append([])
        ws.append(headers)
        ReportesExcel._aplicar_estilo_header(ws, row=4)
        
        # Datos
        consumos = ConsumoTarjeta.objects.filter(
            fecha_consumo__date__gte=fecha_inicio,
            fecha_consumo__date__lte=fecha_fin
        ).select_related('nro_tarjeta').order_by('-fecha_consumo')
        
        for consumo in consumos:
            ws.append([
                consumo.fecha_consumo.strftime('%d/%m/%Y %H:%M'),
                consumo.nro_tarjeta.nro_tarjeta,
                float(consumo.monto_consumido),
                float(consumo.saldo_anterior),
                float(consumo.saldo_posterior),
                consumo.detalle or 'N/A'
            ])
        
        # Formato números
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=3, max_col=5):
            for cell in row:
                cell.number_format = '#,##0'
        
        ReportesExcel._ajustar_columnas(ws)
        
        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_consumos_{date.today()}.xlsx"'
        wb.save(response)
        return response
    
    @staticmethod
    def reporte_clientes():
        """Genera reporte de clientes en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        # Título
        ws['A1'] = 'Reporte de Clientes - Cantina Tita'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Fecha: {date.today().strftime('%d/%m/%Y')}"
        
        # Header
        headers = ['Cliente', 'RUC/CI', 'Tipo Cliente', 'Saldo Actual', 'Total Movimientos', 'Última Actualización']
        ws.append([])
        ws.append(headers)
        ReportesExcel._aplicar_estilo_header(ws, row=4)
        
        # Datos
        clientes = VistaSaldoClientes.objects.all().order_by('-saldo_actual')
        
        for cliente in clientes:
            ws.append([
                cliente.nombre_completo,
                cliente.ruc_ci,
                cliente.tipo_cliente,
                float(cliente.saldo_actual),
                cliente.total_movimientos,
                cliente.ultima_actualizacion.strftime('%d/%m/%Y') if cliente.ultima_actualizacion else 'N/A'
            ])
        
        # Formato números
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = '#,##0'
        
        ReportesExcel._ajustar_columnas(ws)
        
        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_clientes_{date.today()}.xlsx"'
        wb.save(response)
        return response
    
    @staticmethod
    def reporte_cta_corriente_cliente(id_cliente=None, fecha_inicio=None, fecha_fin=None):
        """Genera reporte de cuenta corriente de cliente en Excel usando el nuevo sistema"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Cta Corriente Cliente"
        
        # Título
        if id_cliente:
            try:
                cliente = Cliente.objects.get(pk=id_cliente)
                ws['A1'] = f'Cuenta Corriente - {cliente.nombre_completo}'
                ws['A2'] = f"RUC/CI: {cliente.ruc_ci}"
            except Cliente.DoesNotExist:
                ws['A1'] = 'Cuenta Corriente - Cliente'
                ws['A2'] = 'Cliente no encontrado'
        else:
            ws['A1'] = 'Cuenta Corriente - Todos los Clientes'
        
        ws['A1'].font = Font(bold=True, size=14)
        
        if fecha_inicio and fecha_fin:
            ws['A3'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        else:
            ws['A3'] = f"Fecha: {date.today().strftime('%d/%m/%Y')}"
        
        # Headers
        headers = ['Fecha', 'Cliente', 'RUC/CI', 'Venta #', 'Total Venta', 'Saldo Pendiente', 'Estado Pago']
        ws.append([])
        ws.append(headers)
        ReportesExcel._aplicar_estilo_header(ws, row=ws.max_row)
        
        # Query ventas pendientes
        ventas = Ventas.objects.filter(
            estado_pago__in=['PENDIENTE', 'PARCIAL']
        ).select_related('id_cliente')
        
        if id_cliente:
            ventas = ventas.filter(id_cliente_id=id_cliente)
        
        if fecha_inicio and fecha_fin:
            ventas = ventas.filter(fecha__date__gte=fecha_inicio, fecha__date__lte=fecha_fin)
        
        ventas = ventas.order_by('id_cliente', 'fecha')[:200]
        
        # Datos
        total_pendiente = Decimal('0')
        for venta in ventas:
            ws.append([
                venta.fecha.strftime('%d/%m/%Y %H:%M'),
                venta.id_cliente.nombre_completo,
                venta.id_cliente.ruc_ci,
                venta.id_venta,
                float(venta.monto_total),
                float(venta.saldo_pendiente),
                venta.estado_pago
            ])
            total_pendiente += venta.saldo_pendiente
        
        # Fila de total
        ws.append(['', '', '', '', 'TOTAL PENDIENTE:', float(total_pendiente), ''])
        
        # Formato números
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=5, max_col=6):
            for cell in row:
                cell.number_format = '#,##0'
        
        ReportesExcel._ajustar_columnas(ws)
        
        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="cta_corriente_cliente_{date.today()}.xlsx"'
        wb.save(response)
        return response
    
    @staticmethod
    def reporte_cta_corriente_proveedor(id_proveedor=None, fecha_inicio=None, fecha_fin=None):
        """Genera reporte de cuenta corriente de proveedor en Excel usando el nuevo sistema"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Cta Corriente Proveedor"
        
        # Título
        if id_proveedor:
            try:
                proveedor = Proveedor.objects.get(pk=id_proveedor)
                ws['A1'] = f'Cuenta Corriente - {proveedor.razon_social}'
                ws['A2'] = f"RUC: {proveedor.ruc}"
            except Proveedor.DoesNotExist:
                ws['A1'] = 'Cuenta Corriente - Proveedor'
                ws['A2'] = 'Proveedor no encontrado'
        else:
            ws['A1'] = 'Cuenta Corriente - Todos los Proveedores'
        
        ws['A1'].font = Font(bold=True, size=14)
        
        if fecha_inicio and fecha_fin:
            ws['A3'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        else:
            ws['A3'] = f"Fecha: {date.today().strftime('%d/%m/%Y')}"
        
        # Headers
        headers = ['Fecha', 'Proveedor', 'RUC', 'Compra #', 'Total Compra', 'Saldo Pendiente', 'Estado Pago']
        ws.append([])
        ws.append(headers)
        ReportesExcel._aplicar_estilo_header(ws, row=ws.max_row)
        
        # Query compras pendientes
        compras = Compras.objects.filter(
            estado_pago__in=['PENDIENTE', 'PARCIAL']
        ).select_related('id_proveedor')
        
        if id_proveedor:
            compras = compras.filter(id_proveedor_id=id_proveedor)
        
        if fecha_inicio and fecha_fin:
            compras = compras.filter(fecha__date__gte=fecha_inicio, fecha__date__lte=fecha_fin)
        
        compras = compras.order_by('id_proveedor', 'fecha')[:200]
        
        # Datos
        total_pendiente = Decimal('0')
        for compra in compras:
            ws.append([
                compra.fecha.strftime('%d/%m/%Y %H:%M'),
                compra.id_proveedor.razon_social,
                compra.id_proveedor.ruc,
                compra.id_compra,
                float(compra.monto_total),
                float(compra.saldo_pendiente),
                compra.estado_pago
            ])
            total_pendiente += compra.saldo_pendiente
        
        # Fila de total
        ws.append(['', '', '', '', 'TOTAL PENDIENTE:', float(total_pendiente), ''])
        
        # Formato números
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=5, max_col=6):
            for cell in row:
                cell.number_format = '#,##0'
        
        ReportesExcel._ajustar_columnas(ws)
        
        # Respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="cta_corriente_proveedor_{date.today()}.xlsx"'
        wb.save(response)
        return response
