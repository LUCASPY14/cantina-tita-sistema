from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, F
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.http import JsonResponse, HttpResponse
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

from .models import Producto, Cliente, Proveedor, Categoria, StockUnico
from .forms_productos import ProductoForm, CategoriaForm
from .reportes import ReportesPDF, ReportesExcel
from .cache_reportes import get_reporte_cacheado, ReporteCache, get_datos_dashboard_cacheados
# from .models import Venta, CompraProveedor  # Estos modelos están deshabilitados por ahora


def index(request):
    """Vista principal del sistema"""
    context = {
        'titulo': 'Sistema de Gestión de Cantina'
    }
    return render(request, 'gestion/index.html', context)


@login_required
def dashboard(request):
    """Dashboard con estadísticas del sistema (CACHEADO)"""
    
    # Obtener estadísticas cacheadas (1 minuto)
    try:
        datos = get_datos_dashboard_cacheados()
        context = {
            'total_productos': datos.get('total_productos', 0),
            'productos_bajo_stock': 0,  # Temporalmente deshabilitado
            'total_clientes': datos.get('total_clientes', 0),
            'total_ventas_hoy': datos.get('total_ventas_hoy', 0),
            'cantidad_ventas_hoy': datos.get('ventas_hoy', 0),
            'total_consumos_hoy': datos.get('total_consumos_hoy', 0),
            'cantidad_consumos_hoy': datos.get('consumos_hoy', 0),
            'cache_activo': True,
            'ultima_actualizacion': datos.get('ultima_actualizacion')
        }
    except Exception as e:
        # Fallback sin cache
        total_productos = Producto.objects.filter(activo=True).count()
        total_clientes = Cliente.objects.filter(activo=True).count()
        
        context = {
            'total_productos': total_productos,
            'productos_bajo_stock': 0,
            'total_clientes': total_clientes,
            'total_ventas_hoy': 0,
            'cantidad_ventas_hoy': 0,
            'cache_activo': False
        }
    
    return render(request, 'gestion/dashboard.html', context)


# =============================================================================
# VISTAS PARA REPORTES
# =============================================================================

@login_required
def reporte_ventas_pdf(request):
    """Genera reporte de ventas en PDF con filtros avanzados (CACHEADO)"""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    id_cliente = request.GET.get('id_cliente')
    id_cajero = request.GET.get('id_cajero')
    estado = request.GET.get('estado')
    id_tipo_pago = request.GET.get('id_tipo_pago')
    monto_minimo = request.GET.get('monto_minimo')
    monto_maximo = request.GET.get('monto_maximo')
    
    if fecha_inicio:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    if fecha_fin:
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    # Usar cache (5 minutos)
    return get_reporte_cacheado(
        request,
        'ventas',
        lambda: ReportesPDF.reporte_ventas(
            fecha_inicio, fecha_fin, id_cliente, id_cajero,
            estado, id_tipo_pago, monto_minimo, monto_maximo
        ),
        timeout=300
    )


@login_required
def reporte_ventas_excel(request):
    """Genera reporte de ventas en Excel con filtros avanzados"""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    id_cliente = request.GET.get('id_cliente')
    id_cajero = request.GET.get('id_cajero')
    estado = request.GET.get('estado')
    id_tipo_pago = request.GET.get('id_tipo_pago')
    monto_minimo = request.GET.get('monto_minimo')
    monto_maximo = request.GET.get('monto_maximo')
    
    if fecha_inicio:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    if fecha_fin:
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    return ReportesExcel.reporte_ventas(
        fecha_inicio, fecha_fin, id_cliente, id_cajero,
        estado, id_tipo_pago, monto_minimo, monto_maximo
    )


@login_required
def reporte_productos_pdf(request):
    """Genera reporte de productos en PDF con filtro de categoría (CACHEADO)"""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    id_categoria = request.GET.get('id_categoria')
    
    if fecha_inicio:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    if fecha_fin:
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    # Usar cache (10 minutos)
    return get_reporte_cacheado(
        request,
        'productos',
        lambda: ReportesPDF.reporte_productos_vendidos(fecha_inicio, fecha_fin, id_categoria),
        timeout=600
    )


@login_required
def reporte_productos_excel(request):
    """Genera reporte de productos en Excel con filtro de categoría"""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    id_categoria = request.GET.get('id_categoria')
    
    if fecha_inicio:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    if fecha_fin:
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    return ReportesExcel.reporte_productos_vendidos(fecha_inicio, fecha_fin, id_categoria)


@login_required
def reporte_inventario_pdf(request):
    """Genera reporte de inventario en PDF (CACHEADO)"""
    # Usar cache (30 minutos - inventario cambia lentamente)
    return get_reporte_cacheado(
        request,
        'inventario',
        lambda: ReportesPDF.reporte_inventario(),
        timeout=1800
    )


@login_required
def reporte_inventario_excel(request):
    """Genera reporte de inventario en Excel"""
    return ReportesExcel.reporte_inventario()


@login_required
def reporte_consumos_pdf(request):
    """Genera reporte de consumos en PDF"""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if fecha_inicio:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    if fecha_fin:
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    return ReportesPDF.reporte_consumos_tarjeta(fecha_inicio, fecha_fin)


@login_required
def reporte_consumos_excel(request):
    """Genera reporte de consumos en Excel"""
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if fecha_inicio:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    if fecha_fin:
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    return ReportesExcel.reporte_consumos_tarjeta(fecha_inicio, fecha_fin)


@login_required
def reporte_clientes_pdf(request):
    """Genera reporte de clientes en PDF"""
    return ReportesPDF.reporte_clientes()


@login_required
def reporte_clientes_excel(request):
    """Genera reporte de clientes en Excel"""
    return ReportesExcel.reporte_clientes()


@login_required
@login_required
def reporte_cta_corriente_cliente_pdf(request):
    """Genera reporte de cuenta corriente cliente en PDF"""
    id_cliente = request.GET.get('id_cliente', None)
    fecha_inicio = request.GET.get('fecha_inicio', None)
    fecha_fin = request.GET.get('fecha_fin', None)
    
    if fecha_inicio:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    if fecha_fin:
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    return ReportesPDF.reporte_cta_corriente_cliente(id_cliente, fecha_inicio, fecha_fin)


@login_required
def reporte_cta_corriente_cliente_excel(request):
    """Genera reporte de cuenta corriente cliente en Excel"""
    id_cliente = request.GET.get('id_cliente', None)
    fecha_inicio = request.GET.get('fecha_inicio', None)
    fecha_fin = request.GET.get('fecha_fin', None)
    
    if fecha_inicio:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    if fecha_fin:
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    return ReportesExcel.reporte_cta_corriente_cliente(id_cliente, fecha_inicio, fecha_fin)


@login_required
def reporte_cta_corriente_proveedor_pdf(request):
    """Genera reporte de cuenta corriente proveedor en PDF"""
    id_proveedor = request.GET.get('id_proveedor', None)
    fecha_inicio = request.GET.get('fecha_inicio', None)
    fecha_fin = request.GET.get('fecha_fin', None)
    
    if fecha_inicio:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    if fecha_fin:
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    return ReportesPDF.reporte_cta_corriente_proveedor(id_proveedor, fecha_inicio, fecha_fin)


@login_required
def reporte_cta_corriente_proveedor_excel(request):
    """Genera reporte de cuenta corriente proveedor en Excel"""
    id_proveedor = request.GET.get('id_proveedor', None)
    fecha_inicio = request.GET.get('fecha_inicio', None)
    fecha_fin = request.GET.get('fecha_fin', None)
    
    if fecha_inicio:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    if fecha_fin:
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    return ReportesExcel.reporte_cta_corriente_proveedor(id_proveedor, fecha_inicio, fecha_fin)


# ==================== GESTIÓN DE PRODUCTOS ====================

@login_required
def crear_producto(request):
    """
    Vista para crear un nuevo producto
    Crea automáticamente el registro de stock inicial
    """
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Guardar producto
                    producto = form.save()
                    
                    # Crear registro de stock inicial en 0
                    StockUnico.objects.create(
                        id_producto=producto,
                        stock_actual=Decimal('0.000')
                    )
                    
                    messages.success(
                        request,
                        f'Producto "{producto.descripcion}" creado exitosamente. Stock inicial: 0'
                    )
                    
                    # Redirigir al listado de productos
                    return redirect('inventario_productos')
                    
            except Exception as e:
                messages.error(
                    request,
                    f'Error al crear el producto: {str(e)}'
                )
        else:
            # Mostrar errores de validación
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ProductoForm()
    
    context = {
        'form': form,
        'titulo': 'Crear Producto',
        'accion': 'Crear',
        'url_cancelar': 'inventario_productos'
    }
    
    return render(request, 'gestion/producto_form.html', context)


@login_required
def editar_producto(request, producto_id):
    """
    Vista para editar un producto existente
    """
    producto = get_object_or_404(Producto, pk=producto_id)
    
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        
        if form.is_valid():
            try:
                # Guardar cambios (incluye alérgenos por el método save() del form)
                producto = form.save()
                
                messages.success(
                    request,
                    f'Producto "{producto.descripcion}" actualizado exitosamente'
                )
                
                return redirect('inventario_productos')
                
            except Exception as e:
                messages.error(
                    request,
                    f'Error al actualizar el producto: {str(e)}'
                )
        else:
            # Mostrar errores de validación
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ProductoForm(instance=producto)
    
    context = {
        'form': form,
        'producto': producto,
        'titulo': 'Editar Producto',
        'accion': 'Actualizar',
        'url_cancelar': 'inventario_productos'
    }
    
    return render(request, 'gestion/producto_form.html', context)


@login_required
def eliminar_producto(request, producto_id):
    """
    Vista para desactivar un producto (soft delete)
    No elimina físicamente, solo marca como inactivo
    """
    producto = get_object_or_404(Producto, pk=producto_id)
    
    if request.method == 'POST':
        try:
            producto.activo = False
            producto.save()
            
            messages.success(
                request,
                f'Producto "{producto.descripcion}" desactivado correctamente'
            )
        except Exception as e:
            messages.error(
                request,
                f'Error al desactivar el producto: {str(e)}'
            )
    
    return redirect('inventario_productos')


# ==================== GESTIÓN DE CATEGORÍAS ====================

@login_required
def categorias_lista(request):
    """
    Lista todas las categorías con estructura jerárquica
    """
    categorias = Categoria.objects.filter(activo=True).order_by('nombre')
    
    # Construir árbol jerárquico
    categorias_principales = categorias.filter(id_categoria_padre__isnull=True)
    
    context = {
        'categorias': categorias,
        'categorias_principales': categorias_principales
    }
    
    return render(request, 'gestion/categorias_lista.html', context)


@login_required
def crear_categoria(request):
    """Vista para crear una nueva categoría"""
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        
        if form.is_valid():
            try:
                categoria = form.save()
                
                messages.success(
                    request,
                    f'Categoría "{categoria.nombre}" creada exitosamente'
                )
                
                return redirect('categorias_lista')
                
            except Exception as e:
                messages.error(
                    request,
                    f'Error al crear la categoría: {str(e)}'
                )
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CategoriaForm()
    
    context = {
        'form': form,
        'titulo': 'Crear Categoría',
        'accion': 'Crear',
        'url_cancelar': 'categorias_lista'
    }
    
    return render(request, 'gestion/categoria_form.html', context)


@login_required
def editar_categoria(request, categoria_id):
    """Vista para editar una categoría existente"""
    categoria = get_object_or_404(Categoria, pk=categoria_id)
    
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        
        if form.is_valid():
            try:
                categoria = form.save()
                
                messages.success(
                    request,
                    f'Categoría "{categoria.nombre}" actualizada exitosamente'
                )
                
                return redirect('categorias_lista')
                
            except Exception as e:
                messages.error(
                    request,
                    f'Error al actualizar la categoría: {str(e)}'
                )
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CategoriaForm(instance=categoria)
    
    context = {
        'form': form,
        'categoria': categoria,
        'titulo': 'Editar Categoría',
        'accion': 'Actualizar',
        'url_cancelar': 'categorias_lista'
    }
    
    return render(request, 'gestion/categoria_form.html', context)


@login_required
def eliminar_categoria(request, categoria_id):
    """
    Vista para eliminar una categoría
    Solo permite eliminar si no tiene productos asociados
    """
    categoria = get_object_or_404(Categoria, pk=categoria_id)
    
    if request.method == 'POST':
        # Verificar si tiene productos
        tiene_productos = categoria.productos.exists()
        
        if tiene_productos:
            messages.error(
                request,
                f'No se puede eliminar la categoría "{categoria.nombre}" porque tiene productos asociados. '
                'Desactívala o reasigna los productos primero.'
            )
        else:
            try:
                nombre = categoria.nombre
                categoria.delete()
                
                messages.success(
                    request,
                    f'Categoría "{nombre}" eliminada correctamente'
                )
            except Exception as e:
                messages.error(
                    request,
                    f'Error al eliminar la categoría: {str(e)}'
                )
    
    return redirect('categorias_lista')


# ==================== IMPORTACIÓN/EXPORTACIÓN ====================

@login_required
def importar_productos(request):
    """
    Vista para importar productos masivamente desde CSV/Excel
    """
    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']
        nombre_archivo = archivo.name.lower()
        
        try:
            # Leer datos según formato
            if nombre_archivo.endswith('.csv'):
                datos = _procesar_csv(archivo)
            elif nombre_archivo.endswith(('.xlsx', '.xls')):
                datos = _procesar_excel(archivo)
            else:
                messages.error(request, 'Formato de archivo no soportado. Use CSV o Excel.')
                return redirect('importar_productos')
            
            # Mostrar preview
            if 'confirmar' not in request.POST:
                context = {
                    'datos_preview': datos[:20],  # Primeras 20 filas
                    'total_filas': len(datos),
                    'archivo_nombre': archivo.name
                }
                return render(request, 'inventory/products/import_preview.html', context)
            
            # Importar datos
            resultados = _importar_productos_batch(datos, request.user)
            
            messages.success(
                request,
                f'Importación completada: {resultados["exitosos"]} productos creados, '
                f'{resultados["errores"]} errores'
            )
            
            if resultados['mensajes_error']:
                for error in resultados['mensajes_error'][:10]:  # Primeros 10 errores
                    messages.warning(request, error)
            
            return redirect('inventario_productos')
            
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {str(e)}')
    
    context = {
        'titulo': 'Importar Productos',
        'formato_ejemplo': 'codigo_barra,descripcion,categoria,unidad_medida,impuesto,stock_minimo,activo'
    }
    
    return render(request, 'inventory/products/import.html', context)


@login_required
def exportar_productos_csv(request):
    """Exporta productos a CSV aplicando filtros actuales"""
    # Obtener filtros de la URL
    busqueda = request.GET.get('busqueda', '')
    categoria = request.GET.get('categoria', '')
    
    # Filtrar productos
    productos = Producto.objects.select_related(
        'id_categoria', 'id_unidad_medida', 'id_impuesto', 'stock'
    ).filter(activo=True)
    
    if busqueda:
        productos = productos.filter(
            Q(codigo_barra__icontains=busqueda) | Q(descripcion__icontains=busqueda)
        )
    
    if categoria:
        productos = productos.filter(id_categoria_id=categoria)
    
    # Crear CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="productos_export.csv"'
    response.write('\ufeff')  # BOM para UTF-8
    
    writer = csv.writer(response)
    writer.writerow([
        'Código Barras', 'Descripción', 'Categoría', 'Unidad Medida',
        'Impuesto', 'Stock Actual', 'Stock Mínimo', 'Permite Stock Negativo', 'Activo'
    ])
    
    for p in productos:
        writer.writerow([
            p.codigo_barra or '',
            p.descripcion,
            p.id_categoria.nombre,
            p.id_unidad_medida.nombre,
            p.id_impuesto.descripcion,
            p.stock.stock_actual if hasattr(p, 'stock') else 0,
            p.stock_minimo or 0,
            'Sí' if p.permite_stock_negativo else 'No',
            'Sí' if p.activo else 'No'
        ])
    
    return response


@login_required
def exportar_productos_excel(request):
    """Exporta productos a Excel aplicando filtros actuales"""
    from django.db.models import Q
    
    # Obtener filtros
    busqueda = request.GET.get('busqueda', '')
    categoria = request.GET.get('categoria', '')
    
    # Filtrar productos
    productos = Producto.objects.select_related(
        'id_categoria', 'id_unidad_medida', 'id_impuesto', 'stock'
    ).filter(activo=True)
    
    if busqueda:
        productos = productos.filter(
            Q(codigo_barra__icontains=busqueda) | Q(descripcion__icontains=busqueda)
        )
    
    if categoria:
        productos = productos.filter(id_categoria_id=categoria)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Productos'
    
    # Estilos
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Encabezados
    headers = [
        'Código Barras', 'Descripción', 'Categoría', 'Unidad Medida',
        'Impuesto', 'Stock Actual', 'Stock Mínimo', 'Permite Stock Negativo', 'Activo'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    for row, p in enumerate(productos, 2):
        ws.cell(row=row, column=1, value=p.codigo_barra or '')
        ws.cell(row=row, column=2, value=p.descripcion)
        ws.cell(row=row, column=3, value=p.id_categoria.nombre)
        ws.cell(row=row, column=4, value=p.id_unidad_medida.nombre)
        ws.cell(row=row, column=5, value=p.id_impuesto.descripcion)
        ws.cell(row=row, column=6, value=float(p.stock.stock_actual) if hasattr(p, 'stock') else 0)
        ws.cell(row=row, column=7, value=float(p.stock_minimo) if p.stock_minimo else 0)
        ws.cell(row=row, column=8, value='Sí' if p.permite_stock_negativo else 'No')
        ws.cell(row=row, column=9, value='Sí' if p.activo else 'No')
    
    # Ajustar anchos de columna
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Generar respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="productos_export.xlsx"'
    
    return response


# ==================== FUNCIONES AUXILIARES ====================

def _procesar_csv(archivo):
    """Procesa archivo CSV y retorna lista de diccionarios"""
    import csv
    from io import StringIO
    
    contenido = archivo.read().decode('utf-8-sig')
    lector = csv.DictReader(StringIO(contenido))
    
    return list(lector)


def _procesar_excel(archivo):
    """Procesa archivo Excel y retorna lista de diccionarios"""
    wb = openpyxl.load_workbook(archivo)
    ws = wb.active
    
    # Primera fila como encabezados
    headers = [cell.value for cell in ws[1]]
    
    datos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        datos.append(dict(zip(headers, row)))
    
    return datos


def _importar_productos_batch(datos, usuario):
    """
    Importa productos en batch
    Retorna diccionario con resultados
    """
    from django.db.models import Q
    
    exitosos = 0
    errores = 0
    mensajes_error = []
    
    for idx, fila in enumerate(datos, 1):
        try:
            with transaction.atomic():
                # Buscar categoría
                categoria = Categoria.objects.filter(
                    Q(nombre__iexact=fila.get('categoria', '').strip())
                ).first()
                
                if not categoria:
                    raise ValueError(f'Categoría "{fila.get("categoria")}" no encontrada')
                
                # Buscar unidad de medida
                unidad = UnidadMedida.objects.filter(
                    Q(nombre__iexact=fila.get('unidad_medida', '').strip()) |
                    Q(abreviatura__iexact=fila.get('unidad_medida', '').strip())
                ).first()
                
                if not unidad:
                    unidad = UnidadMedida.objects.first()  # Default
                
                # Buscar impuesto
                impuesto = Impuesto.objects.filter(
                    Q(descripcion__icontains=fila.get('impuesto', '').strip())
                ).first()
                
                if not impuesto:
                    impuesto = Impuesto.objects.first()  # Default
                
                # Crear producto
                producto = Producto.objects.create(
                    codigo_barra=fila.get('codigo_barra', '').strip() or None,
                    descripcion=fila.get('descripcion', '').strip(),
                    id_categoria=categoria,
                    id_unidad_medida=unidad,
                    id_impuesto=impuesto,
                    stock_minimo=Decimal(str(fila.get('stock_minimo', 0) or 0)),
                    permite_stock_negativo=fila.get('activo', '').lower() in ['si', 'sí', 'yes', '1', 'true'],
                    activo=fila.get('activo', '').lower() in ['si', 'sí', 'yes', '1', 'true']
                )
                
                # Crear stock inicial
                StockUnico.objects.create(
                    id_producto=producto,
                    stock_actual=Decimal('0.000')
                )
                
                exitosos += 1
                
        except Exception as e:
            errores += 1
            mensajes_error.append(f'Fila {idx}: {str(e)}')
    
    return {
        'exitosos': exitosos,
        'errores': errores,
        'mensajes_error': mensajes_error
    }






from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator
from django.views.decorators.http import require_http_methods
import json
import csv
from io import StringIO
from openpyxl import Workbook
from .models import *

# ========================= GESTIÓN DE EMPLEADOS =========================

@staff_member_required
def gestionar_empleados(request):
    """Lista y gestiona empleados"""
    empleados = User.objects.filter(is_staff=True).order_by('username')
    return render(request, 'apps/gestion/empleados/gestionar_empleados.html', {
        'empleados': empleados
    })

@staff_member_required
def crear_empleado(request):
    """Crear nuevo empleado"""
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'El nombre de usuario ya existe')
        else:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                is_staff=True
            )
            messages.success(request, f'Empleado {username} creado exitosamente')
            return redirect('gestion:gestionar_empleados')
    
    return render(request, 'apps/gestion/empleados/crear_empleado.html')

@staff_member_required
def perfil_empleado(request, pk):
    """Ver perfil de empleado"""
    empleado = get_object_or_404(User, pk=pk, is_staff=True)
    return render(request, 'apps/gestion/empleados/perfil_empleado.html', {
        'empleado': empleado
    })

@staff_member_required
def cambiar_contrasena_empleado(request, pk):
    """Cambiar contraseña de empleado"""
    empleado = get_object_or_404(User, pk=pk, is_staff=True)
    
    if request.method == 'POST':
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        if new_password == confirm_password:
            empleado.set_password(new_password)
            empleado.save()
            messages.success(request, 'Contraseña actualizada exitosamente')
            return redirect('gestion:perfil_empleado', pk=empleado.pk)
        else:
            messages.error(request, 'Las contraseñas no coinciden')
    
    return render(request, 'apps/gestion/empleados/cambiar_contrasena.html', {
        'empleado': empleado
    })

# ========================= GESTIÓN DE PRODUCTOS =========================

@staff_member_required
def productos_lista(request):
    """Lista de productos"""
    productos = Producto.objects.all().order_by('nombre')
    
    # Búsqueda
    q = request.GET.get('q')
    if q:
        productos = productos.filter(
            Q(nombre__icontains=q) | 
            Q(codigo__icontains=q) |
            Q(categoria__nombre__icontains=q)
        )
    
    # Paginación
    paginator = Paginator(productos, 20)
    page = request.GET.get('page')
    productos = paginator.get_page(page)
    
    return render(request, 'apps/gestion/productos/productos_lista.html', {
        'productos': productos,
        'q': q or ''
    })

@staff_member_required
def crear_producto(request):
    """Crear nuevo producto"""
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        codigo = request.POST.get('codigo')
        precio = request.POST.get('precio')
        categoria_id = request.POST.get('categoria')
        stock_inicial = request.POST.get('stock_inicial', 0)
        
        if Producto.objects.filter(codigo=codigo).exists():
            messages.error(request, 'Ya existe un producto con ese código')
        else:
            producto = Producto.objects.create(
                nombre=nombre,
                codigo=codigo,
                precio=precio,
                categoria_id=categoria_id,
                stock=stock_inicial
            )
            messages.success(request, f'Producto {nombre} creado exitosamente')
            return redirect('gestion:productos_lista')
    
    categorias = CategoriaProducto.objects.all()
    return render(request, 'apps/gestion/productos/crear_producto.html', {
        'categorias': categorias
    })

@staff_member_required
def editar_producto(request, pk):
    """Editar producto existente"""
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        producto.nombre = request.POST.get('nombre')
        producto.codigo = request.POST.get('codigo')
        producto.precio = request.POST.get('precio')
        producto.categoria_id = request.POST.get('categoria')
        producto.stock = request.POST.get('stock')
        producto.save()
        
        messages.success(request, 'Producto actualizado exitosamente')
        return redirect('gestion:productos_lista')
    
    categorias = CategoriaProducto.objects.all()
    return render(request, 'apps/gestion/productos/editar_producto.html', {
        'producto': producto,
        'categorias': categorias
    })

@staff_member_required
def importar_productos(request):
    """Importar productos desde CSV"""
    if request.method == 'POST' and request.FILES.get('archivo'):
        try:
            archivo = request.FILES['archivo']
            decoded_file = archivo.read().decode('utf-8')
            io_string = StringIO(decoded_file)
            reader = csv.DictReader(io_string)
            
            productos_creados = 0
            for row in reader:
                if not Producto.objects.filter(codigo=row['codigo']).exists():
                    categoria, _ = CategoriaProducto.objects.get_or_create(
                        nombre=row.get('categoria', 'Sin categoría')
                    )
                    Producto.objects.create(
                        nombre=row['nombre'],
                        codigo=row['codigo'],
                        precio=float(row['precio']),
                        categoria=categoria,
                        stock=int(row.get('stock', 0))
                    )
                    productos_creados += 1
            
            messages.success(request, f'{productos_creados} productos importados exitosamente')
        except Exception as e:
            messages.error(request, f'Error al importar: {str(e)}')
        
        return redirect('gestion:productos_lista')
    
    return render(request, 'apps/gestion/productos/importar_productos.html')

@staff_member_required
def exportar_productos_excel(request):
    """Exportar productos a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Headers
    headers = ['Código', 'Nombre', 'Categoría', 'Precio', 'Stock']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    productos = Producto.objects.select_related('categoria').all()
    for row, producto in enumerate(productos, 2):
        ws.cell(row=row, column=1, value=producto.codigo)
        ws.cell(row=row, column=2, value=producto.nombre)
        ws.cell(row=row, column=3, value=producto.categoria.nombre if producto.categoria else '')
        ws.cell(row=row, column=4, value=float(producto.precio))
        ws.cell(row=row, column=5, value=producto.stock)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    wb.save(response)
    return response

@staff_member_required
def exportar_productos_csv(request):
    """Exportar productos a CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=productos.csv'
    
    writer = csv.writer(response)
    writer.writerow(['Código', 'Nombre', 'Categoría', 'Precio', 'Stock'])
    
    productos = Producto.objects.select_related('categoria').all()
    for producto in productos:
        writer.writerow([
            producto.codigo,
            producto.nombre,
            producto.categoria.nombre if producto.categoria else '',
            producto.precio,
            producto.stock
        ])
    
    return response

# ========================= GESTIÓN DE CATEGORÍAS =========================

@staff_member_required
def categorias_lista(request):
    """Lista de categorías"""
    categorias = CategoriaProducto.objects.all().order_by('nombre')
    return render(request, 'apps/gestion/productos/categorias_lista.html', {
        'categorias': categorias
    })

@staff_member_required
def crear_categoria(request):
    """Crear nueva categoría"""
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion', '')
        
        if CategoriaProducto.objects.filter(nombre=nombre).exists():
            messages.error(request, 'Ya existe una categoría con ese nombre')
        else:
            CategoriaProducto.objects.create(
                nombre=nombre,
                descripcion=descripcion
            )
            messages.success(request, f'Categoría {nombre} creada exitosamente')
            return redirect('gestion:categorias_lista')
    
    return render(request, 'apps/gestion/productos/crear_categoria.html')

@staff_member_required
def editar_categoria(request, pk):
    """Editar categoría"""
    categoria = get_object_or_404(CategoriaProducto, pk=pk)
    
    if request.method == 'POST':
        categoria.nombre = request.POST.get('nombre')
        categoria.descripcion = request.POST.get('descripcion', '')
        categoria.save()
        
        messages.success(request, 'Categoría actualizada exitosamente')
        return redirect('gestion:categorias_lista')
    
    return render(request, 'apps/gestion/productos/editar_categoria.html', {
        'categoria': categoria
    })

@staff_member_required
@require_http_methods(["POST"])
def eliminar_categoria(request, pk):
    """Eliminar categoría"""
    categoria = get_object_or_404(CategoriaProducto, pk=pk)
    
    if categoria.producto_set.exists():
        messages.error(request, 'No se puede eliminar una categoría que tiene productos asociados')
    else:
        categoria.delete()
        messages.success(request, 'Categoría eliminada exitosamente')
    
    return redirect('gestion:categorias_lista')

# ========================= GESTIÓN DE CLIENTES =========================

@staff_member_required
def clientes_lista(request):
    """Lista de clientes"""
    clientes = Cliente.objects.all().order_by('nombre')
    
    # Búsqueda
    q = request.GET.get('q')
    if q:
        clientes = clientes.filter(
            Q(nombre__icontains=q) | 
            Q(ci__icontains=q) |
            Q(telefono__icontains=q)
        )
    
    # Paginación
    paginator = Paginator(clientes, 20)
    page = request.GET.get('page')
    clientes = paginator.get_page(page)
    
    return render(request, 'apps/gestion/clientes/clientes_lista.html', {
        'clientes': clientes,
        'q': q or ''
    })

# ========================= GESTIÓN DE VENTAS =========================

@staff_member_required
def ventas_lista(request):
    """Lista de ventas"""
    ventas = Ventas.objects.all().order_by('-fecha')
    
    # Filtros
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    cliente_id = request.GET.get('cliente')
    
    if fecha_desde:
        ventas = ventas.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        ventas = ventas.filter(fecha__lte=fecha_hasta)
    if cliente_id:
        ventas = ventas.filter(cliente_id=cliente_id)
    
    # Paginación
    paginator = Paginator(ventas, 20)
    page = request.GET.get('page')
    ventas = paginator.get_page(page)
    
    clientes = Cliente.objects.all()
    
    return render(request, 'apps/gestion/ventas/ventas_lista.html', {
        'ventas': ventas,
        'clientes': clientes,
        'fecha_desde': fecha_desde or '',
        'fecha_hasta': fecha_hasta or '',
        'cliente_id': int(cliente_id) if cliente_id else None
    })

# ========================= FACTURACIÓN ELECTRÓNICA =========================

@staff_member_required
def facturacion_listado(request):
    """Listado de facturas electrónicas"""
    # Placeholder - requiere integración con sistema de facturación
    return render(request, 'apps/gestion/facturacion/listado.html', {
        'facturas': []
    })

@staff_member_required
def facturacion_kude(request):
    """Generar KUDE"""
    # Placeholder - requiere integración con SET
    return render(request, 'apps/gestion/facturacion/kude.html')

@staff_member_required
def facturacion_anular_api(request):
    """API para anular factura"""
    if request.method == 'POST':
        # Placeholder - requiere integración con SET
        return JsonResponse({'success': True, 'message': 'Funcionalidad en desarrollo'})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@staff_member_required
def facturacion_reporte_cumplimiento(request):
    """Reporte de cumplimiento fiscal"""
    # Placeholder - requiere cálculos fiscales
    return render(request, 'apps/gestion/facturacion/reporte_cumplimiento.html')

# ========================= REPORTES =========================

@staff_member_required
def reporte_mensual_completo(request):
    """Reporte mensual completo"""
    # Placeholder - requiere agregaciones complejas
    return render(request, 'apps/gestion/reportes/mensual_completo.html')

# ========================= VALIDACIONES =========================

@staff_member_required
@require_http_methods(["POST"])
def validar_pago_action(request):
    """Validar pago"""
    pago_id = request.POST.get('pago_id')
    # Placeholder - lógica de validación
    return JsonResponse({'success': True, 'message': 'Pago validado'})

# ========================= VIEWS DE DASHBOARD PRINCIPALES =========================

@staff_member_required
def dashboard(request):
    """Dashboard principal de gestión"""
    context = {
        'total_productos': Producto.objects.count(),
        'total_clientes': Cliente.objects.count(),
        'total_empleados': User.objects.filter(is_staff=True).count(),
        'ventas_hoy': Ventas.objects.filter(fecha__date=timezone.now().date()).count(),
    }
    return render(request, 'apps/gestion/dashboard.html', context)

@staff_member_required
def index(request):
    """Página principal de gestión"""
    return redirect('gestion:dashboard')
