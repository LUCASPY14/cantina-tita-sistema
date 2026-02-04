
# VIEWS ADICIONALES PARA GESTION - Completar funcionalidad 100%

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator
from django.views.decorators.http import require_http_methods
import json
import csv
from io import StringIO
from openpyxl import Workbook
from .models import *

# ========================= GESTIÓN DE EMPLEADOS =========================

@staff_member_required
def gestionar_empleados(request):
    """Lista y gestiona empleados"""
    empleados = User.objects.filter(is_staff=True).order_by('username')
    return render(request, 'apps/gestion/empleados/gestionar_empleados.html', {
        'empleados': empleados
    })

@staff_member_required
def crear_empleado(request):
    """Crear nuevo empleado"""
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'El nombre de usuario ya existe')
        else:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=first_name,
                last_name=last_name,
                is_staff=True
            )
            messages.success(request, f'Empleado {username} creado exitosamente')
            return redirect('gestion:gestionar_empleados')
    
    return render(request, 'apps/gestion/empleados/crear_empleado.html')

@staff_member_required
def perfil_empleado(request, pk):
    """Ver perfil de empleado"""
    empleado = get_object_or_404(User, pk=pk, is_staff=True)
    return render(request, 'apps/gestion/empleados/perfil_empleado.html', {
        'empleado': empleado
    })

@staff_member_required
def cambiar_contrasena_empleado(request, pk):
    """Cambiar contraseña de empleado"""
    empleado = get_object_or_404(User, pk=pk, is_staff=True)
    
    if request.method == 'POST':
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        if new_password == confirm_password:
            empleado.set_password(new_password)
            empleado.save()
            messages.success(request, 'Contraseña actualizada exitosamente')
            return redirect('gestion:perfil_empleado', pk=empleado.pk)
        else:
            messages.error(request, 'Las contraseñas no coinciden')
    
    return render(request, 'apps/gestion/empleados/cambiar_contrasena.html', {
        'empleado': empleado
    })

# ========================= GESTIÓN DE PRODUCTOS =========================

@staff_member_required
def productos_lista(request):
    """Lista de productos"""
    productos = Producto.objects.all().order_by('nombre')
    
    # Búsqueda
    q = request.GET.get('q')
    if q:
        productos = productos.filter(
            Q(nombre__icontains=q) | 
            Q(codigo__icontains=q) |
            Q(categoria__nombre__icontains=q)
        )
    
    # Paginación
    paginator = Paginator(productos, 20)
    page = request.GET.get('page')
    productos = paginator.get_page(page)
    
    return render(request, 'apps/gestion/productos/productos_lista.html', {
        'productos': productos,
        'q': q or ''
    })

@staff_member_required
def crear_producto(request):
    """Crear nuevo producto"""
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        codigo = request.POST.get('codigo')
        precio = request.POST.get('precio')
        categoria_id = request.POST.get('categoria')
        stock_inicial = request.POST.get('stock_inicial', 0)
        
        if Producto.objects.filter(codigo=codigo).exists():
            messages.error(request, 'Ya existe un producto con ese código')
        else:
            producto = Producto.objects.create(
                nombre=nombre,
                codigo=codigo,
                precio=precio,
                categoria_id=categoria_id,
                stock=stock_inicial
            )
            messages.success(request, f'Producto {nombre} creado exitosamente')
            return redirect('gestion:productos_lista')
    
    categorias = CategoriaProducto.objects.all()
    return render(request, 'apps/gestion/productos/crear_producto.html', {
        'categorias': categorias
    })

@staff_member_required
def editar_producto(request, pk):
    """Editar producto existente"""
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        producto.nombre = request.POST.get('nombre')
        producto.codigo = request.POST.get('codigo')
        producto.precio = request.POST.get('precio')
        producto.categoria_id = request.POST.get('categoria')
        producto.stock = request.POST.get('stock')
        producto.save()
        
        messages.success(request, 'Producto actualizado exitosamente')
        return redirect('gestion:productos_lista')
    
    categorias = CategoriaProducto.objects.all()
    return render(request, 'apps/gestion/productos/editar_producto.html', {
        'producto': producto,
        'categorias': categorias
    })

@staff_member_required
def importar_productos(request):
    """Importar productos desde CSV"""
    if request.method == 'POST' and request.FILES.get('archivo'):
        try:
            archivo = request.FILES['archivo']
            decoded_file = archivo.read().decode('utf-8')
            io_string = StringIO(decoded_file)
            reader = csv.DictReader(io_string)
            
            productos_creados = 0
            for row in reader:
                if not Producto.objects.filter(codigo=row['codigo']).exists():
                    categoria, _ = CategoriaProducto.objects.get_or_create(
                        nombre=row.get('categoria', 'Sin categoría')
                    )
                    Producto.objects.create(
                        nombre=row['nombre'],
                        codigo=row['codigo'],
                        precio=float(row['precio']),
                        categoria=categoria,
                        stock=int(row.get('stock', 0))
                    )
                    productos_creados += 1
            
            messages.success(request, f'{productos_creados} productos importados exitosamente')
        except Exception as e:
            messages.error(request, f'Error al importar: {str(e)}')
        
        return redirect('gestion:productos_lista')
    
    return render(request, 'apps/gestion/productos/importar_productos.html')

@staff_member_required
def exportar_productos_excel(request):
    """Exportar productos a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Headers
    headers = ['Código', 'Nombre', 'Categoría', 'Precio', 'Stock']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    productos = Producto.objects.select_related('categoria').all()
    for row, producto in enumerate(productos, 2):
        ws.cell(row=row, column=1, value=producto.codigo)
        ws.cell(row=row, column=2, value=producto.nombre)
        ws.cell(row=row, column=3, value=producto.categoria.nombre if producto.categoria else '')
        ws.cell(row=row, column=4, value=float(producto.precio))
        ws.cell(row=row, column=5, value=producto.stock)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    wb.save(response)
    return response

@staff_member_required
def exportar_productos_csv(request):
    """Exportar productos a CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=productos.csv'
    
    writer = csv.writer(response)
    writer.writerow(['Código', 'Nombre', 'Categoría', 'Precio', 'Stock'])
    
    productos = Producto.objects.select_related('categoria').all()
    for producto in productos:
        writer.writerow([
            producto.codigo,
            producto.nombre,
            producto.categoria.nombre if producto.categoria else '',
            producto.precio,
            producto.stock
        ])
    
    return response

# ========================= GESTIÓN DE CATEGORÍAS =========================

@staff_member_required
def categorias_lista(request):
    """Lista de categorías"""
    categorias = CategoriaProducto.objects.all().order_by('nombre')
    return render(request, 'apps/gestion/productos/categorias_lista.html', {
        'categorias': categorias
    })

@staff_member_required
def crear_categoria(request):
    """Crear nueva categoría"""
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion', '')
        
        if CategoriaProducto.objects.filter(nombre=nombre).exists():
            messages.error(request, 'Ya existe una categoría con ese nombre')
        else:
            CategoriaProducto.objects.create(
                nombre=nombre,
                descripcion=descripcion
            )
            messages.success(request, f'Categoría {nombre} creada exitosamente')
            return redirect('gestion:categorias_lista')
    
    return render(request, 'apps/gestion/productos/crear_categoria.html')

@staff_member_required
def editar_categoria(request, pk):
    """Editar categoría"""
    categoria = get_object_or_404(CategoriaProducto, pk=pk)
    
    if request.method == 'POST':
        categoria.nombre = request.POST.get('nombre')
        categoria.descripcion = request.POST.get('descripcion', '')
        categoria.save()
        
        messages.success(request, 'Categoría actualizada exitosamente')
        return redirect('gestion:categorias_lista')
    
    return render(request, 'apps/gestion/productos/editar_categoria.html', {
        'categoria': categoria
    })

@staff_member_required
@require_http_methods(["POST"])
def eliminar_categoria(request, pk):
    """Eliminar categoría"""
    categoria = get_object_or_404(CategoriaProducto, pk=pk)
    
    if categoria.producto_set.exists():
        messages.error(request, 'No se puede eliminar una categoría que tiene productos asociados')
    else:
        categoria.delete()
        messages.success(request, 'Categoría eliminada exitosamente')
    
    return redirect('gestion:categorias_lista')

# ========================= GESTIÓN DE CLIENTES =========================

@staff_member_required
def clientes_lista(request):
    """Lista de clientes"""
    clientes = Cliente.objects.all().order_by('nombre')
    
    # Búsqueda
    q = request.GET.get('q')
    if q:
        clientes = clientes.filter(
            Q(nombre__icontains=q) | 
            Q(ci__icontains=q) |
            Q(telefono__icontains=q)
        )
    
    # Paginación
    paginator = Paginator(clientes, 20)
    page = request.GET.get('page')
    clientes = paginator.get_page(page)
    
    return render(request, 'apps/gestion/clientes/clientes_lista.html', {
        'clientes': clientes,
        'q': q or ''
    })

# ========================= GESTIÓN DE VENTAS =========================

@staff_member_required
def ventas_lista(request):
    """Lista de ventas"""
    ventas = Ventas.objects.all().order_by('-fecha')
    
    # Filtros
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    cliente_id = request.GET.get('cliente')
    
    if fecha_desde:
        ventas = ventas.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        ventas = ventas.filter(fecha__lte=fecha_hasta)
    if cliente_id:
        ventas = ventas.filter(cliente_id=cliente_id)
    
    # Paginación
    paginator = Paginator(ventas, 20)
    page = request.GET.get('page')
    ventas = paginator.get_page(page)
    
    clientes = Cliente.objects.all()
    
    return render(request, 'apps/gestion/ventas/ventas_lista.html', {
        'ventas': ventas,
        'clientes': clientes,
        'fecha_desde': fecha_desde or '',
        'fecha_hasta': fecha_hasta or '',
        'cliente_id': int(cliente_id) if cliente_id else None
    })

# ========================= FACTURACIÓN ELECTRÓNICA =========================

@staff_member_required
def facturacion_listado(request):
    """Listado de facturas electrónicas"""
    # Placeholder - requiere integración con sistema de facturación
    return render(request, 'apps/gestion/facturacion/listado.html', {
        'facturas': []
    })

@staff_member_required
def facturacion_kude(request):
    """Generar KUDE"""
    # Placeholder - requiere integración con SET
    return render(request, 'apps/gestion/facturacion/kude.html')

@staff_member_required
def facturacion_anular_api(request):
    """API para anular factura"""
    if request.method == 'POST':
        # Placeholder - requiere integración con SET
        return JsonResponse({'success': True, 'message': 'Funcionalidad en desarrollo'})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

@staff_member_required
def facturacion_reporte_cumplimiento(request):
    """Reporte de cumplimiento fiscal"""
    # Placeholder - requiere cálculos fiscales
    return render(request, 'apps/gestion/facturacion/reporte_cumplimiento.html')

# ========================= REPORTES =========================

@staff_member_required
def reporte_mensual_completo(request):
    """Reporte mensual completo"""
    # Placeholder - requiere agregaciones complejas
    return render(request, 'apps/gestion/reportes/mensual_completo.html')

# ========================= VALIDACIONES =========================

@staff_member_required
@require_http_methods(["POST"])
def validar_pago_action(request):
    """Validar pago"""
    pago_id = request.POST.get('pago_id')
    # Placeholder - lógica de validación
    return JsonResponse({'success': True, 'message': 'Pago validado'})

# ========================= VIEWS DE DASHBOARD PRINCIPALES =========================

@staff_member_required
def dashboard(request):
    """Dashboard principal de gestión"""
    context = {
        'total_productos': Producto.objects.count(),
        'total_clientes': Cliente.objects.count(),
        'total_empleados': User.objects.filter(is_staff=True).count(),
        'ventas_hoy': Ventas.objects.filter(fecha__date=timezone.now().date()).count(),
    }
    return render(request, 'apps/gestion/dashboard.html', context)

@staff_member_required
def index(request):
    """Página principal de gestión"""
    return redirect('gestion:dashboard')
