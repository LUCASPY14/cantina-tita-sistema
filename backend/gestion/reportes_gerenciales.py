"""
Reportes Gerenciales Avanzados - Cantina Tita
Incluye: Reporte Mensual Completo, Conciliación Bancaria, Dashboard Ejecutivo
"""

from django.http import HttpResponse
from django.db.models import Sum, Count, Q, Avg, F, ExpressionWrapper, DecimalField
from django.db.models.functions import TruncMonth, TruncWeek, TruncDate
from datetime import datetime, date, timedelta
from decimal import Decimal
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

from pos.models import Venta as Ventas, DetalleVenta, PagoVenta as PagosVenta
from gestion.models import (
    Compras, CargasSaldo, ConsumoTarjeta,
    PagosProveedores, StockUnico, Producto,
    Cliente, Empleado, TransaccionOnline
)


class ReportesGerenciales:
    """Reportes ejecutivos para gerencia"""
    
    @staticmethod
    def reporte_mensual_completo(mes=None, año=None):
        """
        Genera reporte ejecutivo mensual completo en Excel
        Incluye: Ventas, Compras, Flujo de Caja, Productos Top, KPIs
        """
        # Configurar período
        if not mes or not año:
            hoy = date.today()
            mes = hoy.month
            año = hoy.year
        
        fecha_inicio = date(año, mes, 1)
        if mes == 12:
            fecha_fin = date(año + 1, 1, 1) - timedelta(days=1)
        else:
            fecha_fin = date(año, mes + 1, 1) - timedelta(days=1)
        
        wb = Workbook()
        
        # === HOJA 1: RESUMEN EJECUTIVO ===
        ws_resumen = wb.active
        ws_resumen.title = "Resumen Ejecutivo"
        
        ReportesGerenciales._crear_hoja_resumen_ejecutivo(
            ws_resumen, fecha_inicio, fecha_fin, mes, año
        )
        
        # === HOJA 2: VENTAS DETALLADAS ===
        ws_ventas = wb.create_sheet("Ventas")
        ReportesGerenciales._crear_hoja_ventas(ws_ventas, fecha_inicio, fecha_fin)
        
        # === HOJA 3: COMPRAS ===
        ws_compras = wb.create_sheet("Compras")
        ReportesGerenciales._crear_hoja_compras(ws_compras, fecha_inicio, fecha_fin)
        
        # === HOJA 4: FLUJO DE CAJA ===
        ws_flujo = wb.create_sheet("Flujo de Caja")
        ReportesGerenciales._crear_hoja_flujo_caja(ws_flujo, fecha_inicio, fecha_fin)
        
        # === HOJA 5: TOP PRODUCTOS ===
        ws_productos = wb.create_sheet("Top Productos")
        ReportesGerenciales._crear_hoja_top_productos(ws_productos, fecha_inicio, fecha_fin)
        
        # === HOJA 6: INDICADORES (KPIs) ===
        ws_kpis = wb.create_sheet("KPIs")
        ReportesGerenciales._crear_hoja_kpis(ws_kpis, fecha_inicio, fecha_fin)
        
        # Respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        nombre_mes = fecha_inicio.strftime('%B_%Y')
        response['Content-Disposition'] = f'attachment; filename="reporte_mensual_{nombre_mes}.xlsx"'
        wb.save(response)
        return response
    
    @staticmethod
    def _crear_hoja_resumen_ejecutivo(ws, fecha_inicio, fecha_fin, mes, año):
        """Crea la hoja de resumen ejecutivo con KPIs principales"""
        
        # Título
        ws['A1'] = f'REPORTE EJECUTIVO - {fecha_inicio.strftime("%B %Y").upper()}'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 30
        
        # Período
        ws['A2'] = f'Período: {fecha_inicio.strftime("%d/%m/%Y")} - {fecha_fin.strftime("%d/%m/%Y")}'
        ws['A2'].font = Font(size=11, italic=True)
        ws.merge_cells('A2:F2')
        
        # Calcular métricas
        ventas_mes = Ventas.objects.filter(
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin,
            estado='PROCESADO'
        )
        
        total_ventas = ventas_mes.aggregate(total=Sum('monto_total'))['total'] or 0
        num_ventas = ventas_mes.count()
        ticket_promedio = total_ventas / num_ventas if num_ventas > 0 else 0
        
        compras_mes = Compras.objects.filter(
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin
        )
        
        total_compras = compras_mes.aggregate(total=Sum('monto_total'))['total'] or 0
        
        recargas_mes = CargasSaldo.objects.filter(
            fecha_carga__date__gte=fecha_inicio,
            fecha_carga__date__lte=fecha_fin
        )
        
        total_recargas = recargas_mes.aggregate(total=Sum('monto'))['total'] or 0
        
        margen_bruto = total_ventas - total_compras
        margen_porcentaje = (margen_bruto / total_ventas * 100) if total_ventas > 0 else 0
        
        # === SECCIÓN: INGRESOS ===
        ws['A4'] = 'INGRESOS'
        ws['A4'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A4'].fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
        ws.merge_cells('A4:C4')
        
        metricas_ingresos = [
            ('Total Ventas', total_ventas),
            ('Total Recargas', total_recargas),
            ('Número de Ventas', num_ventas),
            ('Ticket Promedio', ticket_promedio),
        ]
        
        row = 5
        for nombre, valor in metricas_ingresos:
            ws[f'A{row}'] = nombre
            ws[f'B{row}'] = valor
            ws[f'B{row}'].number_format = 'Gs. #,##0'
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
        
        # === SECCIÓN: EGRESOS ===
        row += 1
        ws[f'A{row}'] = 'EGRESOS'
        ws[f'A{row}'].font = Font(size=14, bold=True, color='FFFFFF')
        ws[f'A{row}'].fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = 'Total Compras'
        ws[f'B{row}'] = total_compras
        ws[f'B{row}'].number_format = 'Gs. #,##0'
        ws[f'B{row}'].font = Font(bold=True)
        
        # === SECCIÓN: RENTABILIDAD ===
        row += 2
        ws[f'A{row}'] = 'RENTABILIDAD'
        ws[f'A{row}'].font = Font(size=14, bold=True, color='FFFFFF')
        ws[f'A{row}'].fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = 'Margen Bruto'
        ws[f'B{row}'] = margen_bruto
        ws[f'B{row}'].number_format = 'Gs. #,##0'
        ws[f'B{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].fill = PatternFill(start_color='D5F4E6', end_color='D5F4E6', fill_type='solid')
        
        row += 1
        ws[f'A{row}'] = 'Margen %'
        ws[f'B{row}'] = margen_porcentaje / 100
        ws[f'B{row}'].number_format = '0.00%'
        ws[f'B{row}'].font = Font(bold=True)
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
    
    @staticmethod
    def _crear_hoja_ventas(ws, fecha_inicio, fecha_fin):
        """Hoja detallada de ventas"""
        ws['A1'] = 'DETALLE DE VENTAS'
        ws['A1'].font = Font(size=14, bold=True)
        
        headers = ['Fecha', 'Cliente', 'Monto', 'Tipo Pago', 'Cajero', 'Estado']
        ws.append(headers)
        
        # Estilo header
        for col in range(1, len(headers) + 1):
            cell = ws.cell(2, col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Datos
        ventas = Ventas.objects.filter(
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin
        ).select_related('id_cliente', 'id_empleado_cajero', 'id_tipo_pago').order_by('-fecha')
        
        for venta in ventas[:1000]:  # Limitar a 1000 registros
            ws.append([
                venta.fecha.strftime('%d/%m/%Y %H:%M'),
                f"{venta.id_cliente.nombres} {venta.id_cliente.apellidos}",
                float(venta.monto_total),
                venta.id_tipo_pago.nombre if venta.id_tipo_pago else 'N/A',
                venta.id_empleado_cajero.nombre if venta.id_empleado_cajero else 'N/A',
                venta.estado
            ])
        
        # Formato de números
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = 'Gs. #,##0'
        
        # Totales
        ws.append([])
        ws.append(['TOTAL', '', f'=SUM(C3:C{ws.max_row-1})', '', '', ''])
        ws[f'A{ws.max_row}'].font = Font(bold=True)
        ws[f'C{ws.max_row}'].font = Font(bold=True)
        ws[f'C{ws.max_row}'].number_format = 'Gs. #,##0'
    
    @staticmethod
    def _crear_hoja_compras(ws, fecha_inicio, fecha_fin):
        """Hoja detallada de compras"""
        ws['A1'] = 'DETALLE DE COMPRAS A PROVEEDORES'
        ws['A1'].font = Font(size=14, bold=True)
        
        headers = ['Fecha', 'Proveedor', 'Nro. Factura', 'Monto', 'Estado Pago']
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(2, col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
        
        compras = Compras.objects.filter(
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin
        ).select_related('id_proveedor').order_by('-fecha')
        
        for compra in compras[:500]:
            ws.append([
                compra.fecha.strftime('%d/%m/%Y'),
                compra.id_proveedor.razon_social,
                compra.nro_factura or 'S/N',
                float(compra.monto_total),
                compra.estado_pago
            ])
        
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = 'Gs. #,##0'
    
    @staticmethod
    def _crear_hoja_flujo_caja(ws, fecha_inicio, fecha_fin):
        """Hoja de flujo de caja diario"""
        ws['A1'] = 'FLUJO DE CAJA'
        ws['A1'].font = Font(size=14, bold=True)
        
        headers = ['Fecha', 'Ingresos Ventas', 'Ingresos Recargas', 'Total Ingresos', 'Egresos Compras', 'Flujo Neto']
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(2, col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
        
        # Calcular por día
        fecha_actual = fecha_inicio
        while fecha_actual <= fecha_fin:
            ventas_dia = Ventas.objects.filter(
                fecha__date=fecha_actual,
                estado='PROCESADO'
            ).aggregate(total=Sum('monto_total'))['total'] or 0
            
            recargas_dia = CargasSaldo.objects.filter(
                fecha_carga__date=fecha_actual
            ).aggregate(total=Sum('monto'))['total'] or 0
            
            compras_dia = Compras.objects.filter(
                fecha__date=fecha_actual
            ).aggregate(total=Sum('monto_total'))['total'] or 0
            
            total_ingresos = ventas_dia + recargas_dia
            flujo_neto = total_ingresos - compras_dia
            
            ws.append([
                fecha_actual.strftime('%d/%m/%Y'),
                float(ventas_dia),
                float(recargas_dia),
                float(total_ingresos),
                float(compras_dia),
                float(flujo_neto)
            ])
            
            fecha_actual += timedelta(days=1)
        
        # Formato números
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=6):
            for cell in row:
                cell.number_format = 'Gs. #,##0'
    
    @staticmethod
    def _crear_hoja_top_productos(ws, fecha_inicio, fecha_fin):
        """Top 50 productos más vendidos"""
        ws['A1'] = 'TOP 50 PRODUCTOS MÁS VENDIDOS'
        ws['A1'].font = Font(size=14, bold=True)
        
        headers = ['Ranking', 'Producto', 'Cantidad', 'Total Vendido', '% del Total']
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(2, col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid')
        
        productos = DetalleVenta.objects.filter(
            id_venta__fecha__date__gte=fecha_inicio,
            id_venta__fecha__date__lte=fecha_fin,
            id_venta__estado='PROCESADO'
        ).values(
            'id_producto__descripcion'
        ).annotate(
            cantidad_total=Sum('cantidad'),
            monto_total=Sum('subtotal_total')
        ).order_by('-monto_total')[:50]
        
        total_general = sum(p['monto_total'] for p in productos)
        
        for idx, producto in enumerate(productos, 1):
            porcentaje = (producto['monto_total'] / total_general * 100) if total_general > 0 else 0
            ws.append([
                idx,
                producto['id_producto__descripcion'],
                float(producto['cantidad_total']),
                float(producto['monto_total']),
                porcentaje / 100
            ])
        
        # Formato
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = 'Gs. #,##0'
        
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = '0.00%'
    
    @staticmethod
    def _crear_hoja_kpis(ws, fecha_inicio, fecha_fin):
        """Indicadores clave de desempeño"""
        ws['A1'] = 'INDICADORES CLAVE (KPIs)'
        ws['A1'].font = Font(size=14, bold=True)
        
        # Calcular KPIs
        ventas = Ventas.objects.filter(
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin,
            estado='PROCESADO'
        )
        
        num_dias = (fecha_fin - fecha_inicio).days + 1
        
        kpis = [
            ['VENTAS', ''],
            ['Total Ventas', ventas.aggregate(t=Sum('monto_total'))['t'] or 0],
            ['Cantidad Transacciones', ventas.count()],
            ['Venta Promedio Diaria', (ventas.aggregate(t=Sum('monto_total'))['t'] or 0) / num_dias],
            ['Ticket Promedio', ventas.aggregate(t=Avg('monto_total'))['t'] or 0],
            ['', ''],
            ['CLIENTES', ''],
            ['Clientes Únicos Atendidos', ventas.values('id_cliente').distinct().count()],
            ['Frecuencia Compra Promedio', ventas.count() / max(ventas.values('id_cliente').distinct().count(), 1)],
            ['', ''],
            ['PRODUCTOS', ''],
            ['Productos Diferentes Vendidos', DetalleVenta.objects.filter(id_venta__in=ventas).values('id_producto').distinct().count()],
            ['Unidades Totales Vendidas', DetalleVenta.objects.filter(id_venta__in=ventas).aggregate(t=Sum('cantidad'))['t'] or 0],
        ]
        
        for item in kpis:
            ws.append(item)
        
        # Formato
        for row in range(2, ws.max_row + 1):
            if ws[f'B{row}'].value:
                ws[f'B{row}'].number_format = 'Gs. #,##0'
                ws[f'B{row}'].font = Font(bold=True)
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    @staticmethod
    def conciliacion_bancaria(fecha_inicio=None, fecha_fin=None):
        """
        Genera reporte de conciliación bancaria
        Compara transacciones registradas vs pagos recibidos
        """
        if not fecha_inicio:
            fecha_inicio = date.today() - timedelta(days=30)
        if not fecha_fin:
            fecha_fin = date.today()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Conciliación"
        
        ws['A1'] = 'CONCILIACIÓN BANCARIA'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f'Período: {fecha_inicio.strftime("%d/%m/%Y")} - {fecha_fin.strftime("%d/%m/%Y")}'
        
        # Headers
        headers = ['Fecha', 'Tipo', 'Referencia', 'Monto Sistema', 'Monto Real', 'Diferencia', 'Estado']
        ws.append([])
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(4, col)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        
        # Recargas online
        recargas = TransaccionOnline.objects.filter(
            fecha_transaccion__date__gte=fecha_inicio,
            fecha_transaccion__date__lte=fecha_fin
        ).order_by('fecha_transaccion')
        
        for recarga in recargas:
            # Simulamos conciliación (en producción comparar con extracto bancario)
            diferencia = 0  # Calcular diferencia real
            estado = 'OK' if diferencia == 0 else 'REVISAR'
            
            ws.append([
                recarga.fecha_transaccion.strftime('%d/%m/%Y'),
                'Recarga Online',
                recarga.nro_transaccion,
                float(recarga.monto),
                float(recarga.monto),  # En producción: monto del banco
                diferencia,
                estado
            ])
        
        # Formato
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=4, max_col=6):
            for cell in row:
                cell.number_format = 'Gs. #,##0'
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="conciliacion_bancaria_{fecha_inicio}_{fecha_fin}.xlsx"'
        wb.save(response)
        return response
