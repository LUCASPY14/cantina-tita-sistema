"""
Views para el sistema POS (Punto de Venta)
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q, F, Sum, Count
from django.db import transaction, models
from django.utils import timezone
from decimal import Decimal
import json

from gestion.models import (
    Producto, Tarjeta, Ventas, DetalleVenta, 
    ConsumoTarjeta, Empleado, StockUnico, Cliente, Hijo,
    CargasSaldo, Proveedor, Categoria, Cajas, CierresCaja,
    MediosPago, TiposPago, PagosVenta, ConciliacionPagos,
    Compras, DetalleCompra, MovimientosStock,
    TarifasComision, DetalleComisionVenta, ListaPrecios, TipoCliente,
    TipoRolGeneral, DocumentosTributarios, Timbrados,
    TarjetaAutorizacion, LogAutorizacion
)

from gestion.seguridad_utils import registrar_auditoria
from gestion.restricciones_utils import (
    analizar_restricciones_producto,
    analizar_carrito_completo
)
from gestion.promociones_utils import (
    calcular_promociones_disponibles,
    registrar_promocion_aplicada
)


import os
import base64
from django.conf import settings
from django.core.files.base import ContentFile


@login_required
def venta_view(request):
    """Vista principal del POS"""
    # Obtener lista de precios por defecto (ID 1 o la primera)
    try:
        from gestion.models import ListaPrecios, PreciosPorLista
        lista_precios = ListaPrecios.objects.filter(activo=True).first()
    except:
        lista_precios = None
    
    # Obtener productos activos con stock y precios
    productos = Producto.objects.filter(
        activo=True
    ).select_related('id_categoria', 'stock').order_by('descripcion')[:50]
    
    # Agregar precio de cada producto
    for producto in productos:
        if lista_precios:
            try:
                precio_obj = PreciosPorLista.objects.get(
                    id_producto=producto,
                    id_lista=lista_precios
                )
                producto.precio_actual = precio_obj.precio_unitario_neto
            except PreciosPorLista.DoesNotExist:
                producto.precio_actual = 0
        else:
            producto.precio_actual = 0
        
        # Marcar si es almuerzo por kilo
        producto.es_por_kilo = 'KILO' in producto.descripcion.upper()
    
    # Obtener tipos de pago disponibles
    tipos_pago = TiposPago.objects.filter(activo=True).order_by('descripcion')
    
    # Obtener medios de pago disponibles
    medios_pago = MediosPago.objects.filter(activo=True).order_by('descripcion')
    
    context = {
        'productos': productos,
        'lista_precios': lista_precios,
        'tipos_pago': tipos_pago,
        'medios_pago': medios_pago,
    }
    return render(request, 'pos/venta.html', context)


@login_required
@require_http_methods(["POST"])
def buscar_productos(request):
    """Búsqueda de productos en tiempo real con HTMX"""
    from gestion.models import ListaPrecios, PreciosPorLista
    
    query = request.POST.get('q', '').strip()
    
    # Obtener lista de precios
    lista_precios = ListaPrecios.objects.filter(activo=True).first()
    
    if query:
        productos = Producto.objects.filter(
            Q(descripcion__icontains=query) | Q(codigo_barra__icontains=query),
            activo=True
        ).select_related('id_categoria', 'stock').order_by('descripcion')[:30]
    else:
        productos = Producto.objects.filter(
            activo=True
        ).select_related('id_categoria', 'stock').order_by('descripcion')[:50]
    
    # Agregar precio actual desde la tabla de precios
    for producto in productos:
        if lista_precios:
            try:
                precio_obj = PreciosPorLista.objects.get(
                    id_producto=producto,
                    id_lista=lista_precios
                )
                producto.precio_actual = precio_obj.precio_unitario_neto
            except PreciosPorLista.DoesNotExist:
                producto.precio_actual = 0
        else:
            producto.precio_actual = 0
        
        # Marcar si es por kilo
        producto.es_por_kilo = 'KILO' in producto.descripcion.upper()
    
    return render(request, 'pos/partials/productos_grid.html', {
        'productos': productos
    })


@login_required
@require_http_methods(["GET"])
@csrf_exempt
def productos_por_categoria(request):
    """Filtrar productos por categoría"""
    from gestion.models import ListaPrecios, PreciosPorLista
    
    categoria = request.GET.get('categoria', 'todos')
    lista_precios = ListaPrecios.objects.filter(activo=True).first()
    
    if categoria == 'todos':
        productos = Producto.objects.filter(activo=True)
    else:
        productos = Producto.objects.filter(
            activo=True,
            id_categoria__nombre__icontains=categoria
        )
    
    productos = productos.select_related('id_categoria', 'stock').order_by('descripcion')[:50]
    
    # Agregar precio actual
    for producto in productos:
        if lista_precios:
            try:
                precio_obj = PreciosPorLista.objects.get(
                    id_producto=producto,
                    id_lista=lista_precios
                )
                producto.precio_actual = precio_obj.precio_unitario_neto
            except PreciosPorLista.DoesNotExist:
                producto.precio_actual = 0
        else:
            producto.precio_actual = 0
        
        producto.es_por_kilo = 'KILO' in producto.descripcion.upper()
    
    return render(request, 'pos/partials/productos_grid.html', {
        'productos': productos
    })


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def buscar_tarjeta(request):
    """Buscar tarjeta de estudiante"""
    nro_tarjeta = request.POST.get('nro_tarjeta', '').strip()
    
    print(f"🔍 Buscando tarjeta: '{nro_tarjeta}'")
    print(f"📋 Método: {request.method}")
    
    if not nro_tarjeta:
        print("❌ No se proporcionó número de tarjeta")
        return render(request, 'pos/partials/tarjeta_info.html', {
            'tarjeta': None
        })
    
    try:
        # Buscar tarjeta activa (estado puede ser 'Activa' o 'ACTIVA')
        # Usar only() para especificar exactamente qué campos queremos
        tarjeta = Tarjeta.objects.select_related(
            'id_hijo',
            'id_hijo__id_cliente_responsable'
        ).only(
            'nro_tarjeta', 'estado', 'saldo_actual',
            'id_hijo__id_hijo', 'id_hijo__nombre', 'id_hijo__apellido', 
            'id_hijo__grado', 'id_hijo__foto_perfil', 'id_hijo__fecha_foto',
            'id_hijo__id_cliente_responsable__id_cliente',
            'id_hijo__id_cliente_responsable__nombres',
            'id_hijo__id_cliente_responsable__apellidos'
        ).filter(nro_tarjeta=nro_tarjeta).filter(
            Q(estado='ACTIVA') | Q(estado='Activa')
        ).first()
        
        if not tarjeta:
            print(f"❌ Tarjeta {nro_tarjeta} no encontrada o inactiva")
            return render(request, 'pos/partials/tarjeta_info.html', {
                'tarjeta': None
            })
        
        print(f"✅ Tarjeta encontrada: {tarjeta.nro_tarjeta}, Estado: {tarjeta.estado}")
        
        # Agregar información del estudiante
        if tarjeta.id_hijo:
            tarjeta.estudiante_nombre = f"{tarjeta.id_hijo.nombre} {tarjeta.id_hijo.apellido}"
            print(f"👤 Estudiante: {tarjeta.estudiante_nombre}")
            
            # Agregar información de foto
            tarjeta.foto_url = None
            tarjeta.tiene_foto = False
            tarjeta.fecha_foto = None
            
            if tarjeta.id_hijo.foto_perfil:
                tarjeta.foto_url = settings.MEDIA_URL + tarjeta.id_hijo.foto_perfil
                tarjeta.tiene_foto = True
                if hasattr(tarjeta.id_hijo, 'fecha_foto'):
                    tarjeta.fecha_foto = tarjeta.id_hijo.fecha_foto
        else:
            tarjeta.estudiante_nombre = "Estudiante"
        
        tarjeta.codigo_tarjeta = nro_tarjeta
        
        return render(request, 'pos/partials/tarjeta_info.html', {
            'tarjeta': tarjeta
        })
        
    except Exception as e:
        print(f"❌ ERROR al buscar tarjeta: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return render(request, 'pos/partials/tarjeta_info.html', {
            'tarjeta': None
        })


@login_required
@csrf_exempt
@require_http_methods(["POST"])
@transaction.atomic
def procesar_venta(request):
    """Procesar una venta desde el POS"""
    try:
        data = json.loads(request.body)
        items = data.get('items', [])
        tarjeta_data = data.get('tarjeta')
        total = Decimal(str(data.get('total', 0)))
        
        # 💰 CAPTURAR PAGOS MIXTOS (nuevo sistema)
        pagos_mixtos = data.get('pagos', [])
        
        # Mantener compatibilidad con sistema anterior (tipo_pago_id único)
        tipo_pago_id = data.get('tipo_pago_id', 1)  # Default: Contado
        medio_pago_id = data.get('medio_pago_id')  # Opcional
        
        # ⚠️ CAPTURAR DATOS DE RESTRICCIONES ALIMENTARIAS
        restricciones_confirmadas = data.get('restricciones_confirmadas', False)
        justificacion_restricciones = data.get('justificacion_restricciones', '')
        
        # 🎉 CAPTURAR DATOS DE PROMOCIÓN SI HAY ALGUNA APLICADA
        promocion_id = data.get('promocion_id')
        descuento_promocion = Decimal(str(data.get('descuento_promocion', 0)))
        
        if not items:
            return JsonResponse({
                'success': False,
                'error': 'El carrito está vacío'
            })
        
        # 💰 VALIDAR PAGOS MIXTOS si existen
        if pagos_mixtos:
            # Calcular comisión total en base al total de productos
            total_comision = Decimal('0')
            for pago_data in pagos_mixtos:
                medio_id = pago_data.get('medio_id')
                if medio_id == 3:  # Tarjeta de Débito
                    total_comision += total * Decimal('0.03')
                elif medio_id == 2:  # Tarjeta de Crédito
                    total_comision += total * Decimal('0.05')
                elif medio_id == 4:  # Giros Tigo
                    total_comision += total * Decimal('0.05')
            
            suma_pagos = sum(Decimal(str(p.get('monto', 0))) for p in pagos_mixtos)
            total_con_comision = total + total_comision
            diferencia = abs(suma_pagos - total_con_comision)
            
            # Tolerancia de 1 guaraní por redondeo
            if diferencia > Decimal('1'):
                return JsonResponse({
                    'success': False,
                    'error': f'La suma de pagos (Gs. {int(suma_pagos):,}) no coincide con el total + comisión (Gs. {int(total_con_comision):,}). Diferencia: Gs. {int(diferencia):,}'
                })
        
        # Obtener empleado actual (si existe en el modelo)
        try:
            empleado = Empleado.objects.get(usuario=request.user.username)
        except Empleado.DoesNotExist:
            # Buscar o crear rol genérico
            rol_generico, _ = TipoRolGeneral.objects.get_or_create(
                nombre_rol='SISTEMA',
                defaults={'descripcion': 'Rol del sistema'}
            )
            
            # Crear o buscar empleado genérico
            empleado, created = Empleado.objects.get_or_create(
                usuario='SISTEMA',
                defaults={
                    'id_rol': rol_generico,
                    'nombre': 'SISTEMA',
                    'apellido': 'POS',
                    'contrasena_hash': 'N/A',
                    'direccion': 'N/A',
                    'ciudad': 'N/A',
                    'telefono': '0',
                    'email': 'sistema@pos.local',
                    'activo': True
                }
            )
        
        # Obtener o crear cliente genérico
        try:
            # Obtener lista de precios por defecto
            lista_precios_default = ListaPrecios.objects.filter(activo=True).first()
            if not lista_precios_default:
                return JsonResponse({
                    'success': False,
                    'error': 'No hay lista de precios activa en el sistema'
                })
            
            # Obtener tipo de cliente por defecto
            tipo_cliente_default = TipoCliente.objects.first()
            if not tipo_cliente_default:
                return JsonResponse({
                    'success': False,
                    'error': 'No hay tipo de cliente configurado en el sistema'
                })
            
            cliente_generico, created = Cliente.objects.get_or_create(
                ruc_ci='0000000',
                defaults={
                    'nombres': 'CLIENTE',
                    'apellidos': 'GENERICO',
                    'id_lista': lista_precios_default,
                    'id_tipo_cliente': tipo_cliente_default,
                    'direccion': 'N/A',
                    'ciudad': 'N/A',
                    'telefono': '0',
                    'email': 'generico@sistema.local',
                    'limite_credito': 0,
                    'activo': True
                }
            )
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al crear cliente genérico: {str(e)}'
            })
        
        # Obtener tarjeta si se especificó
        tarjeta = None
        cliente = cliente_generico  # Por defecto usar cliente genérico
        tiene_tarjeta_exclusiva = False  # Flag para determinar si usa tarjeta exclusiva
        
        if tarjeta_data and tarjeta_data.get('id'):
            try:
                nro_tarjeta = tarjeta_data['id']
                
                # Usar select_for_update() para bloquear la tarjeta durante toda la transacción
                tarjeta = Tarjeta.objects.select_for_update().select_related(
                    'id_hijo',
                    'id_hijo__id_cliente_responsable'
                ).get(nro_tarjeta=nro_tarjeta)
                
                # Usar el cliente de la tarjeta
                if tarjeta.id_hijo and tarjeta.id_hijo.id_cliente_responsable:
                    cliente = tarjeta.id_hijo.id_cliente_responsable
                
                # Si se usa SOLO tarjeta (sin pagos mixtos externos), es consumo exclusivo
                usa_solo_tarjeta = not pagos_mixtos or all(
                    p.get('medio_id') == 6 for p in pagos_mixtos  # ID 6 = Tarjeta Estudiantil
                )
                
                if usa_solo_tarjeta:
                    tiene_tarjeta_exclusiva = True
                    # Validar saldo solo si es consumo exclusivo de tarjeta
                    if tarjeta.saldo_actual < total:
                        return JsonResponse({
                            'success': False,
                            'error': f'Saldo insuficiente. Disponible: Gs. {tarjeta.saldo_actual:,.0f}',
                            'requiere_autorizacion_supervisor': True,  # Indicar que puede autorizarse
                            'monto_faltante': int(total - tarjeta.saldo_actual)
                        })
                    
            except Tarjeta.DoesNotExist:
                pass
            except Exception as e:
                pass        # ================================================
        # DETERMINAR TIPO DE VENTA Y EMISIÓN DE FACTURA
        # ================================================
        # LÓGICA DE NEGOCIO:
        # 1. Con tarjeta exclusiva (solo saldo): NO emite factura → tipo_venta = 'CONTADO'
        # 2. Con pagos externos (efectivo, débito, etc.): SÍ emite factura → tipo_venta = 'CONTADO'
        # 3. Con autorización supervisor (saldo insuficiente): SÍ emite factura → tipo_venta = 'CREDITO'
        
        nro_factura = None
        tipo_venta_final = 'CONTADO'  # Default
        genera_factura_legal = False
        autorizado_por_id = data.get('autorizado_por_id')  # ID del supervisor que autoriza
        motivo_credito = data.get('motivo_credito', '')  # Justificación del crédito
        
        # Verificar si hay pagos con medios externos (que no sean tarjeta exclusiva)
        tiene_pagos_externos = False
        if pagos_mixtos:
            tiene_pagos_externos = any(
                p.get('medio_id') != 6 for p in pagos_mixtos  # ID 6 = Tarjeta Estudiantil
            )
        
        # CASO 1: Venta a CRÉDITO con autorización de supervisor
        if autorizado_por_id:
            tipo_venta_final = 'CREDITO'
            genera_factura_legal = True
            
            # Validar que el supervisor existe
            try:
                supervisor = Empleado.objects.get(id_empleado=autorizado_por_id)
                print(f"✅ Autorización de supervisor: {supervisor.nombre} {supervisor.apellido}")
            except Empleado.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Supervisor no encontrado'
                })
        
        # CASO 2: Venta con pagos externos (efectivo, débito, QR, etc.)
        elif tiene_pagos_externos:
            tipo_venta_final = 'CONTADO'
            genera_factura_legal = True
        
        # CASO 3: Venta solo con tarjeta exclusiva
        elif tiene_tarjeta_exclusiva:
            tipo_venta_final = 'CONTADO'
            genera_factura_legal = False
        
        # CASO 4: Venta sin tarjeta (cliente genérico con efectivo)
        else:
            tipo_venta_final = 'CONTADO'
            genera_factura_legal = True
        
        print(f"📋 TIPO DE VENTA: {tipo_venta_final} | Genera factura legal: {genera_factura_legal}")
        
        # ================================================
        # EMITIR FACTURA LEGAL SI CORRESPONDE
        # ================================================
        if genera_factura_legal:
            try:
                # Buscar timbrado activo
                timbrado_activo = Timbrados.objects.filter(activo=True).first()
                if not timbrado_activo:
                    return JsonResponse({
                        'success': False,
                        'error': 'No hay timbrado activo configurado para emitir facturas'
                    })
                
                # Obtener el último nro_secuencial para este timbrado
                ultimo_doc = DocumentosTributarios.objects.filter(
                    nro_timbrado=timbrado_activo
                ).order_by('-nro_secuencial').first()
                
                nuevo_secuencial = (ultimo_doc.nro_secuencial + 1) if ultimo_doc else 1
                
                # Crear documento tributario
                documento = DocumentosTributarios.objects.create(
                    nro_timbrado=timbrado_activo,
                    nro_secuencial=nuevo_secuencial,
                    fecha_emision=timezone.now(),
                    monto_total=int(total),
                    monto_exento=0,
                    monto_gravado_5=0,
                    monto_iva_5=0,
                    monto_gravado_10=0,
                    monto_iva_10=0
                )
                nro_factura = documento.id_documento
                print(f"📄 FACTURA LEGAL generada: #{nro_factura} (Timbrado: {timbrado_activo.nro_timbrado})")
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Error al crear factura: {str(e)}'
                })
        else:
            nro_factura = None
            print(f"ℹ️ NO se genera factura legal (consumo con tarjeta exclusiva)")
        
        # ================================================
        # CREAR LA VENTA
        # ================================================
        venta = Ventas.objects.create(
            nro_factura_venta=nro_factura,  # NULL si es con tarjeta, ID_Documento si es contado/crédito
            id_cliente=cliente,  # Usar el cliente obtenido (genérico o de tarjeta)
            id_hijo=tarjeta.id_hijo if tarjeta else None,
            id_tipo_pago_id=tipo_pago_id,
            id_empleado_cajero=empleado,
            fecha=timezone.now(),
            monto_total=int(total),
            estado_pago='PAGADA',
            estado='PROCESADO',
            tipo_venta=tipo_venta_final,  # 'CONTADO' o 'CREDITO'
            autorizado_por_id=autorizado_por_id if autorizado_por_id else None,
            motivo_credito=motivo_credito if tipo_venta_final == 'CREDITO' else None,
            genera_factura_legal=genera_factura_legal
        )
        
        # Crear detalles de venta
        for item in items:
            try:
                producto = Producto.objects.select_related(
                    'id_categoria',
                    'stock'
                ).get(id_producto=item['id'])
                
                # Manejar cantidad o peso según el tipo de producto
                if item.get('esPorKilo'):
                    cantidad = Decimal(str(item.get('peso', 1)))
                else:
                    cantidad = Decimal(str(item.get('quantity', 1)))
                    
                precio = Decimal(str(item['price']))
                subtotal = precio * cantidad
                
                # Crear detalle (simplificado para venta con tarjeta, completo para contado/crédito)
                DetalleVenta.objects.create(
                    id_venta=venta,
                    id_producto=producto,
                    cantidad=cantidad,
                    precio_unitario=int(precio),
                    subtotal_total=int(subtotal)
                )
                
                # Actualizar stock
                try:
                    stock = StockUnico.objects.get(id_producto=producto)
                    stock.stock_actual = F('stock_actual') - cantidad
                    stock.save()
                except StockUnico.DoesNotExist:
                    pass
                    
            except Producto.DoesNotExist:
                continue
        
        # Si hay tarjeta Y es uso exclusivo, descontar saldo (proceso simplificado)
        if tarjeta and tiene_tarjeta_exclusiva:
            print(f"DEBUG: Descontando saldo de tarjeta {tarjeta.nro_tarjeta}")
            print(f"DEBUG: Saldo anterior: {tarjeta.saldo_actual}, Total a descontar: {total}")
            
            # Guardar saldo anterior para el registro ANTES de cualquier modificación
            saldo_anterior = tarjeta.saldo_actual
            
            # Calcular saldo posterior
            saldo_posterior = saldo_anterior - total
            
            print(f"DEBUG: Saldo posterior calculado: {saldo_posterior}")
            
            # Registrar consumo PRIMERO, antes de actualizar la tarjeta
            try:
                saldo_ant_int = int(saldo_anterior)
                saldo_post_int = int(saldo_posterior)
                tarjeta_id = tarjeta.nro_tarjeta  # Guardar el ID antes de cualquier modificación
                
                print(f"DEBUG PRE-CREATE: saldo_anterior={saldo_ant_int}, saldo_posterior={saldo_post_int}")
                
                # Usar el ID directamente en lugar del objeto para evitar que Django refresque
                consumo_creado = ConsumoTarjeta.objects.create(
                    nro_tarjeta_id=tarjeta_id,  # Usar _id para evitar refresh del FK
                    fecha_consumo=timezone.now(),
                    monto_consumido=int(total),
                    detalle=f'Venta #{venta.id_venta}',
                    saldo_anterior=saldo_ant_int,
                    saldo_posterior=saldo_post_int
                )
                
                print(f"DEBUG POST-CREATE: Consumo {consumo_creado.id_consumo} creado")
                print(f"DEBUG POST-CREATE: Valores enviados - Ant={saldo_ant_int}, Post={saldo_post_int}")
                
                # Verificar inmediatamente en la BD
                consumo_verificado = ConsumoTarjeta.objects.get(id_consumo=consumo_creado.id_consumo)
                print(f"DEBUG VERIFICACION DB: Consumo {consumo_verificado.id_consumo} en BD - Ant={consumo_verificado.saldo_anterior}, Post={consumo_verificado.saldo_posterior}")
                
                print(f"DEBUG: Consumo registrado exitosamente")
            except Exception as e:
                print(f"ERROR: No se pudo registrar el consumo: {e}")
                raise  # Re-lanzar para que la transacción haga rollback
            
            # AHORA actualizar el saldo de la tarjeta (después del consumo)
            tarjeta.saldo_actual = saldo_posterior
            tarjeta.save()
            print(f"DEBUG: Saldo de tarjeta actualizado a {tarjeta.saldo_actual}")
        
        # ⚠️ REGISTRAR EN AUDITORÍA SI SE CONFIRMARON RESTRICCIONES ALIMENTARIAS
        if restricciones_confirmadas and tarjeta and tarjeta.id_hijo:
            hijo = tarjeta.id_hijo
            descripcion = f'Venta #{venta.id_venta} procesada con RESTRICCIONES ALIMENTARIAS confirmadas'
            if justificacion_restricciones:
                descripcion += f' - Justificación del cajero: {justificacion_restricciones}'
            descripcion += f' - Estudiante: {hijo.descripcions} {hijo.apellidos} (Tarjeta #{tarjeta.nro_tarjeta})'
            descripcion += f' - Restricciones activas: {hijo.restricciones_compra[:100]}...' if len(hijo.restricciones_compra or '') > 100 else f' - Restricciones: {hijo.restricciones_compra}'
            
            registrar_auditoria(
                request=request,
                operacion='VENTA_CON_RESTRICCIONES',
                tipo_usuario='CAJERO',
                tabla_afectada='ventas',
                id_registro=venta.id_venta,
                descripcion=descripcion,
                resultado='EXITOSO'
            )
            print(f"⚠️ AUDITORÍA: Venta con restricciones confirmadas registrada - Venta #{venta.id_venta}")
        
        # 🎉 REGISTRAR PROMOCIÓN APLICADA si hay alguna
        if promocion_id and descuento_promocion > 0:
            try:
                from .promociones_utils import registrar_promocion_aplicada
                registrar_promocion_aplicada(venta.id_venta, promocion_id, float(descuento_promocion))
                print(f"🎉 PROMOCIÓN: Registrada promoción #{promocion_id} en venta #{venta.id_venta} - Descuento: Gs. {descuento_promocion:,.0f}")
            except Exception as e:
                # No fallar la venta si hay error al registrar promoción
                print(f"⚠️ ERROR al registrar promoción aplicada: {e}")
        
        # 💰 REGISTRAR PAGOS MIXTOS si existen
        if pagos_mixtos:
            for pago_data in pagos_mixtos:
                try:
                    medio_id = pago_data.get('medio_id')
                    monto_pago = Decimal(str(pago_data.get('monto', 0)))
                    
                    # Cargar el medio de pago
                    try:
                        medio_pago = MediosPago.objects.get(id_medio_pago=medio_id)
                    except MediosPago.DoesNotExist:
                        print(f"⚠️ ERROR: Medio de pago #{medio_id} no encontrado")
                        continue
                    
                    # VALIDAR REFERENCIA DE TRANSACCIÓN para medios que la requieren
                    requiere_referencia = medio_id in [2, 3, 4, 5]  # Débito, Crédito, Tigo, Transferencia
                    referencia = pago_data.get('referencia', '').strip()
                    
                    if requiere_referencia and not referencia:
                        return JsonResponse({
                            'success': False,
                            'error': f'El medio de pago {medio_pago.descripcion} requiere un código de transacción'
                        })
                    
                    # Calcular comisión si el medio de pago la genera
                    comision = Decimal('0')
                    porcentaje_comision = Decimal('0')
                    if medio_pago.genera_comision:
                        # Usar porcentajes fijos según el medio de pago
                        if medio_id == 3:  # Tarjeta de Débito
                            porcentaje_comision = Decimal('0.03')  # 3%
                        elif medio_id == 2:  # Tarjeta de Crédito
                            porcentaje_comision = Decimal('0.05')  # 5%
                        elif medio_id == 4:  # Giros Tigo
                            porcentaje_comision = Decimal('0.05')  # 5%
                        
                        if porcentaje_comision > 0:
                            comision = total * porcentaje_comision  # Calcular sobre el total de productos
                            print(f"💳 Comisión calculada: {porcentaje_comision * 100}% sobre Gs. {int(total):,} = Gs. {int(comision):,}")
                        else:
                            print(f"⚠️ No hay porcentaje de comisión definido para {medio_pago.descripcion}")
                    
                    # Crear registro de pago
                    pago_venta = PagosVenta.objects.create(
                        id_venta=venta,
                        id_medio_pago=medio_pago,
                        nro_tarjeta_usada=tarjeta if (tarjeta and medio_id == 6) else None,  # Solo asociar tarjeta si es pago con tarjeta estudiantil
                        monto_aplicado=int(monto_pago),
                        referencia_transaccion=pago_data.get('referencia'),  # Capturar referencia de transacción si existe
                        fecha_pago=timezone.now()
                    )
                    
                    # Registrar comisión en detalle_comision_venta si hay comisión
                    if comision > 0:
                        # Crear o obtener tarifa de comisión con el porcentaje aplicado
                        tarifa_aplicada, created = TarifasComision.objects.get_or_create(
                            id_medio_pago=medio_pago,
                            porcentaje_comision=porcentaje_comision,
                            activo=True,
                            defaults={
                                'fecha_inicio_vigencia': timezone.now(),
                                'monto_fijo_comision': Decimal('0')
                            }
                        )
                        
                        DetalleComisionVenta.objects.create(
                            id_pago_venta=pago_venta,
                            id_tarifa=tarifa_aplicada,
                            monto_comision_calculada=comision,
                            porcentaje_aplicado=porcentaje_comision
                        )
                        print(f"📊 Comisión registrada en detalle_comision_venta: Gs. {int(comision):,}")
                    
                    print(f"💰 Pago registrado: {medio_pago.descripcion} - Gs. {int(monto_pago):,} (Comisión: Gs. {int(comision):,})")
                    
                except MediosPago.DoesNotExist:
                    print(f"⚠️ ERROR: Medio de pago #{medio_id} no encontrado")
                except Exception as e:
                    print(f"⚠️ ERROR al registrar pago: {e}")
                    import traceback
                    traceback.print_exc()
                    # No fallar la venta, solo registrar el error
        
        # ========================================
        # 🖨️ IMPRIMIR TICKET EN IMPRESORA TÉRMICA
        # ========================================
        ticket_impreso = False
        error_impresion = None
        
        try:
            from gestion.impresora_manager import ImpresoraTermica
            
            impresora = ImpresoraTermica()
            
            # Preparar datos del ticket
            ticket_data = {
                'venta_id': venta.id_venta,
                'fecha': venta.fecha.strftime('%d/%m/%Y %H:%M'),
                'cliente': f"{cliente.nombres} {cliente.apellidos}",
                'cajero': f"{empleado.nombre} {empleado.apellido}",
                'items': [],
                'subtotal': float(total),
                'descuento': float(descuento_promocion) if descuento_promocion > 0 else 0,
                'total': float(total),
                'nro_factura': nro_factura,
                'tipo_venta': tipo_venta_final
            }
            
            # Agregar items
            for item in items:
                ticket_data['items'].append({
                    'descripcion': item.get('name', 'Producto'),
                    'cantidad': item.get('quantity', 1),
                    'precio': item.get('price', 0),
                    'subtotal': item.get('quantity', 1) * item.get('price', 0)
                })
            
            # Agregar información de tarjeta si existe
            if tarjeta:
                ticket_data['tarjeta'] = {
                    'nro': tarjeta.nro_tarjeta,
                    'titular': f"{tarjeta.id_hijo.descripcions} {tarjeta.id_hijo.apellidos}" if tarjeta.id_hijo else 'N/A',
                    'saldo_anterior': float(tarjeta.saldo_actual + total) if tiene_tarjeta_exclusiva else None,
                    'saldo_actual': float(tarjeta.saldo_actual) if tiene_tarjeta_exclusiva else None
                }
            
            # Agregar información de pagos mixtos
            if pagos_mixtos:
                ticket_data['pagos'] = []
                for pago in pagos_mixtos:
                    medio_id = pago.get('medio_id')
                    try:
                        medio = MediosPago.objects.get(id_medio_pago=medio_id)
                        ticket_data['pagos'].append({
                            'medio': medio.descripcion,
                            'monto': float(pago.get('monto', 0)),
                            'referencia': pago.get('referencia', '')
                        })
                    except:
                        pass
            
            # Intentar imprimir
            if impresora.imprimir_ticket(ticket_data):
                ticket_impreso = True
                print(f"✅ Ticket impreso correctamente para venta #{venta.id_venta}")
            else:
                error_impresion = "La impresora no respondió"
                print(f"⚠️ No se pudo imprimir el ticket para venta #{venta.id_venta}")
                
        except ImportError:
            error_impresion = "Módulo de impresora no disponible"
            print(f"⚠️ Módulo de impresora no disponible")
        except Exception as e:
            error_impresion = str(e)
            print(f"⚠️ Error al imprimir ticket: {e}")
            # NO fallar la venta si hay error de impresión
        
        # Preparar respuesta con información actualizada
        response_data = {
            'success': True,
            'venta_id': venta.id_venta,
            'total': float(total),
            'message': '¡Venta procesada exitosamente!',
            'ticket_impreso': ticket_impreso
        }
        
        if error_impresion:
            response_data['warning'] = f'Venta exitosa pero no se pudo imprimir: {error_impresion}'
        
        # Si hay tarjeta, incluir información del saldo actualizado
        if tarjeta:
            tarjeta.refresh_from_db()  # Refrescar desde BD para obtener el saldo actualizado
            response_data['saldo_actual'] = float(tarjeta.saldo_actual)
            response_data['tarjeta_nro'] = tarjeta.nro_tarjeta
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def dashboard_view(request):
    """Vista del dashboard con estadísticas y gráficos - Optimizada con cache"""
    from datetime import datetime, timedelta
    from decimal import Decimal
    from django.core.cache import cache
    import json

    # Cache key para dashboard
    cache_key = f'dashboard_stats_{datetime.now().date()}'
    cached_data = cache.get(cache_key)

    if cached_data and not request.GET.get('refresh'):
        return render(request, 'pos/dashboard.html', cached_data)

    today = datetime.now().date()
    week_ago = today - timedelta(days=7)
    month_start = today.replace(day=1)

    # === Estadísticas principales - Optimizadas ===

    # Ventas de hoy con select_related para empleado
    ventas_hoy = Ventas.objects.filter(
        fecha__date=today
    ).select_related('id_empleado_cajero')

    stats_hoy = ventas_hoy.aggregate(
        total=Sum('monto_total'),
        cantidad=Count('id_venta')
    )

    # Ventas del mes
    ventas_mes = Ventas.objects.filter(fecha__date__gte=month_start)
    stats_mes = ventas_mes.aggregate(
        total=Sum('monto_total'),
        cantidad=Count('id_venta')
    )

    # Items vendidos hoy - Optimizado con una sola query
    items_vendidos = DetalleVenta.objects.filter(
        id_venta__fecha__date=today
    ).aggregate(total=Sum('cantidad'))['total'] or 0

    # Ventas de ayer para comparación
    yesterday = today - timedelta(days=1)
    ventas_ayer = Ventas.objects.filter(fecha__date=yesterday).aggregate(
        cantidad=Count('id_venta')
    )['cantidad'] or 0

    # Promedio por venta hoy
    total_hoy = stats_hoy['total'] or Decimal('0')
    cantidad_hoy = stats_hoy['cantidad'] or 1
    promedio = total_hoy / cantidad_hoy if cantidad_hoy > 0 else Decimal('0')

    stats = {
        'ventas_hoy': stats_hoy['cantidad'] or 0,
        'total_hoy': total_hoy,
        'ventas_mes': stats_mes['cantidad'] or 0,
        'total_mes': stats_mes['total'] or Decimal('0'),
        'items_vendidos': items_vendidos,
        'promedio': promedio,
        'ventas_ayer': ventas_ayer  # Para comparación
    }
    
    # === Ventas por hora (hoy) ===
    from django.db.models.functions import ExtractHour
    
    ventas_por_hora_raw = ventas_hoy.annotate(
        hora=ExtractHour('fecha')
    ).values('hora').annotate(
        total=Count('id_venta')
    ).order_by('hora')
    
    # Crear array de 24 horas con valores 0
    ventas_hora_dict = {i: 0 for i in range(24)}
    for item in ventas_por_hora_raw:
        hora = int(item['hora'])
        ventas_hora_dict[hora] = item['total']
    
    ventas_por_hora = {
        'labels': [f"{h:02d}:00" for h in range(24)],
        'data': [ventas_hora_dict[h] for h in range(24)]
    }
    
    # === Top 10 productos vendidos (hoy) ===
    top_productos_raw = DetalleVenta.objects.filter(
        id_venta__fecha__date=today
    ).select_related('id_producto').values(
        'id_producto__descripcion'
    ).annotate(
        total=Sum('cantidad')
    ).order_by('-total')[:10]
    
    top_productos = {
        'labels': [item['id_producto__descripcion'][:20] for item in top_productos_raw],
        'data': [item['total'] for item in top_productos_raw]
    }
    
    # === Ventas últimos 7 días ===
    from django.db.models.functions import TruncDate
    
    ventas_semanales_raw = Ventas.objects.filter(
        fecha__date__gte=week_ago
    ).annotate(
        dia=TruncDate('fecha')
    ).values('dia').annotate(
        total=Sum('monto_total'),
        cantidad=Count('id_venta')
    ).order_by('dia')
    
    # Crear dict con todos los días
    ventas_semana_dict = {}
    for i in range(7):
        dia = (week_ago + timedelta(days=i)).isoformat()
        ventas_semana_dict[dia] = 0
    
    for item in ventas_semanales_raw:
        dia_str = str(item['dia'])
        ventas_semana_dict[dia_str] = float(item['total'] or 0)
    
    ventas_semanales = {
        'labels': [
            (week_ago + timedelta(days=i)).strftime('%d/%m')
            for i in range(7)
        ],
        'data': [
            ventas_semana_dict[(week_ago + timedelta(days=i)).isoformat()]
            for i in range(7)
        ]
    }
    
    # === Ventas por categoría (hoy) ===
    ventas_categoria_raw = DetalleVenta.objects.filter(
        id_venta__fecha__date=today
    ).select_related(
        'id_producto',
        'id_producto__id_categoria'
    ).values(
        'id_producto__id_categoria__nombre'
    ).annotate(
        total=Sum('cantidad')
    ).order_by('-total')
    
    ventas_por_categoria = {
        'labels': [item['id_producto__id_categoria__nombre'] or 'Sin categoría' 
                   for item in ventas_categoria_raw],
        'data': [item['total'] for item in ventas_categoria_raw]
    }
    
    # === Últimas 10 ventas ===
    ultimas_ventas = Ventas.objects.select_related(
        'id_empleado_cajero'
    ).filter(
        fecha__date=today
    ).annotate(
        empleado_nombre=F('id_empleado_cajero__nombre')
    ).order_by('-fecha')[:10]
    
    # === Stock bajo ===
    stock_bajo = Producto.objects.filter(
        activo=True,
        stock_minimo__isnull=False
    ).annotate(
        stock_actual_val=F('stock__stock_actual')
    ).filter(
        stock_actual_val__lt=F('stock_minimo')
    ).select_related('id_categoria', 'stock', 'id_unidad_de_medida').annotate(
        stock_actual=F('stock__stock_actual')
    ).order_by('stock__stock_actual')[:10]
    
    # === NOTIFICACIONES DE VALIDACIÓN (ADMINISTRADOR) ===
    
    # Cargas de saldo pendientes de validación
    cargas_pendientes = CargasSaldo.objects.filter(
        estado='PENDIENTE'
    ).select_related(
        'nro_tarjeta',
        'nro_tarjeta__id_hijo',
        'id_cliente_origen'
    ).order_by('-fecha_carga')[:20]
    
    # Pagos pendientes de validación (transferencias bancarias)
    pagos_pendientes = Ventas.objects.filter(
        motivo_credito__icontains='PAGO_PENDIENTE_TRANSFERENCIA'
    ).select_related(
        'id_cliente',
        'id_empleado_cajero'
    ).order_by('-fecha')[:20]
    
    # Contadores para notificaciones
    total_cargas_pendientes = CargasSaldo.objects.filter(estado='PENDIENTE').count()
    total_pagos_pendientes = Ventas.objects.filter(
        motivo_credito__icontains='PAGO_PENDIENTE_TRANSFERENCIA'
    ).count()
    
    context = {
        'stats': stats,
        'ventas_por_hora': json.dumps(ventas_por_hora),
        'top_productos': json.dumps(top_productos),
        'ventas_semanales': json.dumps(ventas_semanales),
        'ventas_por_categoria': json.dumps(ventas_por_categoria),
        'ultimas_ventas': ultimas_ventas,
        'stock_bajo': stock_bajo,
        # Notificaciones de validación
        'cargas_pendientes': cargas_pendientes,
        'pagos_pendientes': pagos_pendientes,
        'total_cargas_pendientes': total_cargas_pendientes,
        'total_pagos_pendientes': total_pagos_pendientes,
    }

    # Cachear el contexto por 5 minutos
    cache.set(cache_key, context, 300)

    return render(request, 'pos/dashboard.html', context)


@login_required
def historial_view(request):
    """Historial de ventas"""
    ventas = Ventas.objects.select_related(
        'id_empleado_cajero'
    ).prefetch_related(
        'detalles',
        'detalles__id_producto'
    ).annotate(
        items_count=Count('detalles')
    ).order_by('-fecha')[:100]
    
    context = {
        'ventas': ventas,
    }
    return render(request, 'pos/historial.html', context)


@login_required
def reportes_view(request):
    """Reportes y estadísticas"""
    from datetime import datetime, timedelta
    from decimal import Decimal
    
    # Obtener parámetros
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    tipo_reporte = request.GET.get('tipo_reporte', 'ventas')
    
    # Fechas por defecto: mes actual
    if not fecha_desde or not fecha_hasta:
        hoy = datetime.now().date()
        fecha_desde = hoy.replace(day=1).isoformat()
        fecha_hasta = hoy.isoformat()
    
    # Convertir a objetos date
    try:
        fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    except ValueError:
        fecha_desde_obj = datetime.now().date().replace(day=1)
        fecha_hasta_obj = datetime.now().date()
        fecha_desde = fecha_desde_obj.isoformat()
        fecha_hasta = fecha_hasta_obj.isoformat()
    
    # Generar datos según tipo de reporte
    datos = []
    columnas = []
    titulo_tabla = ""
    stats = {}
    
    if tipo_reporte == 'ventas':
        titulo_tabla = "Ventas por Período"
        columnas = ['Fecha', 'ID Venta', 'Empleado', 'Items', 'Total (Gs.)']
        
        ventas = Ventas.objects.filter(
            fecha__date__gte=fecha_desde_obj,
            fecha__date__lte=fecha_hasta_obj
        ).select_related('id_empleado_cajero').annotate(
            items_count=Count('detalles')
        ).order_by('-fecha')
        
        for venta in ventas:
            datos.append([
                venta.fecha.strftime('%d/%m/%Y %H:%M'),
                venta.id_venta,
                f"{venta.id_empleado_cajero.nombre} {venta.id_empleado_cajero.apellido}" if venta.id_empleado_cajero else 'N/A',
                venta.items_count,
                f"{int(venta.monto_total):,}".replace(',', '.')
            ])
        
        # Estadísticas
        total_ventas = ventas.aggregate(
            total=Sum('monto_total'),
            cantidad=Count('id_venta')
        )
        stats = {
            'total': total_ventas['cantidad'] or 0,
            'monto_total': total_ventas['total'] or Decimal('0'),
            'promedio': (total_ventas['total'] / total_ventas['cantidad']) if total_ventas['cantidad'] > 0 else Decimal('0')
        }
    
    elif tipo_reporte == 'productos':
        titulo_tabla = "Productos Más Vendidos"
        columnas = ['Producto', 'Código', 'Cantidad Vendida', 'Total Vendido (Gs.)', 'Ventas']
        
        productos = DetalleVenta.objects.filter(
            id_venta__fecha__date__gte=fecha_desde_obj,
            id_venta__fecha__date__lte=fecha_hasta_obj
        ).select_related('id_producto').values(
            'id_producto__descripcion',
            'id_producto__codigo_barra'
        ).annotate(
            cantidad_total=Sum('cantidad'),
            monto_total=Sum('subtotal_total'),
            num_ventas=Count('id_venta', distinct=True)
        ).order_by('-cantidad_total')[:50]
        
        for p in productos:
            datos.append([
                p['id_producto__descripcion'],
                p['id_producto__codigo_barra'],
                p['cantidad_total'],
                f"{int(p['monto_total']):,}".replace(',', '.'),
                p['num_ventas']
            ])
        
        # Estadísticas
        totales = productos.aggregate(
            cantidad=Sum('cantidad_total'),
            monto=Sum('monto_total')
        )
        stats = {
            'total': len(productos),
            'monto_total': totales['monto'] or Decimal('0'),
            'promedio': (totales['monto'] / len(productos)) if len(productos) > 0 else Decimal('0')
        }
    
    elif tipo_reporte == 'empleados':
        titulo_tabla = "Desempeño de Empleados"
        columnas = ['Empleado', 'Rol', 'Ventas', 'Total Vendido (Gs.)', 'Promedio (Gs.)']
        
        empleados = Ventas.objects.filter(
            fecha__date__gte=fecha_desde_obj,
            fecha__date__lte=fecha_hasta_obj
        ).select_related(
            'id_empleado_cajero',
            'id_empleado_cajero__id_rol'
        ).values(
            'id_empleado_cajero__nombre',
            'id_empleado_cajero__apellido',
            'id_empleado_cajero__id_rol__nombre_rol'
        ).annotate(
            num_ventas=Count('id_venta'),
            monto_total=Sum('monto_total')
        ).order_by('-monto_total')
        
        for emp in empleados:
            promedio_emp = (emp['monto_total'] / emp['num_ventas']) if emp['num_ventas'] > 0 else 0
            datos.append([
                f"{emp['id_empleado_cajero__nombre']} {emp['id_empleado_cajero__apellido']}",
                emp['id_empleado_cajero__id_rol__nombre_rol'] or 'N/A',
                emp['num_ventas'],
                f"{int(emp['monto_total']):,}".replace(',', '.'),
                f"{int(promedio_emp):,}".replace(',', '.')
            ])
        
        # Estadísticas
        totales = empleados.aggregate(
            ventas=Sum('num_ventas'),
            monto=Sum('monto_total')
        )
        stats = {
            'total': len(empleados),
            'monto_total': totales['monto'] or Decimal('0'),
            'promedio': (totales['monto'] / totales['ventas']) if totales['ventas'] > 0 else Decimal('0')
        }
    
    elif tipo_reporte == 'stock':
        titulo_tabla = "Reporte de Stock Actual"
        columnas = ['Producto', 'Código', 'Stock Actual', 'Stock Mínimo', 'Estado']
        
        productos_stock = Producto.objects.filter(
            activo=True
        ).select_related('stock').order_by('descripcion')
        
        for p in productos_stock:
            try:
                stock_actual = p.stock.stock_actual
                stock_minimo = p.stock_minimo or 0
                
                if stock_actual == 0:
                    estado = '❌ Agotado'
                elif p.stock_minimo and stock_actual < p.stock_minimo:
                    estado = '⚠️ Bajo'
                else:
                    estado = '✅ OK'
                
                datos.append([
                    p.descripcion,
                    p.codigo,
                    stock_actual,
                    stock_minimo,
                    estado
                ])
            except:
                pass
        
        stats = {'total': len(datos)}
    
    elif tipo_reporte == 'tarjetas':
        titulo_tabla = "Consumos por Tarjeta"
        columnas = ['Tarjeta', 'Estudiante', 'Consumos', 'Total Consumido (Gs.)', 'Saldo Actual (Gs.)']
        
        tarjetas = ConsumoTarjeta.objects.filter(
            fecha_consumo__date__gte=fecha_desde_obj,
            fecha_consumo__date__lte=fecha_hasta_obj
        ).values(
            'nro_tarjeta__nro_tarjeta',
            'nro_tarjeta__id_hijo__nombre',
            'nro_tarjeta__id_hijo__apellido',
            'nro_tarjeta__saldo_actual'
        ).annotate(
            num_consumos=Count('id_consumo_tarjeta'),
            total_consumido=Sum('monto_consumido')
        ).order_by('-total_consumido')[:50]
        
        for t in tarjetas:
            datos.append([
                t['nro_tarjeta__nro_tarjeta'],
                f"{t['nro_tarjeta__id_hijo__nombre']} {t['nro_tarjeta__id_hijo__apellido']}",
                t['num_consumos'],
                f"{int(t['total_consumido']):,}".replace(',', '.'),
                f"{int(t['nro_tarjeta__saldo_actual']):,}".replace(',', '.')
            ])
        
        # Estadísticas
        totales = tarjetas.aggregate(
            consumos=Sum('num_consumos'),
            monto=Sum('total_consumido')
        )
        stats = {
            'total': len(tarjetas),
            'monto_total': totales['monto'] or Decimal('0'),
            'promedio': (totales['monto'] / totales['consumos']) if totales['consumos'] > 0 else Decimal('0')
        }
    
    context = {
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'tipo_reporte': tipo_reporte,
        'datos': datos,
        'columnas': columnas,
        'titulo_tabla': titulo_tabla,
        'stats': stats,
    }
    
    return render(request, 'pos/reportes.html', context)


@login_required
def exportar_reporte(request):
    """Exportar reporte a Excel o PDF"""
    from datetime import datetime
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    import io
    
    # Obtener parámetros
    formato = request.GET.get('formato', 'excel')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    tipo_reporte = request.GET.get('tipo_reporte', 'ventas')
    
    # Convertir fechas
    try:
        fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    except:
        fecha_desde_obj = datetime.now().date().replace(day=1)
        fecha_hasta_obj = datetime.now().date()
    
    # Obtener datos (reutilizar lógica de reportes_view)
    datos, columnas, titulo = obtener_datos_reporte(tipo_reporte, fecha_desde_obj, fecha_hasta_obj)
    
    # Asegurar que el título no esté vacío
    if not titulo or titulo.strip() == '':
        titulo = f'Reporte de {tipo_reporte.title()}'
    
    if formato == 'excel':
        return exportar_excel(datos, columnas, titulo, fecha_desde_obj, fecha_hasta_obj)
    elif formato == 'pdf':
        return exportar_pdf(datos, columnas, titulo, fecha_desde_obj, fecha_hasta_obj)
    else:
        return HttpResponse('Formato no soportado', status=400)


def obtener_datos_reporte(tipo_reporte, fecha_desde, fecha_hasta):
    """Función auxiliar para obtener datos del reporte"""
    from decimal import Decimal
    
    datos = []
    columnas = []
    titulo = ""
    
    if tipo_reporte == 'ventas':
        titulo = "Reporte de Ventas"
        columnas = ['Fecha', 'ID Venta', 'Empleado', 'Items', 'Total (Gs.)']
        
        ventas = Ventas.objects.filter(
            fecha__date__gte=fecha_desde,
            fecha__date__lte=fecha_hasta
        ).select_related('id_empleado').order_by('-fecha')
        
        for venta in ventas:
            items_count = DetalleVenta.objects.filter(id_venta=venta).count()
            datos.append([
                venta.fecha.strftime('%d/%m/%Y %H:%M'),
                str(venta.id_venta),
                f"{venta.id_empleado.nombre} {venta.id_empleado.apellido}" if venta.id_empleado else 'N/A',
                str(items_count),
                f"{int(venta.monto_total):,}".replace(',', '.')
            ])
    
    elif tipo_reporte == 'productos':
        titulo = "Productos Más Vendidos"
        columnas = ['Producto', 'Código', 'Cantidad', 'Total (Gs.)', 'Ventas']
        
        productos = DetalleVenta.objects.filter(
            id_venta__fecha__date__gte=fecha_desde,
            id_venta__fecha__date__lte=fecha_hasta
        ).values(
            'id_producto__descripcion',
            'id_producto__codigo_barra'
        ).annotate(
            cantidad_total=Sum('cantidad'),
            monto_total=Sum('subtotal_total'),
            num_ventas=Count('id_venta', distinct=True)
        ).order_by('-cantidad_total')[:50]
        
        for p in productos:
            datos.append([
                p['id_producto__descripcion'],
                p['id_producto__codigo_barra'],
                str(p['cantidad_total']),
                f"{int(p['monto_total']):,}".replace(',', '.'),
                str(p['num_ventas'])
            ])
    
    elif tipo_reporte == 'empleados':
        titulo = "Desempeño de Empleados"
        columnas = ['Empleado', 'Rol', 'Ventas', 'Total (Gs.)', 'Promedio (Gs.)']
        
        empleados = Ventas.objects.filter(
            fecha__date__gte=fecha_desde,
            fecha__date__lte=fecha_hasta
        ).values(
            'id_empleado__nombre',
            'id_empleado__apellido',
            'id_empleado__id_rol__nombre_rol'
        ).annotate(
            num_ventas=Count('id_venta'),
            monto_total=Sum('monto_total'),
            promedio=Sum('monto_total') / Count('id_venta')
        ).order_by('-monto_total')
        
        for emp in empleados:
            datos.append([
                f"{emp['id_empleado__nombre']} {emp['id_empleado__apellido']}",
                emp['id_empleado__id_rol__nombre_rol'] or 'N/A',
                str(emp['num_ventas']),
                f"{int(emp['monto_total']):,}".replace(',', '.'),
                f"{int(emp['promedio']):,}".replace(',', '.')
            ])
    
    elif tipo_reporte == 'stock':
        titulo = "Reporte de Stock"
        columnas = ['Producto', 'Código', 'Stock Actual', 'Stock Mínimo', 'Estado']
        
        productos = Producto.objects.filter(estado='activo').select_related('stock')
        
        for p in productos:
            try:
                stock_actual = p.stock.stock_actual
                stock_minimo = p.stock_minimo or 0
                
                if stock_actual == 0:
                    estado = 'Agotado'
                elif p.stock_minimo and stock_actual < p.stock_minimo:
                    estado = 'Bajo'
                else:
                    estado = 'OK'
                
                datos.append([
                    p.descripcion,
                    p.codigo,
                    str(stock_actual),
                    str(stock_minimo),
                    estado
                ])
            except:
                pass
    
    elif tipo_reporte == 'tarjetas':
        titulo = "Consumos por Tarjeta"
        columnas = ['Tarjeta', 'Estudiante', 'Consumos', 'Total (Gs.)', 'Saldo (Gs.)']
        
        tarjetas = ConsumoTarjeta.objects.filter(
            fecha_consumo__date__gte=fecha_desde,
            fecha_consumo__date__lte=fecha_hasta
        ).values(
            'nro_tarjeta__nro_tarjeta',
            'nombre',
            'apellido',
            'nro_tarjeta__saldo_actual'
        ).annotate(
            num_consumos=Count('id_consumo_tarjeta'),
            total_consumido=Sum('monto_consumido')
        ).order_by('-total_consumido')[:50]
        
        for t in tarjetas:
            datos.append([
                t['nro_tarjeta__nro_tarjeta'],
                f"{t['nombre']} {t['apellido']}",
                str(t['num_consumos']),
                f"{int(t['total_consumido']):,}".replace(',', '.'),
                f"{int(t['nro_tarjeta__saldo_actual']):,}".replace(',', '.')
            ])
    
    return datos, columnas, titulo


def exportar_excel(datos, columnas, titulo, fecha_desde, fecha_hasta):
    """Exportar reporte a Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # Max 31 caracteres
    
    # Título principal
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = f"{titulo}\n{fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}"
    title_cell.font = Font(size=16, bold=True, color='FFFFFF')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color='FF6B35', end_color='FF6B35', fill_type='solid')
    ws.row_dimensions[1].height = 40
    
    # Encabezados
    header_fill = PatternFill(start_color='4ECDC4', end_color='4ECDC4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col_num, column_title in enumerate(columnas, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for row_num, row_data in enumerate(datos, 4):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Ajustar anchos de columna
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{titulo}_{fecha_desde.strftime('%Y%m%d')}_{fecha_hasta.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def exportar_pdf(datos, columnas, titulo, fecha_desde, fecha_hasta):
    """Exportar reporte a PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from django.http import HttpResponse
    import io
    
    # Crear buffer
    buffer = io.BytesIO()
    
    # Crear documento (landscape para tablas anchas)
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#FF6B35'),
        alignment=TA_CENTER,
        spaceAfter=30
    )
    
    # Título
    title_text = f"{titulo}<br/>{fecha_desde.strftime('%d/%m/%Y')} - {fecha_hasta.strftime('%d/%m/%Y')}"
    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Preparar datos para tabla
    table_data = [columnas] + datos
    
    # Crear tabla
    table = Table(table_data)
    table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4ECDC4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Datos
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        
        # Bordes
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#FF6B35')),
        
        # Filas alternadas
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')])
    ]))
    
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    
    # Preparar respuesta HTTP
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = f"{titulo}_{fecha_desde.strftime('%Y%m%d')}_{fecha_hasta.strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response



@login_required
def ticket_view(request, venta_id):
    """Vista para imprimir ticket de venta"""
    try:
        venta = Ventas.objects.select_related(
            'id_empleado_cajero',
            'id_cliente',
            'id_hijo',
            'id_tipo_pago'
        ).get(id_venta=venta_id)
        
        detalles = DetalleVenta.objects.filter(id_venta=venta).select_related(
            'id_producto',
            'id_producto__id_categoria'
        )
        
        # Verificar si fue pago con tarjeta
        tarjeta = None
        saldo_anterior = None
        saldo_actual = None
        consumo = None
        
        # Intentar obtener información de la tarjeta si es venta con hijo
        if venta.id_hijo:
            try:
                tarjeta = venta.id_hijo.tarjeta
                
                # Buscar el consumo por el detalle exacto (más confiable)
                consumo = ConsumoTarjeta.objects.filter(
                    nro_tarjeta=tarjeta,
                    detalle=f'Venta #{venta.id_venta}'
                ).order_by('-fecha_consumo').first()
                
                if consumo:
                    saldo_anterior = consumo.saldo_anterior
                    saldo_actual = consumo.saldo_posterior
                else:
                    # Si no se encuentra el consumo, usar el saldo actual de la tarjeta
                    # y calcular el anterior
                    saldo_actual = tarjeta.saldo_actual
                    saldo_anterior = saldo_actual + venta.monto_total
            except Exception as e:
                pass
        
        # Obtener datos de la empresa para el ticket
        try:
            from gestion.models import DatosEmpresa
            empresa = DatosEmpresa.objects.first()
        except:
            empresa = None
        
        # 💰 Obtener pagos mixtos si existen
        pagos_venta = PagosVenta.objects.filter(id_venta=venta).select_related('id_medio_pago')
        
        # Calcular comisión total
        total_comision = Decimal('0')
        for pago in pagos_venta:
            try:
                comision_detalle = DetalleComisionVenta.objects.filter(id_pago_venta=pago).first()
                if comision_detalle:
                    total_comision += comision_detalle.monto_comision_calculada
            except:
                pass
        
        # Calcular total a pagar (monto_total + comisión)
        total_a_pagar = venta.monto_total + total_comision
        
        context = {
            'venta': venta,
            'detalles': detalles,
            'tarjeta': tarjeta,
            'saldo_anterior': saldo_anterior,
            'saldo_actual': saldo_actual,
            'consumo': consumo,
            'empresa': empresa,
            'pagos_venta': pagos_venta,  # Nuevo: lista de pagos
            'total_comision': total_comision,
            'total_a_pagar': total_a_pagar,  # Total a pagar incluyendo comisión
        }
        
        return render(request, 'pos/ticket.html', context)
        
    except Ventas.DoesNotExist:
        return JsonResponse({'error': 'Venta no encontrada'}, status=404)


@login_required
def recargas_view(request):
    """Vista principal del módulo de recargas"""
    from datetime import datetime
    
    hoy = datetime.now().date()
    
    # Estadísticas del día
    recargas_hoy = CargasSaldo.objects.filter(fecha_carga__date=hoy)
    stats_hoy = recargas_hoy.aggregate(
        total=Sum('monto_cargado'),
        cantidad=Count('id_carga')
    )
    
    total_hoy = stats_hoy['total'] or Decimal('0')
    cantidad_hoy = stats_hoy['cantidad'] or 1
    promedio_hoy = total_hoy / cantidad_hoy if cantidad_hoy > 0 else Decimal('0')
    
    stats = {
        'recargas_hoy': stats_hoy['cantidad'] or 0,
        'total_hoy': total_hoy,
        'promedio_hoy': promedio_hoy
    }
    
    # Últimas 10 recargas del día
    ultimas_recargas = CargasSaldo.objects.filter(
        fecha_carga__date=hoy
    ).select_related('nro_tarjeta', 'nro_tarjeta__id_hijo').order_by('-fecha_carga')[:10]
    
    context = {
        'stats': stats,
        'ultimas_recargas': ultimas_recargas,
    }
    
    return render(request, 'pos/recargas.html', context)


@login_required
def procesar_recarga(request):
    """Procesar recarga de tarjeta"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        
        nro_tarjeta = data.get('nro_tarjeta')
        monto = Decimal(str(data.get('monto', 0)))
        forma_pago = data.get('forma_pago', 'efectivo')
        observaciones = data.get('observaciones', '')
        
        # Validaciones
        if not nro_tarjeta or monto < 1000:
            return JsonResponse({'error': 'Datos inválidos'}, status=400)
        
        # Buscar tarjeta
        try:
            tarjeta = Tarjeta.objects.get(nro_tarjeta=nro_tarjeta)
        except Tarjeta.DoesNotExist:
            return JsonResponse({'error': 'Tarjeta no encontrada'}, status=404)
        
        # Verificar estado
        if tarjeta.estado == 'bloqueada':
            return JsonResponse({'error': 'Tarjeta bloqueada'}, status=400)
        
        # Obtener datos del hijo
        hijo = tarjeta.id_hijo
        
        # Calcular comisión si aplica
        comision = Decimal('0')
        if forma_pago in ['tarjeta_credito', 'tarjeta_debito', 'giros_tigo']:
            # Buscar medio de pago
            try:
                nombre_medio = {
                    'tarjeta_credito': 'Tarjeta de Crédito',
                    'tarjeta_debito': 'Tarjeta de Débito',
                    'giros_tigo': 'Giros Tigo'
                }.get(forma_pago, '')
                
                medio_pago = MediosPago.objects.filter(descripcion=nombre_medio).first()
                
                if medio_pago:
                    # Buscar tarifa de comisión
                    tarifa = TarifasComision.objects.filter(
                        id_medio_pago=medio_pago,
                        activo=True
                    ).first()
                    
                    if tarifa:
                        # Calcular: monto_fijo + (monto * porcentaje / 100)
                        comision = tarifa.monto_fijo + (monto * tarifa.porcentaje / Decimal('100'))
                        comision = comision.quantize(Decimal('0.01'))
            except Exception as e:
                print(f"Error calculando comisión: {e}")
                # Continuar sin comisión si hay error
        
        # Guardar saldo anterior
        saldo_anterior = tarjeta.saldo_actual
        saldo_posterior = saldo_anterior + monto
        
        # Registrar recarga
        recarga = CargasSaldo.objects.create(
            nro_tarjeta=tarjeta,
            fecha=timezone.now(),
            monto=monto,
            forma_pago=forma_pago,
            saldo_anterior=saldo_anterior,
            saldo_posterior=saldo_posterior,
            nombre=hijo.nombre,
            apellido=hijo.apellido,
            id_empleado=request.user if hasattr(request.user, 'id_empleado') else None,
            observaciones=observaciones[:200] if observaciones else None
        )
        
        # Actualizar saldo de tarjeta
        tarjeta.saldo_actual = F('saldo_actual') + monto
        tarjeta.save()
        
        # Refrescar tarjeta para obtener nuevo saldo
        tarjeta.refresh_from_db()
        
        # Si hay comisión, registrarla (esto sería en DetalleComisionVenta pero para recargas)
        # Por ahora solo la calculamos y mostramos
        
        response_data = {
            'success': True,
            'message': 'Recarga procesada exitosamente',
            'recarga_id': recarga.id_carga_saldo,
            'nuevo_saldo': float(tarjeta.saldo_actual),
            'monto': float(monto)
        }
        
        if comision > 0:
            response_data['comision'] = float(comision)
            response_data['message'] += f' (Comisión POS: Gs. {int(comision):,})'
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def historial_recargas_view(request):
    """Vista del historial de recargas"""
    from datetime import datetime, timedelta
    from django.core.paginator import Paginator
    
    # Obtener filtros
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    buscar = request.GET.get('buscar', '')
    
    # Fechas por defecto: últimos 30 días
    if not fecha_desde or not fecha_hasta:
        hoy = datetime.now().date()
        fecha_desde = (hoy - timedelta(days=30)).isoformat()
        fecha_hasta = hoy.isoformat()
    
    # Convertir a objetos date
    try:
        fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    except ValueError:
        fecha_desde_obj = (datetime.now() - timedelta(days=30)).date()
        fecha_hasta_obj = datetime.now().date()
        fecha_desde = fecha_desde_obj.isoformat()
        fecha_hasta = fecha_hasta_obj.isoformat()
    
    # Consulta base
    recargas_query = CargasSaldo.objects.filter(
        fecha_carga__date__gte=fecha_desde_obj,
        fecha_carga__date__lte=fecha_hasta_obj
    ).select_related('nro_tarjeta', 'nro_tarjeta__id_hijo')
    
    # Filtro de búsqueda
    if buscar:
        recargas_query = recargas_query.filter(
            Q(nro_tarjeta__nro_tarjeta__icontains=buscar) |
            Q(nro_tarjeta__id_hijo__nombre__icontains=buscar) |
            Q(nro_tarjeta__id_hijo__apellido__icontains=buscar)
        )
    
    # Ordenar por fecha descendente
    recargas_query = recargas_query.order_by('-fecha_carga')
    
    # Estadísticas del período
    stats_periodo = recargas_query.aggregate(
        total_recargas=Count('id_carga'),
        monto_total=Sum('monto_cargado'),
        tarjetas_unicas=Count('nro_tarjeta', distinct=True)
    )
    
    promedio = (stats_periodo['monto_total'] / stats_periodo['total_recargas']) if stats_periodo['total_recargas'] > 0 else Decimal('0')
    
    stats = {
        'total_recargas': stats_periodo['total_recargas'] or 0,
        'monto_total': stats_periodo['monto_total'] or Decimal('0'),
        'promedio': promedio,
        'tarjetas_unicas': stats_periodo['tarjetas_unicas'] or 0
    }
    
    # Paginación
    paginator = Paginator(recargas_query, 50)
    page_number = request.GET.get('page', 1)
    recargas = paginator.get_page(page_number)
    
    # Query string para paginación
    query_params = request.GET.copy()
    if 'page' in query_params:
        query_params.pop('page')
    query_string = '&' + query_params.urlencode() if query_params else ''
    
    context = {
        'recargas': recargas,
        'stats': stats,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'buscar': buscar,
        'query_string': query_string,
    }
    
    return render(request, 'pos/historial_recargas.html', context)


@login_required
def comprobante_recarga_view(request, recarga_id):
    """Vista para imprimir comprobante de recarga"""
    try:
        recarga = CargasSaldo.objects.select_related(
            'nro_tarjeta',
            'nro_tarjeta__id_hijo',
            'nro_tarjeta__id_hijo__id_cliente_responsable'
        ).get(id_carga=recarga_id)
        
        context = {
            'recarga': recarga,
        }
        
        return render(request, 'pos/comprobante_recarga.html', context)
        
    except CargasSaldo.DoesNotExist:
        return JsonResponse({'error': 'Recarga no encontrada'}, status=404)


# ==================== CUENTA CORRIENTE ====================

@login_required
def cuenta_corriente_view(request):
    """Vista principal de cuenta corriente"""
    # Obtener filtros
    buscar = request.GET.get('buscar', '').strip()
    estado_filter = request.GET.get('estado', '')
    con_credito = request.GET.get('con_credito', '')
    
    # Query base - contar hijos por cliente
    from django.db.models import Count
    clientes = Cliente.objects.annotate(
        num_hijos=Count('hijos')
    ).select_related('id_lista', 'id_tipo_cliente')
    
    # Aplicar filtros
    if buscar:
        clientes = clientes.filter(
            Q(nombres__icontains=buscar) | 
            Q(apellidos__icontains=buscar) |
            Q(ruc_ci__icontains=buscar)
        )
    
    if estado_filter:
        clientes = clientes.filter(activo=(estado_filter == 'activo'))
    
    if con_credito == 'si':
        clientes = clientes.filter(limite_credito__gt=0)
    elif con_credito == 'no':
        clientes = clientes.filter(Q(limite_credito__isnull=True) | Q(limite_credito=0))
    
    # Estadísticas generales
    from django.db.models import Count as CountFunc
    stats = Cliente.objects.aggregate(
        total_clientes=CountFunc('id_cliente'),
        limite_total=Sum('limite_credito')
    )
    
    clientes_con_credito = Cliente.objects.filter(limite_credito__gt=0).count()
    
    context = {
        'clientes': clientes.order_by('-fecha_registro')[:100],
        'stats': stats,
        'clientes_con_credito': clientes_con_credito,
        'buscar': buscar,
        'estado': estado_filter,
        'con_credito': con_credito,
    }
    
    return render(request, 'pos/cuenta_corriente.html', context)


@login_required
def cc_detalle_view(request, cliente_id):
    """Vista de detalle de cuenta corriente de un cliente"""
    try:
        cliente = Cliente.objects.select_related(
            'id_lista', 'id_tipo_cliente'
        ).get(id_cliente=cliente_id)
        
        # Obtener hijos del cliente
        hijos = Hijo.objects.filter(id_cliente_responsable=cliente).select_related('tarjeta')
        
        # Obtener ventas relacionadas con las tarjetas de los hijos
        ventas = Ventas.objects.filter(
            id_hijo__id_cliente_responsable=cliente
        ).select_related('id_hijo', 'id_empleado_cajero').order_by('-fecha')[:50]
        
        # Obtener recargas del cliente
        recargas = CargasSaldo.objects.filter(
            nro_tarjeta__id_hijo__id_cliente_responsable=cliente
        ).select_related('nro_tarjeta').order_by('-fecha_carga')[:50]
        
        # Calcular totales
        total_ventas = ventas.aggregate(total=Sum('monto_total'))['total'] or 0
        total_recargas = recargas.aggregate(total=Sum('monto_cargado'))['total'] or 0
        
        context = {
            'cliente': cliente,
            'hijos': hijos,
            'ventas': ventas,
            'recargas': recargas,
            'total_ventas': total_ventas,
            'total_recargas': total_recargas,
        }
        
        return render(request, 'pos/cc_detalle.html', context)
        
    except Cliente.DoesNotExist:
        return JsonResponse({'error': 'Cliente no encontrado'}, status=404)


@login_required
def cuenta_corriente_unificada(request, cliente_id):
    """Vista unificada de cuenta corriente con timeline, tabla y gráficos"""
    try:
        from datetime import datetime, timedelta
        from decimal import Decimal
        
        # Obtener cliente
        cliente = Cliente.objects.select_related(
            'id_lista', 'id_tipo_cliente'
        ).get(id_cliente=cliente_id)
        
        # Obtener hijos con tarjetas
        hijos = Hijo.objects.filter(
            id_cliente_responsable=cliente
        ).select_related('tarjeta')
        
        # Obtener filtros
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        tipo_movimiento = request.GET.get('tipo_movimiento', '')
        
        # Fechas por defecto: último mes
        if not fecha_desde or not fecha_hasta:
            fecha_hasta_dt = datetime.now()
            fecha_desde_dt = fecha_hasta_dt - timedelta(days=30)
            fecha_desde = fecha_desde_dt.strftime('%Y-%m-%d')
            fecha_hasta = fecha_hasta_dt.strftime('%Y-%m-%d')
        
        # Query de ventas
        ventas_query = Ventas.objects.filter(
            id_hijo__id_cliente_responsable=cliente
        ).select_related(
            'id_hijo', 
            'id_empleado_cajero'
        ).prefetch_related(
            'detalles',
            'detalles__id_producto'
        )
        
        # Query de recargas
        recargas_query = CargasSaldo.objects.filter(
            nro_tarjeta__id_hijo__id_cliente_responsable=cliente
        ).select_related(
            'nro_tarjeta',
            'nro_tarjeta__id_hijo',
            'id_empleado'
        )
        
        # Aplicar filtros de fecha
        if fecha_desde:
            ventas_query = ventas_query.filter(fecha__date__gte=fecha_desde)
            recargas_query = recargas_query.filter(fecha_carga__date__gte=fecha_desde)
        if fecha_hasta:
            ventas_query = ventas_query.filter(fecha__date__lte=fecha_hasta)
            recargas_query = recargas_query.filter(fecha_carga__date__lte=fecha_hasta)
        
        # Obtener datos
        ventas = ventas_query.order_by('-fecha')
        recargas = recargas_query.order_by('-fecha_carga')
        
        # Construir lista unificada de movimientos
        movimientos = []
        saldo_acumulado = Decimal('0')
        
        # Agregar recargas (ABONOS - aumentan saldo)
        for recarga in recargas:
            if tipo_movimiento and tipo_movimiento != 'recarga':
                continue
            saldo_acumulado += recarga.monto_cargado
            movimientos.append({
                'fecha': recarga.fecha_carga,
                'tipo': 'ABONO',
                'descripcion': f'Recarga de saldo - Tarjeta {recarga.nro_tarjeta.nro_tarjeta}',
                'monto': recarga.monto_cargado,
                'saldo_acumulado': saldo_acumulado,
                'estudiante': f"{recarga.nro_tarjeta.id_hijo.nombre} {recarga.nro_tarjeta.id_hijo.apellido}" if recarga.nro_tarjeta.id_hijo else None,
                'empleado': f"{recarga.id_empleado.nombre} {recarga.id_empleado.apellido}" if recarga.id_empleado else None,
                'referencia': f'Recarga #{recarga.id_carga}',
                'items': None
            })
        
        # Agregar ventas (CARGOS - disminuyen saldo)
        for venta in ventas:
            if tipo_movimiento and tipo_movimiento != 'venta':
                continue
            saldo_acumulado -= venta.monto_total
            
            # Obtener items de la venta
            items = []
            for detalle in venta.detalles.all():
                items.append({
                    'producto': detalle.id_producto.descripcion,
                    'cantidad': detalle.cantidad,
                    'precio_unitario': detalle.precio_unitario,
                    'subtotal': detalle.subtotal_total
                })
            
            movimientos.append({
                'fecha': venta.fecha,
                'tipo': 'CARGO',
                'descripcion': f'Venta #{venta.id_venta}',
                'monto': venta.monto_total,
                'saldo_acumulado': saldo_acumulado,
                'estudiante': f"{venta.id_hijo.nombre} {venta.id_hijo.apellido}" if venta.id_hijo else None,
                'empleado': f"{venta.id_empleado_cajero.nombre} {venta.id_empleado_cajero.apellido}" if venta.id_empleado_cajero else None,
                'referencia': f'Venta #{venta.id_venta}',
                'items': items
            })
        
        # Ordenar por fecha descendente
        movimientos.sort(key=lambda x: x['fecha'], reverse=True)
        
        # Recalcular saldo acumulado en orden correcto
        saldo = Decimal('0')
        for mov in reversed(movimientos):
            if mov['tipo'] == 'CARGO':
                saldo -= mov['monto']
            else:
                saldo += mov['monto']
            mov['saldo_acumulado'] = saldo
        
        # Calcular totales
        total_ventas = ventas.count()
        monto_ventas = ventas.aggregate(total=Sum('monto_total'))['total'] or Decimal('0')
        total_recargas = recargas.count()
        monto_recargas = recargas.aggregate(total=Sum('monto_cargado'))['total'] or Decimal('0')
        
        total_cargos = sum(m['monto'] for m in movimientos if m['tipo'] == 'CARGO')
        total_abonos = sum(m['monto'] for m in movimientos if m['tipo'] == 'ABONO')
        saldo_actual = total_abonos - total_cargos
        
        # Crédito disponible
        credito_disponible = Decimal('0')
        if cliente.limite_credito and cliente.limite_credito > 0:
            credito_disponible = cliente.limite_credito + saldo_actual
            if credito_disponible < 0:
                credito_disponible = Decimal('0')
        
        # Preparar datos para gráfico (JSON)
        movimientos_json = json.dumps([{
            'fecha': m['fecha'].strftime('%d/%m/%Y'),
            'saldo': float(m['saldo_acumulado'])
        } for m in reversed(movimientos)])
        
        context = {
            'cliente': cliente,
            'hijos': hijos,
            'movimientos': movimientos,
            'movimientos_json': movimientos_json,
            'total_ventas': total_ventas,
            'monto_ventas': monto_ventas,
            'total_recargas': total_recargas,
            'monto_recargas': monto_recargas,
            'total_cargos': total_cargos,
            'total_abonos': total_abonos,
            'saldo_actual': saldo_actual,
            'credito_disponible': credito_disponible,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'tipo_movimiento': tipo_movimiento,
        }
        
        return render(request, 'pos/cuenta_corriente_unificada.html', context)
        
    except Cliente.DoesNotExist:
        return JsonResponse({'error': 'Cliente no encontrado'}, status=404)


@login_required
@require_http_methods(["POST"])
def cc_registrar_pago(request):
    """Registrar una recarga como 'pago' del responsable"""
    try:
        data = json.loads(request.body)
        cliente_id = data.get('cliente_id')
        tarjeta_id = data.get('tarjeta_id')  # Tarjeta del hijo donde se aplicará
        monto = Decimal(data.get('monto', 0))
        forma_pago = data.get('forma_pago', 'efectivo')
        observaciones = data.get('observaciones', 'Pago del responsable')
        
        # Validaciones
        if monto <= 0:
            return JsonResponse({
                'success': False,
                'error': 'El monto debe ser mayor a 0'
            }, status=400)
        
        # Obtener cliente y tarjeta
        cliente = Cliente.objects.get(id_cliente=cliente_id)
        tarjeta = Tarjeta.objects.get(nro_tarjeta=tarjeta_id)
        
        # Verificar que la tarjeta pertenece a un hijo del cliente
        if not Hijo.objects.filter(id_cliente_responsable=cliente, nro_tarjeta=tarjeta).exists():
            return JsonResponse({
                'success': False,
                'error': 'La tarjeta no pertenece a este cliente'
            }, status=400)
        
        # Obtener empleado actual
        empleado = Empleado.objects.filter(activo=True).first()
        
        # Crear recarga
        saldo_anterior = tarjeta.saldo_actual
        saldo_posterior = saldo_anterior + monto
        
        recarga = CargasSaldo.objects.create(
            nro_tarjeta=tarjeta,
            fecha=timezone.now(),
            monto=monto,
            forma_pago=forma_pago,
            saldo_anterior=saldo_anterior,
            saldo_posterior=saldo_posterior,
            observaciones=observaciones,
            id_empleado=empleado
        )
        
        # Actualizar saldo de tarjeta
        tarjeta.saldo_actual = F('saldo_actual') + monto
        tarjeta.save()
        tarjeta.refresh_from_db()
        
        return JsonResponse({
            'success': True,
            'recarga_id': recarga.id_carga_saldo,
            'monto': float(monto),
            'nuevo_saldo': float(tarjeta.saldo_actual),
            'mensaje': f'Pago registrado. Nuevo saldo: Gs. {tarjeta.saldo_actual:,.0f}'
        })
        
    except (Cliente.DoesNotExist, Tarjeta.DoesNotExist):
        return JsonResponse({
            'success': False,
            'error': 'Cliente o tarjeta no encontrados'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def cc_estado_cuenta(request, cliente_id):
    """Vista de estado de cuenta imprimible"""
    try:
        cliente = Cliente.objects.get(id_cliente=cliente_id)
        
        # Obtener fechas del filtro
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        
        # Obtener movimientos (ventas y pagos)
        ventas = Ventas.objects.filter(
            id_hijo__id_cliente_responsable=cliente
        ).select_related('id_hijo', 'id_empleado_cajero').order_by('-fecha')
        
        recargas = CargasSaldo.objects.filter(
            nro_tarjeta__id_hijo__id_cliente_responsable=cliente
        ).select_related('nro_tarjeta').order_by('-fecha_carga')
        
        # Aplicar filtros de fecha si existen
        if fecha_desde:
            ventas = ventas.filter(fecha__gte=fecha_desde)
            recargas = recargas.filter(fecha__gte=fecha_desde)
        
        if fecha_hasta:
            ventas = ventas.filter(fecha__lte=fecha_hasta)
            recargas = recargas.filter(fecha__lte=fecha_hasta)
        
        # Combinar y ordenar movimientos
        movimientos = []
        
        for venta in ventas:
            movimientos.append({
                'fecha': venta.fecha,
                'hora': venta.hora,
                'tipo': 'Venta',
                'descripcion': f'Venta #{venta.id_venta}',
                'cargo': venta.total,
                'abono': 0,
                'empleado': venta.id_empleado.nombre if venta.id_empleado else 'N/A'
            })
        
        for recarga in recargas:
            movimientos.append({
                'fecha': recarga.fecha,
                'tipo': 'Pago/Recarga',
                'descripcion': f'Recarga tarjeta {recarga.nro_tarjeta.nro_tarjeta}',
                'cargo': 0,
                'abono': recarga.monto,
                'empleado': recarga.id_empleado.nombre if recarga.id_empleado else 'N/A'
            })
        
        # Ordenar por fecha
        movimientos.sort(key=lambda x: x['fecha'], reverse=True)
        
        # Calcular totales
        total_cargos = sum(m['cargo'] for m in movimientos)
        total_abonos = sum(m['abono'] for m in movimientos)
        
        context = {
            'cliente': cliente,
            'movimientos': movimientos,
            'total_cargos': total_cargos,
            'total_abonos': total_abonos,
            'saldo': total_cargos - total_abonos,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
        }
        
        return render(request, 'pos/cc_estado_cuenta.html', context)
        
    except Cliente.DoesNotExist:
        return JsonResponse({'error': 'Cliente no encontrado'}, status=404)


# ==================== PROVEEDORES ====================

@login_required
def proveedores_view(request):
    """Vista principal de gestión de proveedores"""
    # Obtener filtros
    buscar = request.GET.get('buscar', '').strip()
    estado_filter = request.GET.get('estado', '')
    
    # Query base
    proveedores = Proveedor.objects.all()
    
    # Aplicar filtros
    if buscar:
        proveedores = proveedores.filter(
            Q(razon_social__icontains=buscar) |
            Q(ruc__icontains=buscar)
        )
    
    if estado_filter:
        proveedores = proveedores.filter(activo=(estado_filter == 'activo'))
    
    # Estadísticas
    from django.db.models import Count as CountFunc
    stats = {
        'total_proveedores': proveedores.count(),
        'proveedores_activos': proveedores.filter(activo=True).count(),
    }
    
    context = {
        'proveedores': proveedores.order_by('-fecha_registro'),
        'stats': stats,
        'buscar': buscar,
        'estado': estado_filter,
    }
    
    return render(request, 'pos/proveedores.html', context)


@login_required
def proveedor_detalle_view(request, proveedor_id):
    """Vista de detalle de un proveedor"""
    try:
        proveedor = Proveedor.objects.get(id_proveedor=proveedor_id)
        
        context = {
            'proveedor': proveedor,
        }
        
        return render(request, 'pos/proveedor_detalle.html', context)
        
    except Proveedor.DoesNotExist:
        return JsonResponse({'error': 'Proveedor no encontrado'}, status=404)


@login_required
@require_http_methods(["POST"])
def proveedor_crear(request):
    """Crear nuevo proveedor"""
    try:
        data = json.loads(request.body)
        
        # Validar datos requeridos
        required_fields = ['ruc', 'razon_social']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({
                    'success': False,
                    'error': f'El campo {field} es requerido'
                }, status=400)
        
        # Verificar que el RUC no exista
        if Proveedor.objects.filter(ruc=data['ruc']).exists():
            return JsonResponse({
                'success': False,
                'error': 'Ya existe un proveedor con ese RUC'
            }, status=400)
        
        # Crear proveedor
        proveedor = Proveedor.objects.create(
            ruc=data['ruc'],
            razon_social=data['razon_social'],
            telefono=data.get('telefono', ''),
            email=data.get('email', ''),
            direccion=data.get('direccion', ''),
            ciudad=data.get('ciudad', ''),
            activo=data.get('activo', True)
        )
        
        return JsonResponse({
            'success': True,
            'proveedor_id': proveedor.id_proveedor,
            'mensaje': f'Proveedor {proveedor.razon_social} creado exitosamente'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def proveedor_editar(request, proveedor_id):
    """Editar proveedor existente"""
    try:
        proveedor = Proveedor.objects.get(id_proveedor=proveedor_id)
        data = json.loads(request.body)
        
        # Actualizar campos
        if 'razon_social' in data:
            proveedor.razon_social = data['razon_social']
        if 'telefono' in data:
            proveedor.telefono = data['telefono']
        if 'email' in data:
            proveedor.email = data['email']
        if 'direccion' in data:
            proveedor.direccion = data['direccion']
        if 'ciudad' in data:
            proveedor.ciudad = data['ciudad']
        if 'activo' in data:
            proveedor.activo = data['activo']
        
        proveedor.save()
        
        return JsonResponse({
            'success': True,
            'mensaje': 'Proveedor actualizado exitosamente'
        })
        
    except Proveedor.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Proveedor no encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def proveedor_eliminar(request, proveedor_id):
    """Desactivar proveedor (soft delete)"""
    try:
        proveedor = Proveedor.objects.get(id_proveedor=proveedor_id)
        proveedor.activo = False
        proveedor.save()
        
        return JsonResponse({
            'success': True,
            'mensaje': f'Proveedor {proveedor.razon_social} desactivado'
        })
        
    except Proveedor.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Proveedor no encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ==================== INVENTARIO AVANZADO ====================

@login_required
def inventario_dashboard(request):
    """Dashboard principal de inventario con alertas y estadísticas"""
    # Productos con stock bajo (menor al mínimo)
    productos_stock_bajo = Producto.objects.filter(
        activo=True,
        stock_minimo__isnull=False
    ).annotate(
        stock_actual_val=F('stock__stock_actual')
    ).filter(
        stock_actual_val__lt=F('stock_minimo')
    ).select_related('id_categoria', 'stock')[:20]
    
    # Productos sin stock
    productos_sin_stock = Producto.objects.filter(
        activo=True,
        stock__stock_actual__lte=0
    ).select_related('id_categoria', 'stock')[:20]
    
    # Estadísticas generales
    from django.db.models import Count as CountFunc, Avg
    stats = Producto.objects.filter(activo=True).aggregate(
        total_productos=CountFunc('id_producto'),
        stock_promedio=Avg('stock__stock_actual')
    )
    
    # Productos más vendidos (últimos 30 días)
    from datetime import timedelta
    fecha_limite = timezone.now() - timedelta(days=30)
    
    productos_mas_vendidos = DetalleVenta.objects.filter(
        id_venta__fecha__gte=fecha_limite
    ).values(
        'id_producto__descripcion',
        'id_producto__codigo_barra'
    ).annotate(
        total_vendido=Sum('cantidad')
    ).order_by('-total_vendido')[:10]
    
    # Categorías con más productos
    categorias_stock = Categoria.objects.annotate(
        total_productos=CountFunc('productos', filter=Q(productos__activo=True)),
        stock_total=Sum('productos__stock__stock_actual')
    ).filter(total_productos__gt=0).order_by('-stock_total')[:10]
    
    context = {
        'productos_stock_bajo': productos_stock_bajo,
        'productos_sin_stock': productos_sin_stock,
        'stats': stats,
        'productos_mas_vendidos': productos_mas_vendidos,
        'categorias_stock': categorias_stock,
        'alertas_count': productos_stock_bajo.count() + productos_sin_stock.count(),
    }
    
    return render(request, 'pos/inventario_dashboard.html', context)


@login_required
def inventario_productos(request):
    """Vista de listado de productos con stock"""
    # Obtener filtros
    buscar = request.GET.get('buscar', '').strip()
    categoria_id = request.GET.get('categoria', '')
    estado_stock = request.GET.get('estado_stock', '')
    
    # Query base
    productos = Producto.objects.filter(activo=True).select_related(
        'id_categoria', 'id_unidad_de_medida', 'stock'
    ).annotate(
        stock_actual_val=F('stock__stock_actual')
    )
    
    # Aplicar filtros
    if buscar:
        productos = productos.filter(
            Q(descripcion__icontains=buscar) |
            Q(codigo__icontains=buscar)
        )
    
    if categoria_id:
        productos = productos.filter(id_categoria_id=categoria_id)
    
    if estado_stock == 'bajo':
        productos = productos.filter(
            stock_minimo__isnull=False,
            stock_actual_val__lt=F('stock_minimo')
        )
    elif estado_stock == 'sin_stock':
        productos = productos.filter(stock_actual_val__lte=0)
    elif estado_stock == 'normal':
        productos = productos.filter(
            Q(stock_minimo__isnull=True) |
            Q(stock_actual_val__gte=F('stock_minimo'))
        ).exclude(stock_actual_val__lte=0)
    
    # Categorías para filtro
    categorias = Categoria.objects.all()
    
    context = {
        'productos': productos.order_by('descripcion'),
        'categorias': categorias,
        'buscar': buscar,
        'categoria_id': categoria_id,
        'estado_stock': estado_stock,
    }
    
    return render(request, 'pos/inventario_productos.html', context)


@login_required
def kardex_producto(request, producto_id):
    """Kardex completo de un producto (historial de movimientos)"""
    try:
        producto = Producto.objects.select_related(
            'id_categoria', 'id_unidad_de_medida', 'stock'
        ).get(id_producto=producto_id)
        
        # Obtener fechas del filtro
        fecha_desde = request.GET.get('fecha_desde')
        fecha_hasta = request.GET.get('fecha_hasta')
        
        # Movimientos de salida (ventas)
        ventas = DetalleVenta.objects.filter(
            id_producto=producto
        ).select_related('id_venta', 'id_venta__id_empleado')
        
        if fecha_desde:
            ventas = ventas.filter(id_venta__fecha__gte=fecha_desde)
        if fecha_hasta:
            ventas = ventas.filter(id_venta__fecha__lte=fecha_hasta)
        
        ventas = ventas.order_by('-id_venta__fecha', '-id_venta__hora')[:100]
        
        # Construir kardex
        movimientos = []
        for venta_detalle in ventas:
            movimientos.append({
                'fecha': venta_detalle.id_venta.fecha,
                'hora': venta_detalle.id_venta.hora,
                'tipo': 'Salida',
                'descripcion': f'Venta #{venta_detalle.id_venta.id_venta}',
                'cantidad': venta_detalle.cantidad,
                'tipo_movimiento': 'venta',
                'empleado': venta_detalle.id_venta.id_empleado.nombre if venta_detalle.id_venta.id_empleado else 'N/A'
            })
        
        # Ordenar por fecha descendente
        movimientos.sort(key=lambda x: (x['fecha'], x.get('hora', '')), reverse=True)
        
        # Calcular totales
        total_entradas = 0
        total_salidas = sum(m['cantidad'] for m in movimientos if m['tipo'] == 'Salida')
        
        context = {
            'producto': producto,
            'movimientos': movimientos,
            'total_entradas': total_entradas,
            'total_salidas': total_salidas,
            'saldo_actual': producto.stock.stock_actual if hasattr(producto, 'stock') else 0,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
        }
        
        return render(request, 'pos/kardex_producto.html', context)
        
    except Producto.DoesNotExist:
        return JsonResponse({'error': 'Producto no encontrado'}, status=404)


@login_required
@require_http_methods(["GET", "POST"])
def ajuste_inventario_view(request):
    """Vista para realizar ajustes de inventario"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            producto_id = data.get('producto_id')
            tipo_ajuste = data.get('tipo_ajuste')  # 'suma' o 'resta'
            cantidad = Decimal(data.get('cantidad', 0))
            motivo = data.get('motivo', '')
            
            # Validaciones
            if cantidad <= 0:
                return JsonResponse({
                    'success': False,
                    'error': 'La cantidad debe ser mayor a 0'
                }, status=400)
            
            # Obtener producto y stock
            producto = Producto.objects.get(id_producto=producto_id)
            stock = StockUnico.objects.get(id_producto=producto)
            
            stock_anterior = stock.stock_actual
            
            # Aplicar ajuste
            if tipo_ajuste == 'suma':
                stock.stock_actual = F('stock_actual') + cantidad
                descripcion_movimiento = f'Ajuste de inventario: +{cantidad} {producto.id_unidad_de_medida.abreviatura}'
            elif tipo_ajuste == 'resta':
                stock.stock_actual = F('stock_actual') - cantidad
                descripcion_movimiento = f'Ajuste de inventario: -{cantidad} {producto.id_unidad_de_medida.abreviatura}'
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Tipo de ajuste inválido'
                }, status=400)
            
            stock.save()
            stock.refresh_from_db()
            
            # Aquí podrías registrar el ajuste en una tabla de auditoría
            # AjusteInventario.objects.create(...)
            
            return JsonResponse({
                'success': True,
                'stock_anterior': float(stock_anterior),
                'cantidad_ajuste': float(cantidad),
                'stock_nuevo': float(stock.stock_actual),
                'mensaje': f'Ajuste realizado. Nuevo stock: {stock.stock_actual} {producto.id_unidad_de_medida.abreviatura}'
            })
            
        except Producto.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Producto no encontrado'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    # GET - Mostrar formulario
    productos = Producto.objects.filter(activo=True).select_related(
        'id_categoria', 'id_unidad_de_medida', 'stock'
    ).order_by('descripcion')
    
    context = {
        'productos': productos,
    }
    
    return render(request, 'pos/ajuste_inventario.html', context)


@login_required
def alertas_inventario(request):
    """Vista de alertas de inventario"""
    # Productos con stock bajo
    productos_stock_bajo = Producto.objects.filter(
        activo=True,
        stock_minimo__isnull=False
    ).annotate(
        stock_actual_val=F('stock__stock_actual')
    ).filter(
        stock_actual_val__lt=F('stock_minimo')
    ).select_related('id_categoria', 'stock', 'id_unidad_de_medida')
    
    # Productos sin stock
    productos_sin_stock = Producto.objects.filter(
        activo=True,
        stock__stock_actual__lte=0
    ).select_related('id_categoria', 'stock', 'id_unidad_de_medida')
    
    # Productos críticos (menos del 50% del stock mínimo)
    productos_criticos = Producto.objects.filter(
        activo=True,
        stock_minimo__isnull=False
    ).annotate(
        stock_actual_val=F('stock__stock_actual')
    ).filter(
        stock_actual_val__lt=F('stock_minimo') * 0.5
    ).select_related('id_categoria', 'stock', 'id_unidad_de_medida')
    
    context = {
        'productos_stock_bajo': productos_stock_bajo,
        'productos_sin_stock': productos_sin_stock,
        'productos_criticos': productos_criticos,
        'total_alertas': (
            productos_stock_bajo.count() + 
            productos_sin_stock.count() + 
            productos_criticos.count()
        ),
    }
    
    return render(request, 'pos/alertas_inventario.html', context)


@login_required
@require_http_methods(["POST"])
def actualizar_stock_masivo(request):
    """Actualización masiva de stock (para inventario físico)"""
    try:
        data = json.loads(request.body)
        ajustes = data.get('ajustes', [])  # Lista de {producto_id, nuevo_stock}
        
        if not ajustes:
            return JsonResponse({
                'success': False,
                'error': 'No se proporcionaron ajustes'
            }, status=400)
        
        actualizados = 0
        errores = []
        
        for ajuste in ajustes:
            try:
                producto_id = ajuste.get('producto_id')
                nuevo_stock = Decimal(ajuste.get('nuevo_stock', 0))
                
                stock = StockUnico.objects.get(id_producto_id=producto_id)
                stock.stock_actual = nuevo_stock
                stock.save()
                actualizados += 1
                
            except Exception as e:
                errores.append(f'Producto {producto_id}: {str(e)}')
        
        return JsonResponse({
            'success': True,
            'actualizados': actualizados,
            'errores': errores,
            'mensaje': f'{actualizados} productos actualizados'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ==================== SISTEMA DE ALERTAS ====================

@login_required
def alertas_sistema_view(request):
    """Dashboard de alertas del sistema"""
    from datetime import timedelta
    
    # 1. ALERTAS DE SALDO BAJO EN TARJETAS (≤ 10,000 Gs)
    SALDO_MINIMO = 10000
    tarjetas_saldo_bajo = Tarjeta.objects.filter(
        estado='Activa',
        saldo_actual__lte=SALDO_MINIMO
    ).select_related('id_hijo', 'id_hijo__id_cliente_responsable').order_by('saldo_actual')
    
    # 2. ALERTAS DE STOCK BAJO
    productos_stock_bajo = Producto.objects.filter(
        activo=True,
        stock_minimo__isnull=False
    ).annotate(
        stock_actual_val=F('stock__stock_actual')
    ).filter(
        stock_actual_val__lt=F('stock_minimo')
    ).select_related('id_categoria', 'stock')[:20]
    
    # 3. ALERTAS DE PRODUCTOS SIN STOCK
    productos_sin_stock = Producto.objects.filter(
        activo=True,
        stock__stock_actual__lte=0
    ).select_related('id_categoria', 'stock')[:20]
    
    # 4. ALERTAS DE TARJETAS POR VENCER (próximos 30 días)
    fecha_limite = timezone.now().date() + timedelta(days=30)
    tarjetas_por_vencer = Tarjeta.objects.filter(
        estado='Activa',
        fecha_vencimiento__lte=fecha_limite,
        fecha_vencimiento__gte=timezone.now().date()
    ).select_related('id_hijo', 'id_hijo__id_cliente_responsable').order_by('fecha_vencimiento')
    
    # 5. ALERTAS DE TARJETAS BLOQUEADAS
    tarjetas_bloqueadas = Tarjeta.objects.filter(
        estado='Bloqueada'
    ).select_related('id_hijo', 'id_hijo__id_cliente_responsable').order_by('-fecha_creacion')[:20]
    
    # Estadísticas de alertas
    total_alertas = (
        tarjetas_saldo_bajo.count() +
        productos_stock_bajo.count() +
        productos_sin_stock.count() +
        tarjetas_por_vencer.count() +
        tarjetas_bloqueadas.count()
    )
    
    context = {
        'tarjetas_saldo_bajo': tarjetas_saldo_bajo,
        'saldo_minimo': SALDO_MINIMO,
        'productos_stock_bajo': productos_stock_bajo,
        'productos_sin_stock': productos_sin_stock,
        'tarjetas_por_vencer': tarjetas_por_vencer,
        'tarjetas_bloqueadas': tarjetas_bloqueadas,
        'total_alertas': total_alertas,
        'alertas_criticas': tarjetas_saldo_bajo.filter(saldo_actual__lte=5000).count(),
    }
    
    return render(request, 'pos/alertas_sistema.html', context)


@login_required
def alertas_tarjetas_saldo_view(request):
    """Vista específica para alertas de saldo bajo en tarjetas"""
    # Obtener parámetros de filtro
    saldo_max = request.GET.get('saldo_max', 10000)
    buscar = request.GET.get('buscar', '').strip()
    
    try:
        saldo_max = int(saldo_max)
    except ValueError:
        saldo_max = 10000
    
    # Query base
    tarjetas = Tarjeta.objects.filter(
        estado='Activa',
        saldo_actual__lte=saldo_max
    ).select_related(
        'id_hijo',
        'id_hijo__id_cliente_responsable'
    )
    
    # Aplicar búsqueda
    if buscar:
        tarjetas = tarjetas.filter(
            Q(nro_tarjeta__icontains=buscar) |
            Q(id_hijo__nombre__icontains=buscar) |
            Q(id_hijo__apellido__icontains=buscar) |
            Q(id_hijo__id_cliente_responsable__nombres__icontains=buscar) |
            Q(id_hijo__id_cliente_responsable__apellidos__icontains=buscar)
        )
    
    tarjetas = tarjetas.order_by('saldo_actual')
    
    # Estadísticas
    total_tarjetas = tarjetas.count()
    criticas = tarjetas.filter(saldo_actual__lte=5000).count()
    sin_saldo = tarjetas.filter(saldo_actual__lte=0).count()
    
    context = {
        'tarjetas': tarjetas,
        'saldo_max': saldo_max,
        'buscar': buscar,
        'total_tarjetas': total_tarjetas,
        'criticas': criticas,
        'sin_saldo': sin_saldo,
    }
    
    return render(request, 'pos/alertas_tarjetas_saldo.html', context)


@login_required
@require_http_methods(["POST"])
def marcar_alerta_vista(request):
    """Marcar una alerta como vista (para futuro con tabla de alertas)"""
    try:
        data = json.loads(request.body)
        alerta_id = data.get('alerta_id')
        tipo = data.get('tipo')  # 'tarjeta', 'producto', etc.
        
        # Por ahora solo retornamos success
        # En el futuro se puede crear tabla de alertas_historial
        
        return JsonResponse({
            'success': True,
            'mensaje': 'Alerta marcada como vista'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def enviar_notificacion_saldo(request, tarjeta_id):
    """Enviar notificación de saldo bajo al responsable"""
    try:
        from gestion.notificaciones import notificar_saldo_bajo
        
        tarjeta = Tarjeta.objects.select_related(
            'id_hijo',
            'id_hijo__id_cliente_responsable'
        ).get(nro_tarjeta=tarjeta_id)
        
        responsable = tarjeta.id_hijo.id_cliente_responsable
        
        # Determinar canales disponibles
        canales = []
        if responsable.email:
            canales.append('email')
        if responsable.telefono:
            canales.append('sms')  # También se puede agregar 'whatsapp'
        
        if not canales:
            return JsonResponse({
                'success': False,
                'error': 'El cliente no tiene email ni teléfono configurado'
            }, status=400)
        
        # Enviar notificación por todos los canales disponibles
        resultados = notificar_saldo_bajo(tarjeta, canales)
        
        # Verificar si al menos un canal fue exitoso
        exito = any(resultados.values())
        
        if exito:
            canales_exitosos = [canal for canal, result in resultados.items() if result]
            return JsonResponse({
                'success': True,
                'mensaje': f'Notificación enviada por: {", ".join(canales_exitosos)}',
                'resultados': resultados
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'No se pudo enviar la notificación por ningún canal',
                'resultados': resultados
            }, status=500)
        
    except Tarjeta.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Tarjeta no encontrada'
        }, status=404)
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'error': str(e),
            'trace': traceback.format_exc()
        }, status=500)


# ==================== SISTEMA DE CAJAS ====================

@login_required
def cajas_dashboard_view(request):
    """Dashboard principal de cajas"""
    # Obtener todas las cajas
    cajas = Cajas.objects.filter(activo=True)
    
    # Obtener caja abierta del usuario actual
    caja_abierta = None
    try:
        empleado = Empleado.objects.get(usuario=request.user.username)
        caja_abierta = CierresCaja.objects.filter(
            id_empleado=empleado,
            estado='Abierta'
        ).select_related('id_caja').first()
    except Empleado.DoesNotExist:
        pass
    
    # Estadísticas del día
    hoy = timezone.now().date()
    cierres_hoy = CierresCaja.objects.filter(
        fecha_hora_apertura__date=hoy
    ).select_related('id_caja', 'id_empleado')
    
    total_ventas_hoy = Ventas.objects.filter(
        fecha__date=hoy,
        estado='Completada'
    ).aggregate(total=Sum('monto_total'))['total'] or 0
    
    context = {
        'cajas': cajas,
        'caja_abierta': caja_abierta,
        'cierres_hoy': cierres_hoy,
        'total_ventas_hoy': total_ventas_hoy,
    }
    return render(request, 'pos/cajas_dashboard.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def apertura_caja_view(request):
    """Apertura de caja"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            id_caja = data.get('id_caja')
            monto_inicial = Decimal(data.get('monto_inicial', '0'))
            
            # Validar empleado
            try:
                empleado = Empleado.objects.get(usuario=request.user.username)
            except Empleado.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Usuario no asociado a un empleado'
                }, status=400)
            
            # Verificar que no tenga caja abierta
            caja_existente = CierresCaja.objects.filter(
                id_empleado=empleado,
                estado='Abierta'
            ).first()
            
            if caja_existente:
                return JsonResponse({
                    'success': False,
                    'error': f'Ya tiene la caja {caja_existente.id_caja.nombre_caja} abierta'
                }, status=400)
            
            # Validar caja
            caja = Cajas.objects.get(id_caja=id_caja, activo=True)
            
            # Crear apertura
            cierre = CierresCaja.objects.create(
                id_caja=caja,
                id_empleado=empleado,
                fecha_apertura=timezone.now(),
                monto_inicial=monto_inicial,
                estado='Abierta'
            )
            
            return JsonResponse({
                'success': True,
                'mensaje': f'Caja {caja.nombre_caja} abierta correctamente',
                'id_cierre': cierre.id_cierre_caja
            })
            
        except Cajas.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Caja no encontrada'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    # GET
    cajas_disponibles = Cajas.objects.filter(activo=True)
    return render(request, 'pos/apertura_caja.html', {
        'cajas': cajas_disponibles
    })


@login_required
@require_http_methods(["GET", "POST"])
def cierre_caja_view(request):
    """Cierre de caja"""
    try:
        empleado = Empleado.objects.get(usuario=request.user.username)
        caja_abierta = CierresCaja.objects.filter(
            id_empleado=empleado,
            estado='Abierta'
        ).select_related('id_caja').first()
        
        if not caja_abierta:
            return render(request, 'pos/cierre_caja.html', {
                'error': 'No tiene ninguna caja abierta'
            })
        
        if request.method == 'POST':
            try:
                data = json.loads(request.body)
                monto_final = Decimal(data.get('monto_final', '0'))
                observaciones = data.get('observaciones', '')
                
                # Calcular ventas del turno
                ventas_turno = Ventas.objects.filter(
                    fecha_venta__gte=caja_abierta.fecha_apertura,
                    estado='Completada'
                ).aggregate(total=Sum('monto_total'))['total'] or 0
                
                # Calcular diferencia
                monto_esperado = caja_abierta.monto_inicial + ventas_turno
                diferencia = monto_final - monto_esperado
                
                # Actualizar cierre
                caja_abierta.fecha_cierre = timezone.now()
                caja_abierta.monto_final = monto_final
                caja_abierta.diferencia = diferencia
                caja_abierta.observaciones = observaciones
                caja_abierta.estado = 'Cerrada'
                caja_abierta.save()
                
                return JsonResponse({
                    'success': True,
                    'mensaje': 'Caja cerrada correctamente',
                    'diferencia': float(diferencia),
                    'monto_esperado': float(monto_esperado)
                })
                
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                }, status=500)
        
        # GET - Calcular totales para mostrar
        ventas_turno = Ventas.objects.filter(
            fecha_venta__gte=caja_abierta.fecha_apertura,
            estado='Completada'
        )
        
        total_ventas = ventas_turno.aggregate(total=Sum('monto_total'))['total'] or 0
        cantidad_ventas = ventas_turno.count()
        monto_esperado = caja_abierta.monto_inicial + total_ventas
        
        context = {
            'caja_abierta': caja_abierta,
            'total_ventas': total_ventas,
            'cantidad_ventas': cantidad_ventas,
            'monto_esperado': monto_esperado,
        }
        
        return render(request, 'pos/cierre_caja.html', context)
        
    except Empleado.DoesNotExist:
        return render(request, 'pos/cierre_caja.html', {
            'error': 'Usuario no asociado a un empleado'
        })


@login_required
def arqueo_caja_view(request):
    """Arqueo de caja - conteo de efectivo"""
    try:
        empleado = Empleado.objects.get(usuario=request.user.username)
        caja_abierta = CierresCaja.objects.filter(
            id_empleado=empleado,
            estado='Abierta'
        ).select_related('id_caja').first()
        
        if not caja_abierta:
            return render(request, 'pos/arqueo_caja.html', {
                'error': 'No tiene ninguna caja abierta'
            })
        
        # Denominaciones de billetes y monedas (Paraguay)
        denominaciones = [
            {'valor': 100000, 'tipo': 'billete', 'nombre': '100.000 Gs'},
            {'valor': 50000, 'tipo': 'billete', 'nombre': '50.000 Gs'},
            {'valor': 20000, 'tipo': 'billete', 'nombre': '20.000 Gs'},
            {'valor': 10000, 'tipo': 'billete', 'nombre': '10.000 Gs'},
            {'valor': 5000, 'tipo': 'billete', 'nombre': '5.000 Gs'},
            {'valor': 2000, 'tipo': 'billete', 'nombre': '2.000 Gs'},
            {'valor': 1000, 'tipo': 'moneda', 'nombre': '1.000 Gs'},
            {'valor': 500, 'tipo': 'moneda', 'nombre': '500 Gs'},
            {'valor': 100, 'tipo': 'moneda', 'nombre': '100 Gs'},
            {'valor': 50, 'tipo': 'moneda', 'nombre': '50 Gs'},
        ]
        
        context = {
            'caja_abierta': caja_abierta,
            'denominaciones': denominaciones,
        }
        
        return render(request, 'pos/arqueo_caja.html', context)
        
    except Empleado.DoesNotExist:
        return render(request, 'pos/arqueo_caja.html', {
            'error': 'Usuario no asociado a un empleado'
        })


@login_required
def conciliacion_pagos_view(request):
    """Conciliación de pagos por medio de pago"""
    fecha = request.GET.get('fecha', timezone.now().date())
    
    # Obtener medios de pago
    medios_pago = MediosPago.objects.filter(activo=True)
    
    # Calcular totales por medio de pago
    resumen = []
    for medio in medios_pago:
        pagos = PagosVenta.objects.filter(
            id_venta__fecha__date=fecha,
            id_medio_pago=medio,
            id_venta__estado='Completada'
        )
        
        total = pagos.aggregate(total=Sum('monto_aplicado'))['total'] or 0
        cantidad = pagos.count()
        
        resumen.append({
            'medio': medio,
            'total': total,
            'cantidad': cantidad
        })
    
    # Total general
    total_general = sum(item['total'] for item in resumen)
    
    context = {
        'fecha': fecha,
        'resumen': resumen,
        'total_general': total_general,
    }
    
    return render(request, 'pos/conciliacion_pagos.html', context)


# ==================== SISTEMA DE COMPRAS ====================

@login_required
def compras_dashboard_view(request):
    """Dashboard de compras"""
    # Últimas compras
    compras_recientes = Compras.objects.select_related(
        'id_proveedor'
    ).order_by('-fecha')[:20]
    
    # Estadísticas
    hoy = timezone.now().date()
    mes_actual = hoy.replace(day=1)
    
    compras_mes = Compras.objects.filter(fecha__gte=mes_actual)
    total_mes = compras_mes.aggregate(total=Sum('monto_total'))['total'] or 0
    cantidad_mes = compras_mes.count()
    
    # Compras pendientes de pago (usar nuevo sistema)
    compras_pendientes = Compras.objects.filter(
        estado_pago__in=['PENDIENTE', 'PARCIAL']
    ).count()
    
    # Deuda total con proveedores (suma de saldos pendientes)
    deuda_total = Compras.objects.filter(
        estado_pago__in=['PENDIENTE', 'PARCIAL']
    ).aggregate(total=Sum('saldo_pendiente'))['total'] or 0
    
    context = {
        'compras_recientes': compras_recientes,
        'total_mes': total_mes,
        'cantidad_mes': cantidad_mes,
        'compras_pendientes': compras_pendientes,
        'deuda_total': deuda_total,
    }
    
    return render(request, 'pos/compras_dashboard.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def nueva_compra_view(request):
    """Registrar nueva compra"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Validar empleado
            try:
                empleado = Empleado.objects.get(usuario=request.user.username)
            except Empleado.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Usuario no asociado a un empleado'
                }, status=400)
            
            # Validar proveedor
            proveedor = Proveedor.objects.get(id_proveedor=data['id_proveedor'])
            
            # Crear compra
            compra = Compras.objects.create(
                id_proveedor=proveedor,
                fecha=timezone.now(),
                nro_factura=data.get('nro_factura', ''),
                monto_total=Decimal(data['total']),
                observaciones=data.get('observaciones', '')
            )
            
            # Crear detalle de compra
            total_calculado = Decimal('0')
            for item in data['items']:
                producto = Producto.objects.get(id_producto=item['id_producto'])
                cantidad = Decimal(item['cantidad'])
                precio_compra = Decimal(item['precio_compra'])
                subtotal_item = cantidad * precio_compra
                
                DetalleCompra.objects.create(
                    id_compra=compra,
                    id_producto=producto,
                    cantidad=cantidad,
                    precio_compra=precio_compra,
                    subtotal=subtotal_item
                )
                
                total_calculado += subtotal_item
            
            # El saldo pendiente y estado ya se establecen automáticamente en el modelo
            # compra.saldo_pendiente = compra.total (por defecto)
            # compra.estado_pago = 'PENDIENTE' (por defecto)
            
            return JsonResponse({
                'success': True,
                'mensaje': 'Compra registrada correctamente',
                'id_compra': compra.id_compra
            })
            
        except Proveedor.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Proveedor no encontrado'
            }, status=404)
        except Producto.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Producto no encontrado'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    # GET
    proveedores = Proveedor.objects.filter(activo=True).order_by('razon_social')
    productos = Producto.objects.filter(activo=True).order_by('descripcion')
    
    context = {
        'proveedores': proveedores,
        'productos': productos,
    }
    
    return render(request, 'pos/nueva_compra.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def recepcion_mercaderia_view(request, id_compra):
    """Recepción de mercadería - ingresa stock"""
    try:
        compra = Compras.objects.select_related('id_proveedor').get(id_compra=id_compra)
        
        if request.method == 'POST':
            try:
                data = json.loads(request.body)
                
                # Validar que la compra esté pendiente
                if compra.estado != 'Pendiente':
                    return JsonResponse({
                        'success': False,
                        'error': 'La compra ya fue recibida'
                    }, status=400)
                
                # Validar empleado
                try:
                    empleado = Empleado.objects.get(usuario=request.user.username)
                except Empleado.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': 'Usuario no asociado a un empleado'
                    }, status=400)
                
                # Procesar cada item recibido
                for item_data in data['items']:
                    detalle = DetalleCompra.objects.get(
                        id_detalle_compra=item_data['id_detalle']
                    )
                    
                    cantidad_recibida = Decimal(item_data['cantidad_recibida'])
                    
                    # Actualizar stock
                    stock, created = StockUnico.objects.get_or_create(
                        id_producto=detalle.id_producto,
                        defaults={'stock_actual': Decimal('0'), 'stock_minimo': Decimal('10')}
                    )
                    
                    stock_anterior = stock.stock_actual
                    stock.stock_actual += cantidad_recibida
                    stock.ultima_actualizacion = timezone.now()
                    stock.save()
                    
                    # Registrar movimiento
                    MovimientosStock.objects.create(
                        id_producto=detalle.id_producto,
                        tipo_movimiento='Entrada',
                        cantidad=cantidad_recibida,
                        stock_anterior=stock_anterior,
                        stock_nuevo=stock.stock_actual,
                        motivo=f'Compra #{compra.id_compra} - Factura {compra.nro_factura}',
                        id_empleado=empleado,
                        fecha_movimiento=timezone.now()
                    )
                
                # Actualizar estado de compra
                compra.estado = 'Recibida'
                compra.fecha_recepcion = timezone.now()
                compra.save()
                
                return JsonResponse({
                    'success': True,
                    'mensaje': 'Mercadería recibida e ingresada a stock'
                })
                
            except DetalleCompra.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Detalle de compra no encontrado'
                }, status=404)
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                }, status=500)
        
        # GET - Obtener detalles de la compra
        detalles = DetalleCompra.objects.filter(
            id_compra=compra
        ).select_related('id_producto')
        
        context = {
            'compra': compra,
            'detalles': detalles,
        }
        
        return render(request, 'pos/recepcion_mercaderia.html', context)
        
    except Compras.DoesNotExist:
        return render(request, 'pos/recepcion_mercaderia.html', {
            'error': 'Compra no encontrada'
        })


@login_required
def deuda_proveedores_view(request):
    """Reporte de deuda con proveedores"""
    # Obtener compras con saldo pendiente agrupadas por proveedor
    from django.db.models import Q
    deudas = Compras.objects.filter(
        Q(estado_pago='PENDIENTE') | Q(estado_pago='PARCIAL'),
        saldo_pendiente__gt=0
    ).values(
        'id_proveedor__id_proveedor',
        'id_proveedor__razon_social'
    ).annotate(
        saldo=Sum('saldo_pendiente'),
        cantidad_compras=Count('id_compra')
    ).order_by('-saldo')
    
    # Total de deuda
    total_deuda = deudas.aggregate(total=Sum('saldo'))['total'] or 0
    
    context = {
        'deudas': deudas,
        'total_deuda': total_deuda,
    }
    
    return render(request, 'pos/deuda_proveedores.html', context)


# ==================== SISTEMA DE COMISIONES ====================

@login_required
def comisiones_dashboard_view(request):
    """Dashboard de comisiones"""
    # Obtener tarifas activas
    tarifas = TarifasComision.objects.filter(activo=True).select_related('id_medio_pago')
    
    # Estadísticas del mes
    hoy = timezone.now().date()
    mes_actual = hoy.replace(day=1)
    
    # Comisiones del mes - corregido: usar id_pago_venta en lugar de id_venta
    comisiones_mes = DetalleComisionVenta.objects.filter(
        id_pago_venta__fecha_pago__gte=mes_actual
    ).aggregate(
        total_comision=Sum('monto_comision_calculada'),
        cantidad=Count('id_detalle_comision')
    )
    
    # Comisiones por medio de pago
    comisiones_por_medio = []
    medios_comision = MediosPago.objects.filter(
        genera_comision=True,
        activo=True
    )
    
    for medio in medios_comision:
        # Obtener comisiones a través de la relación correcta
        comisiones = DetalleComisionVenta.objects.filter(
            id_pago_venta__fecha_pago__gte=mes_actual,
            id_pago_venta__id_medio_pago=medio
        ).aggregate(
            total_comision=Sum('monto_comision_calculada'),
            cantidad=Count('id_detalle_comision')
        )
        
        if comisiones['total_comision'] and comisiones['total_comision'] > 0:
            comisiones_por_medio.append({
                'medio': medio,
                'total_comision': comisiones['total_comision'] or 0,
                'cantidad': comisiones['cantidad'] or 0
            })
    
    context = {
        'tarifas': tarifas,
        'total_comision_mes': comisiones_mes['total_comision'] or 0,
        'total_transacciones_mes': comisiones_mes['cantidad'] or 0,
        'comisiones_por_medio': comisiones_por_medio,
    }
    
    return render(request, 'pos/comisiones_dashboard.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def configurar_tarifas_view(request):
    """
    Configuración avanzada de tarifas de comisión
    Permite crear, editar y gestionar tarifas con validaciones
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Validar datos
            if not data.get('id_medio_pago'):
                return JsonResponse({
                    'success': False,
                    'error': 'Debe seleccionar un medio de pago'
                }, status=400)
            
            # Obtener medio de pago
            medio_pago = MediosPago.objects.get(id_medio_pago=data['id_medio_pago'])
            
            # Validar que el medio genere comisión
            if not medio_pago.genera_comision:
                return JsonResponse({
                    'success': False,
                    'error': f'{medio_pago.descripcion} no genera comisión'
                }, status=400)
            
            # Convertir porcentaje a decimal (de % a fracción)
            porcentaje_input = Decimal(str(data.get('porcentaje', '0')))
            porcentaje_decimal = porcentaje_input / 100  # 2.5% → 0.025
            
            # Validar rango de porcentaje
            if porcentaje_decimal < 0 or porcentaje_decimal > 1:
                return JsonResponse({
                    'success': False,
                    'error': 'El porcentaje debe estar entre 0% y 100%'
                }, status=400)
            
            monto_fijo = Decimal(str(data.get('monto_fijo', '0')))
            
            # Validar que al menos uno tenga valor
            if porcentaje_decimal == 0 and monto_fijo == 0:
                return JsonResponse({
                    'success': False,
                    'error': 'Debe especificar un porcentaje y/o monto fijo'
                }, status=400)
            
            # Desactivar tarifas anteriores del mismo medio
            TarifasComision.objects.filter(
                id_medio_pago=medio_pago,
                activo=True
            ).update(
                activo=False,
                fecha_fin_vigencia=timezone.now()
            )
            
            # Crear nueva tarifa
            tarifa = TarifasComision.objects.create(
                id_medio_pago=medio_pago,
                porcentaje_comision=porcentaje_decimal,
                monto_fijo_comision=monto_fijo if monto_fijo > 0 else None,
                fecha_inicio_vigencia=timezone.now(),
                fecha_fin_vigencia=None,
                activo=True
            )
            
            # Ejemplo de cálculo
            monto_ejemplo = 100000
            comision_ejemplo = float(monto_fijo) + (float(monto_ejemplo) * float(porcentaje_decimal))
            
            return JsonResponse({
                'success': True,
                'mensaje': f'Tarifa configurada exitosamente para {medio_pago.descripcion}',
                'tarifa': {
                    'id': tarifa.id_tarifa,
                    'medio': medio_pago.descripcion,
                    'porcentaje': f'{float(porcentaje_input):.2f}%',
                    'monto_fijo': f'Gs {float(monto_fijo):,.0f}' if monto_fijo > 0 else 'No',
                    'ejemplo': f'Gs {comision_ejemplo:,.0f}'
                }
            })
            
        except MediosPago.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Medio de pago no encontrado'
            }, status=404)
        except ValueError as e:
            return JsonResponse({
                'success': False,
                'error': f'Valor numérico inválido: {str(e)}'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error al guardar: {str(e)}'
            }, status=500)
    
    # GET - Mostrar formulario
    # Obtener medios de pago que generan comisión
    medios_comision = MediosPago.objects.filter(
        genera_comision=True,
        activo=True
    ).order_by('descripcion')
    
    # Obtener tarifas activas
    tarifas_existentes = TarifasComision.objects.filter(
        activo=True
    ).select_related('id_medio_pago').order_by('id_medio_pago__descripcion')
    
    # Formatear para el template
    tarifas_formateadas = []
    for tarifa in tarifas_existentes:
        porcentaje_display = float(tarifa.porcentaje_comision) * 100
        monto_fijo_display = float(tarifa.monto_fijo_comision) if tarifa.monto_fijo_comision else 0
        
        # Ejemplo de cálculo
        monto_ejemplo = 100000
        comision_ejemplo = monto_fijo_display + (monto_ejemplo * float(tarifa.porcentaje_comision))
        
        tarifas_formateadas.append({
            'id': tarifa.id_tarifa,
            'medio': tarifa.id_medio_pago.descripcion,
            'porcentaje': porcentaje_display,
            'monto_fijo': monto_fijo_display,
            'ejemplo_100k': comision_ejemplo,
            'fecha_inicio': tarifa.fecha_inicio_vigencia,
        })
    
    # Estadísticas de comisiones
    hoy = timezone.now().date()
    mes_actual = hoy.replace(day=1)
    
    stats_mes = DetalleComisionVenta.objects.filter(
        id_pago_venta__fecha_pago__gte=mes_actual
    ).aggregate(
        total_comisiones=Sum('monto_comision_calculada'),
        total_transacciones=Count('id_detalle_comision')
    )
    
    context = {
        'medios_comision': medios_comision,
        'tarifas_existentes': tarifas_formateadas,
        'stats_mes': stats_mes,
        'mes_actual': mes_actual,
    }
    
    return render(request, 'pos/configurar_tarifas.html', context)


@login_required
def reporte_comisiones_view(request):
    """Reporte de comisiones por período"""
    # Obtener fechas del filtro
    fecha_desde = request.GET.get('fecha_desde', timezone.now().date().replace(day=1))
    fecha_hasta = request.GET.get('fecha_hasta', timezone.now().date())
    
    # Comisiones del período
    comisiones = DetalleComisionVenta.objects.filter(
        id_pago_venta__fecha_pago__date__gte=fecha_desde,
        id_pago_venta__fecha_pago__date__lte=fecha_hasta
    ).select_related('id_pago_venta', 'id_tarifa').order_by('-id_pago_venta__fecha_pago')
    
    # Resumen
    resumen = comisiones.aggregate(
        total_comision=Sum('monto_comision_calculada'),
        cantidad=Count('id_detalle_comision')
    )
    
    # Por medio de pago (usando la tarifa)
    por_medio = []
    # Agrupar por medio de pago a través de la tarifa
    por_medio_data = comisiones.values(
        'id_tarifa__id_medio_pago__descripcion'
    ).annotate(
        total_comision=Sum('monto_comision_calculada'),
        cantidad=Count('id_detalle_comision')
    ).order_by('-total_comision')
    
    for item in por_medio_data:
        if item['total_comision']:
            por_medio.append({
                'medio': item['id_tarifa__id_medio_pago__descripcion'],
                'total_comision': item['total_comision'],
                'cantidad': item['cantidad']
            })
    
    context = {
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'comisiones': comisiones[:100],  # Limitar a 100 registros
        'total_comision': resumen['total_comision'] or 0,
        'cantidad_total': resumen['cantidad'] or 0,
        'por_medio': por_medio,
    }
    
    return render(request, 'pos/reporte_comisiones.html', context)


@require_http_methods(["POST"])
@login_required
def calcular_comision_recarga(monto, id_medio_pago):
    """
    Función auxiliar para calcular comisión en recargas
    Esta función debe ser llamada desde la vista de recargas
    """
    try:
        medio_pago = MediosPago.objects.get(id_medio_pago=id_medio_pago)
        
        # Solo aplicar comisión a estos medios
        if medio_pago.descripcion not in ['Tarjeta de Crédito', 'Tarjeta de Débito', 'Giros Tigo']:
            return Decimal('0')
        
        # Obtener tarifa
        tarifa = TarifasComision.objects.filter(
            id_medio_pago=medio_pago,
            activo=True
        ).first()
        
        if not tarifa:
            return Decimal('0')
        
        # Calcular comisión: monto_fijo + (monto * porcentaje / 100)
        comision = tarifa.monto_fijo + (Decimal(monto) * tarifa.porcentaje / Decimal('100'))
        
        return comision.quantize(Decimal('0.01'))
        
    except MediosPago.DoesNotExist:
        return Decimal('0')
    except Exception:
        return Decimal('0')


# ==================== SISTEMA DE ALMUERZOS ====================

@login_required
def almuerzos_dashboard_view(request):
    """Dashboard principal del sistema de almuerzos"""
    from gestion.models import PlanesAlmuerzo, SuscripcionesAlmuerzo, RegistroConsumoAlmuerzo, PagosAlmuerzoMensual
    from django.db.models import Count, Sum, Q
    
    hoy = timezone.now().date()
    
    # Estadísticas del día actual
    consumos_hoy = RegistroConsumoAlmuerzo.objects.filter(fecha_consumo=hoy).count()
    
    suscripciones_activas = SuscripcionesAlmuerzo.objects.filter(
        Q(estado='Activa') & 
        (Q(fecha_fin__gte=hoy) | Q(fecha_fin__isnull=True))
    ).count()
    
    # Ingresos del mes
    mes_actual = hoy.month
    anio_actual = hoy.year
    inicio_mes = hoy.replace(day=1)
    
    ingresos_mes = PagosAlmuerzoMensual.objects.filter(
        fecha_pago__year=anio_actual,
        fecha_pago__month=mes_actual,
        estado='Pagado'
    ).aggregate(total=Sum('monto_pagado'))['total'] or 0
    
    # Asistencia semanal (últimos 7 días)
    hace_7_dias = hoy - timezone.timedelta(days=7)
    consumos_semana = RegistroConsumoAlmuerzo.objects.filter(
        fecha_consumo__gte=hace_7_dias
    ).count()
    
    if suscripciones_activas > 0:
        asistencia_semanal = (consumos_semana / (suscripciones_activas * 7)) * 100
    else:
        asistencia_semanal = 0
    
    # Planes activos con número de suscriptores
    planes_activos = PlanesAlmuerzo.objects.filter(activo=True).annotate(
        num_suscriptores=Count('suscripcionesalmuerzo', filter=Q(suscripcionesalmuerzo__estado='Activa'))
    )
    
    # Suscripciones próximas a vencer (7 días)
    fecha_limite = hoy + timezone.timedelta(days=7)
    suscripciones_proximas = SuscripcionesAlmuerzo.objects.filter(
        estado='Activa',
        fecha_fin__lte=fecha_limite,
        fecha_fin__gte=hoy
    ).select_related('id_hijo', 'id_plan_almuerzo')[:10]
    
    # Agregar días restantes
    for sus in suscripciones_proximas:
        sus.dias_restantes = (sus.fecha_fin - hoy).days
    
    # Consumo de últimos 7 días
    consumo_semanal = []
    for i in range(7):
        dia = hoy - timezone.timedelta(days=6-i)
        total = RegistroConsumoAlmuerzo.objects.filter(fecha_consumo=dia).count()
        
        if suscripciones_activas > 0:
            porcentaje = (total / suscripciones_activas) * 100
        else:
            porcentaje = 0
            
        consumo_semanal.append({
            'fecha': dia,
            'total': total,
            'porcentaje': round(porcentaje, 1),
            'ingreso': 0  # Calcular basado en precio promedio si se necesita
        })
    
    context = {
        'stats': {
            'consumos_hoy': consumos_hoy,
            'suscripciones_activas': suscripciones_activas,
            'ingresos_mes': ingresos_mes,
            'asistencia_semanal': round(asistencia_semanal, 1),
            'fecha_hoy': hoy.strftime('%d/%m/%Y'),
            'mes_actual': hoy.strftime('%B %Y')
        },
        'planes_activos': planes_activos,
        'suscripciones_proximas': suscripciones_proximas,
        'consumo_semanal': consumo_semanal,
    }
    
    return render(request, 'gestion/almuerzos_dashboard.html', context)


@login_required
def planes_almuerzo_view(request):
    """Gestión de planes de almuerzo"""
    from gestion.models import PlanesAlmuerzo
    from django.db.models import Count, Q
    
    planes = PlanesAlmuerzo.objects.annotate(
        num_suscriptores=Count('suscripcionesalmuerzo', filter=Q(suscripcionesalmuerzo__estado='Activa'))
    ).order_by('-activo', 'nombre_plan')
    
    context = {
        'planes': planes,
    }
    
    return render(request, 'gestion/planes_almuerzo.html', context)


@login_required
@require_http_methods(["POST"])
def crear_plan_almuerzo(request):
    """Crear nuevo plan de almuerzo"""
    from gestion.models import PlanesAlmuerzo
    
    try:
        data = json.loads(request.body)
        
        plan = PlanesAlmuerzo.objects.create(
            nombre_plan=data['nombre_plan'],
            descripcion=data.get('descripcion', ''),
            precio_mensual=Decimal(data['precio']),
            dias_semana_incluidos=data.get('dias_incluidos', 'Lunes a Viernes'),
            activo=data.get('activo', True),
            fecha_creacion=timezone.now()
        )
        
        return JsonResponse({
            'success': True,
            'mensaje': 'Plan creado correctamente',
            'id_plan': plan.id_plan_almuerzo
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def editar_plan_almuerzo(request, plan_id):
    """Editar plan de almuerzo existente"""
    from gestion.models import PlanesAlmuerzo
    
    try:
        plan = PlanesAlmuerzo.objects.get(id_plan_almuerzo=plan_id)
        data = json.loads(request.body)
        
        if 'nombre_plan' in data:
            plan.nombre_plan = data['nombre_plan']
        if 'descripcion' in data:
            plan.descripcion = data['descripcion']
        if 'precio' in data:
            plan.precio_mensual = Decimal(data['precio'])
        if 'dias_incluidos' in data:
            plan.dias_semana_incluidos = data['dias_incluidos']
        if 'activo' in data:
            plan.activo = data['activo']
        
        plan.save()
        
        return JsonResponse({
            'success': True,
            'mensaje': 'Plan actualizado correctamente'
        })
        
    except PlanesAlmuerzo.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Plan no encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def suscripciones_almuerzo_view(request):
    """Gestión de suscripciones de almuerzos"""
    from gestion.models import SuscripcionesAlmuerzo, PlanesAlmuerzo, Hijo
    from django.db.models import Q
    
    hoy = timezone.now().date()
    
    # Filtros
    estado_filtro = request.GET.get('estado', '')
    busqueda = request.GET.get('estudiante', '')
    grado_filtro = request.GET.get('grado', '')
    plan_filtro = request.GET.get('plan', '')
    
    suscripciones = SuscripcionesAlmuerzo.objects.select_related(
        'id_hijo', 'id_plan_almuerzo', 'id_hijo__id_cliente_responsable'
    )
    
    if estado_filtro:
        suscripciones = suscripciones.filter(estado=estado_filtro)
    
    if busqueda:
        suscripciones = suscripciones.filter(
            Q(id_hijo__nombre__icontains=busqueda) |
            Q(id_hijo__apellido__icontains=busqueda)
        )
    
    if grado_filtro:
        suscripciones = suscripciones.filter(id_hijo__grado=grado_filtro)
    
    if plan_filtro:
        suscripciones = suscripciones.filter(id_plan_almuerzo_id=plan_filtro)
    
    suscripciones = suscripciones.order_by('-fecha_inicio')[:100]
    
    # Agregar días restantes
    for sus in suscripciones:
        if sus.fecha_fin:
            sus.dias_restantes = (sus.fecha_fin - hoy).days
        else:
            sus.dias_restantes = 999
    
    # Estadísticas
    activas = SuscripcionesAlmuerzo.objects.filter(
        estado='Activa',
        fecha_fin__gte=hoy
    ).count()
    
    por_vencer = SuscripcionesAlmuerzo.objects.filter(
        estado='Activa',
        fecha_fin__lte=hoy + timezone.timedelta(days=7),
        fecha_fin__gte=hoy
    ).count()
    
    vencidas = SuscripcionesAlmuerzo.objects.filter(
        fecha_fin__lt=hoy
    ).count()
    
    # Datos para formulario
    planes_activos = PlanesAlmuerzo.objects.filter(activo=True)
    grados = Hijo.objects.values_list('grado', flat=True).distinct().order_by('grado')
    
    context = {
        'suscripciones': suscripciones,
        'planes': planes_activos,
        'grados': grados,
        'estadisticas': {
            'activas': activas,
            'por_vencer': por_vencer,
            'vencidas': vencidas,
            'ingreso': 0  # Calcular si es necesario
        },
    }
    
    return render(request, 'gestion/suscripciones_almuerzo.html', context)


@login_required
@require_http_methods(["POST"])
def crear_suscripcion_almuerzo(request):
    """Crear nueva suscripción a plan de almuerzo"""
    from gestion.models import SuscripcionesAlmuerzo, PlanesAlmuerzo, Hijo
    
    try:
        data = json.loads(request.body)
        
        hijo = Hijo.objects.get(id_hijo=data['id_hijo'])
        plan = PlanesAlmuerzo.objects.get(id_plan_almuerzo=data['id_plan_almuerzo'])
        
        # Calcular fecha fin (30 días por defecto para plan mensual)
        if 'fecha_inicio' in data:
            from datetime import datetime
            fecha_inicio = datetime.strptime(data['fecha_inicio'], '%Y-%m-%d').date()
        else:
            fecha_inicio = timezone.now().date()
        
        fecha_fin = fecha_inicio + timezone.timedelta(days=30)
        
        # Verificar si ya tiene suscripción activa
        suscripcion_existente = SuscripcionesAlmuerzo.objects.filter(
            id_hijo=hijo,
            estado='Activa',
            fecha_fin__gte=fecha_inicio
        ).first()
        
        if suscripcion_existente:
            return JsonResponse({
                'success': False,
                'error': 'El estudiante ya tiene una suscripción activa'
            }, status=400)
        
        # Crear suscripción
        suscripcion = SuscripcionesAlmuerzo.objects.create(
            id_hijo=hijo,
            id_plan_almuerzo=plan,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            estado='Activa'
        )
        
        return JsonResponse({
            'success': True,
            'mensaje': f'Suscripción creada para {hijo.nombres} {hijo.apellidos}',
            'id_suscripcion': suscripcion.id_suscripcion
        })
        
    except Hijo.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Estudiante no encontrado'
        }, status=404)
    except PlanesAlmuerzo.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Plan no encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def registro_consumo_almuerzo_view(request):
    """Registro diario de consumo de almuerzos"""
    from gestion.models import RegistroConsumoAlmuerzo, SuscripcionesAlmuerzo
    from django.db.models import Q
    
    hoy = timezone.now().date()
    
    # Consumos del día
    consumos_hoy = RegistroConsumoAlmuerzo.objects.filter(
        fecha_consumo=hoy
    ).select_related(
        'id_hijo',
        'id_suscripcion__id_plan_almuerzo'
    ).order_by('id_hijo__apellidos')
    
    # Suscripciones activas disponibles para registrar
    suscripciones_activas = SuscripcionesAlmuerzo.objects.filter(
        Q(estado='Activa') &
        (Q(fecha_fin__gte=hoy) | Q(fecha_fin__isnull=True))
    ).select_related('id_hijo', 'id_plan_almuerzo').order_by('id_hijo__apellidos')
    
    # Excluir los que ya consumieron hoy
    ids_consumidos = consumos_hoy.values_list('id_hijo_id', flat=True)
    suscripciones_disponibles = suscripciones_activas.exclude(
        id_hijo_id__in=ids_consumidos
    )
    
    context = {
        'consumos_hoy': consumos_hoy,
        'suscripciones_disponibles': suscripciones_disponibles,
        'total_consumos': consumos_hoy.count(),
        'total_disponibles': suscripciones_disponibles.count(),
    }
    
    return render(request, 'gestion/registro_consumo_almuerzo.html', context)


@login_required
@require_http_methods(["POST"])
def registrar_consumo_almuerzo(request):
    """Registrar consumo de almuerzo"""
    from gestion.models import RegistroConsumoAlmuerzo, SuscripcionesAlmuerzo
    
    try:
        data = json.loads(request.body)
        
        # Soportar registro músltiple
        if 'registros' in data:
            registros = data['registros']
        else:
            # Registro simple
            registros = [{
                'id_suscripcion': data['id_suscripcion']
            }]
        
        registrados = 0
        hoy = timezone.now().date()
        hora_actual = timezone.now().time()
        
        for reg_data in registros:
            suscripcion = SuscripcionesAlmuerzo.objects.select_related(
                'id_hijo', 'id_plan_almuerzo'
            ).get(id_suscripcion=reg_data['id_suscripcion'])
            
            # Validar que la suscripción esté activa
            if suscripcion.estado != 'Activa':
                continue
            
            # Validar que no esté vencida
            if suscripcion.fecha_fin and suscripcion.fecha_fin < hoy:
                continue
            
            # Verificar si ya consumió hoy
            consumo_existente = RegistroConsumoAlmuerzo.objects.filter(
                id_hijo=suscripcion.id_hijo,
                fecha_consumo=hoy
            ).first()
            
            if consumo_existente:
                continue
            
            # Registrar consumo
            RegistroConsumoAlmuerzo.objects.create(
                id_hijo=suscripcion.id_hijo,
                id_suscripcion=suscripcion,
                fecha_consumo=hoy,
                hora_registro=hora_actual
            )
            
            registrados += 1
        
        return JsonResponse({
            'success': True,
            'mensaje': f'{registrados} consumo(s) registrado(s)',
            'registrados': registrados
        })
        
    except SuscripcionesAlmuerzo.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Suscripción no encontrada'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def menu_diario_view(request):
    """Gestión de menús diarios"""
    from gestion.models import SuscripcionesAlmuerzo, RegistroConsumoAlmuerzo
    from django.db.models import Q
    
    # Obtener fecha del request o usar hoy
    fecha_str = request.GET.get('fecha', '')
    if fecha_str:
        from datetime import datetime
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    else:
        fecha = timezone.now().date()
    
    # Suscripciones activas para la fecha
    suscripciones_activas = SuscripcionesAlmuerzo.objects.filter(
        Q(estado='Activa') &
        (Q(fecha_fin__gte=fecha) | Q(fecha_fin__isnull=True)),
        fecha_inicio__lte=fecha
    ).count()
    
    # Consumo esperado (estimado 85%)
    consumo_esperado = 85
    
    context = {
        'fecha': fecha,
        'suscripciones_activas': suscripciones_activas,
        'consumo_esperado': consumo_esperado,
        'menu': {
            'entrada': 'Ensalada mixta',
            'principal': 'Pollo a la plancha con vegetales',
            'acompanamiento': 'Arroz integral',
            'bebida': 'Jugo natural de naranja',
            'postre': 'Fruta de estación',
            'notas': ''
        }
    }
    
    return render(request, 'gestion/menu_diario.html', context)


@login_required
def facturacion_mensual_almuerzos_view(request):
    """Facturación mensual de almuerzos"""
    from gestion.models import PagosAlmuerzoMensual, SuscripcionesAlmuerzo, RegistroConsumoAlmuerzo
    from django.db.models import Sum, Count
    
    # Obtener mes y año del filtro
    mes = int(request.GET.get('mes', timezone.now().month))
    anio = int(request.GET.get('anio', timezone.now().year))
    
    # Crear fecha para filtrar
    from datetime import datetime
    fecha_inicio = datetime(anio, mes, 1).date()
    if mes == 12:
        fecha_fin = datetime(anio + 1, 1, 1).date()
    else:
        fecha_fin = datetime(anio, mes + 1, 1).date()
    
    # Pagos del mes
    pagos_mes = PagosAlmuerzoMensual.objects.filter(
        mes_pagado__gte=fecha_inicio,
        mes_pagado__lt=fecha_fin
    ).select_related(
        'id_suscripcion__id_hijo',
        'id_suscripcion__id_plan_almuerzo'
    ).order_by('-fecha_pago')
    
    # Estadísticas
    total_facturado = pagos_mes.aggregate(total=Sum('monto_pagado'))['total'] or 0
    total_suscripciones = pagos_mes.count()
    
    # Verificar si ya se facturó este mes
    ya_facturado = pagos_mes.exists()
    
    # Resumen por plan
    from collections import defaultdict
    resumen_por_plan = defaultdict(lambda: {'cantidad': 0, 'total': 0, 'total_consumos': 0})
    
    for pago in pagos_mes:
        plan_nombre = pago.id_suscripcion.id_plan_almuerzo.nombre_plan
        resumen_por_plan[plan_nombre]['cantidad'] += 1
        resumen_por_plan[plan_nombre]['total'] += pago.monto_pagado
    
    resumen_por_plan_list = [
        {'plan': k, **v} for k, v in resumen_por_plan.items()
    ]
    
    # Histórico (últimos 6 meses)
    historico = []
    for i in range(6):
        mes_hist = mes - i
        anio_hist = anio
        if mes_hist <= 0:
            mes_hist += 12
            anio_hist -= 1
        
        fecha_hist_inicio = datetime(anio_hist, mes_hist, 1).date()
        if mes_hist == 12:
            fecha_hist_fin = datetime(anio_hist + 1, 1, 1).date()
        else:
            fecha_hist_fin = datetime(anio_hist, mes_hist + 1, 1).date()
        
        pagos_hist = PagosAlmuerzoMensual.objects.filter(
            mes_pagado__gte=fecha_hist_inicio,
            mes_pagado__lt=fecha_hist_fin
        )
        
        total = pagos_hist.aggregate(total=Sum('monto_pagado'))['total'] or 0
        pagado = pagos_hist.filter(estado='Pagado').aggregate(total=Sum('monto_pagado'))['total'] or 0
        
        if total > 0:
            tasa = (pagado / total) * 100
        else:
            tasa = 0
        
        historico.append({
            'mes': mes_hist,
            'anio': anio_hist,
            'suscripciones': pagos_hist.count(),
            'total_facturado': total,
            'total_pagado': pagado,
            'pendiente': total - pagado,
            'tasa_cobro': round(tasa, 1)
        })
    
    context = {
        'pagos_mes': pagos_mes,
        'estadisticas': {
            'total_facturado': total_facturado,
            'total_suscripciones': total_suscripciones,
        },
        'ya_facturado': ya_facturado,
        'resumen_por_plan': resumen_por_plan_list,
        'historico': historico,
    }
    
    return render(request, 'gestion/facturacion_mensual_almuerzos.html', context)


@login_required
@require_http_methods(["POST"])
def generar_facturacion_mensual(request):
    """Generar facturación mensual automática para todas las suscripciones activas"""
    from gestion.models import PagosAlmuerzoMensual, SuscripcionesAlmuerzo, RegistroConsumoAlmuerzo
    
    try:
        data = json.loads(request.body)
        mes = int(data.get('mes', timezone.now().month))
        anio = int(data.get('anio', timezone.now().year))
        
        # Crear fecha del mes
        from datetime import datetime
        fecha_mes = datetime(anio, mes, 1).date()
        fecha_inicio = fecha_mes
        if mes == 12:
            fecha_fin = datetime(anio + 1, 1, 1).date()
        else:
            fecha_fin = datetime(anio, mes + 1, 1).date()
        
        # Verificar si ya existe facturación para este mes
        facturacion_existente = PagosAlmuerzoMensual.objects.filter(
            mes_pagado__gte=fecha_inicio,
            mes_pagado__lt=fecha_fin
        ).exists()
        
        if facturacion_existente:
            return JsonResponse({
                'success': False,
                'error': f'Ya existe facturación para {mes}/{anio}'
            }, status=400)
        
        # Obtener suscripciones activas del mes
        suscripciones = SuscripcionesAlmuerzo.objects.filter(
            estado='Activa',
            fecha_inicio__lte=fecha_fin,
            fecha_fin__gte=fecha_inicio
        ).select_related('id_hijo', 'id_plan_almuerzo', 'id_hijo__id_cliente')
        
        facturas_creadas = 0
        total_facturado = 0
        
        for suscripcion in suscripciones:
            # Contar días consumidos en el mes
            consumos = RegistroConsumoAlmuerzo.objects.filter(
                id_hijo=suscripcion.id_hijo,
                fecha_consumo__gte=fecha_inicio,
                fecha_consumo__lt=fecha_fin
            ).count()
            
            if consumos > 0:
                # Calcular monto (precio mensual completo)
                monto = suscripcion.id_plan_almuerzo.precio_mensual
                
                # Crear pago
                pago = PagosAlmuerzoMensual.objects.create(
                    id_suscripcion=suscripcion,
                    fecha_pago=timezone.now(),
                    monto_pagado=monto,
                    mes_pagado=fecha_mes,
                    estado='Pendiente'
                )
                
                # El sistema nuevo maneja automáticamente el saldo pendiente
                # a través de PagosAlmuerzoMensual.estado = 'Pendiente'
                # No se requiere entrada manual en cuenta corriente
                
                facturas_creadas += 1
                total_facturado += monto
        
        return JsonResponse({
            'success': True,
            'mensaje': f'Se generaron {facturas_creadas} facturas',
            'procesadas': facturas_creadas,
            'total_facturado': float(total_facturado)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def reportes_almuerzos_view(request):
    """Reportes del sistema de almuerzos"""
    from gestion.models import (
        PlanesAlmuerzo, SuscripcionesAlmuerzo, 
        RegistroConsumoAlmuerzo, PagosAlmuerzoMensual
    )
    from django.db.models import Sum, Count, Avg
    
    # Período del reporte
    fecha_desde = request.GET.get('fecha_desde', timezone.now().date().replace(day=1))
    fecha_hasta = request.GET.get('fecha_hasta', timezone.now().date())
    
    # Consumos por día
    consumos_por_dia = RegistroConsumoAlmuerzo.objects.filter(
        fecha_consumo__gte=fecha_desde,
        fecha_consumo__lte=fecha_hasta
    ).values('fecha_consumo').annotate(
        total=Count('id_registro_consumo')
    ).order_by('fecha_consumo')
    
    # Suscripciones por plan
    suscripciones_por_plan = SuscripcionesAlmuerzo.objects.filter(
        fecha_inicio__gte=fecha_desde,
        fecha_inicio__lte=fecha_hasta
    ).values(
        'id_plan_almuerzo__nombre_plan',
        'id_plan_almuerzo__precio_mensual'
    ).annotate(
        cantidad=Count('id_suscripcion')
    ).order_by('-cantidad')
    
    # Calcular ingresos proyectados por plan
    ingresos_por_plan = []
    for item in suscripciones_por_plan:
        ingresos_por_plan.append({
            'id_plan_almuerzo__nombre_plan': item['id_plan_almuerzo__nombre_plan'],
            'cantidad': item['cantidad'],
            'total_ingresos': item['cantidad'] * item['id_plan_almuerzo__precio_mensual']
        })
    
    # Pagos realizados
    pagos_periodo = PagosAlmuerzoMensual.objects.filter(
        fecha_pago__date__gte=fecha_desde,
        fecha_pago__date__lte=fecha_hasta
    )
    
    total_pagado = pagos_periodo.aggregate(total=Sum('monto_pagado'))['total'] or 0
    
    # Tasa de asistencia
    suscripciones_activas = SuscripcionesAlmuerzo.objects.filter(
        estado='Activa',
        fecha_fin__gte=fecha_desde
    ).count()
    
    dias_periodo = (fecha_hasta - fecha_desde).days + 1
    consumos_esperados = suscripciones_activas * dias_periodo
    consumos_reales = RegistroConsumoAlmuerzo.objects.filter(
        fecha_consumo__gte=fecha_desde,
        fecha_consumo__lte=fecha_hasta
    ).count()
    
    tasa_asistencia = (consumos_reales / consumos_esperados * 100) if consumos_esperados > 0 else 0
    
    # Top 10 consumidores
    top_consumidores = RegistroConsumoAlmuerzo.objects.filter(
        fecha_consumo__gte=fecha_desde,
        fecha_consumo__lte=fecha_hasta
    ).values(
        'id_hijo__nombre',
        'id_hijo__apellido'
    ).annotate(
        total_consumos=Count('id_registro_consumo')
    ).order_by('-total_consumos')[:10]
    
    context = {
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'consumos_por_dia': consumos_por_dia,
        'ingresos_por_plan': ingresos_por_plan,
        'tasa_asistencia': round(tasa_asistencia, 2),
        'consumos_reales': consumos_reales,
        'consumos_esperados': consumos_esperados,
        'total_pagado': total_pagado,
        'top_consumidores': top_consumidores,
    }
    
    return render(request, 'gestion/reportes_almuerzos.html', context)


# =============================================================================
# SISTEMA DE AUTORIZACIONES PARA ANULAR VENTAS Y RECARGAS
# =============================================================================

@login_required
@require_http_methods(["POST"])
def anular_venta(request, venta_id):
    """
    Anula una venta después de validar autorización
    """
    try:
        # Obtener el log de autorización
        log_id = request.POST.get('log_id')
        motivo = request.POST.get('motivo', 'Sin motivo especificado')
        
        if not log_id:
            return JsonResponse({
                'success': False,
                'message': 'Se requiere autorización para anular'
            })
        
        # Verificar que el log sea reciente (últimos 30 segundos)
        try:
            log = LogAutorizacion.objects.get(
                id_log=log_id,
                resultado='EXITOSO',
                fecha_hora__gte=timezone.now() - timezone.timedelta(seconds=30)
            )
        except LogAutorizacion.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Autorización expirada. Escanee nuevamente la tarjeta.'
            })
        
        # Obtener la venta
        venta = get_object_or_404(Ventas, id_venta=venta_id)
        
        # Verificar que no esté ya anulada
        if venta.estado == 'ANULADA':
            return JsonResponse({
                'success': False,
                'message': 'Esta venta ya está anulada'
            })
        
        with transaction.atomic():
            # Guardar información para el log
            cliente_nombre = venta.id_cliente.nombre_completo if venta.id_cliente else 'Genérico'
            total = venta.monto_total
            fecha = venta.fecha
            
            # Si fue venta con tarjeta, devolver el saldo
            if venta.id_hijo and venta.id_hijo.tarjeta:
                tarjeta = venta.id_hijo.tarjeta
                
                # Buscar el consumo asociado
                consumo = ConsumoTarjeta.objects.filter(
                    nro_tarjeta=tarjeta,
                    detalle__contains=f'Venta #{venta.id_venta}'
                ).first()
                
                if consumo:
                    # Devolver el monto a la tarjeta
                    tarjeta.saldo_actual = F('saldo_actual') + consumo.monto_consumido
                    tarjeta.save()
                    
                    # Crear registro de devolución en ConsumoTarjeta
                    tarjeta.refresh_from_db()
                    ConsumoTarjeta.objects.create(
                        nro_tarjeta=tarjeta,
                        fecha_consumo=timezone.now(),
                        monto_consumido=-consumo.monto_consumido,  # Negativo para devolución
                        detalle=f'Devolución - Anulación Venta #{venta.id_venta}. Motivo: {motivo}',
                        saldo_anterior=tarjeta.saldo_actual - consumo.monto_consumido,
                        saldo_posterior=tarjeta.saldo_actual
                    )
            
            # Restaurar stock de productos vendidos
            detalles = DetalleVenta.objects.filter(id_venta=venta)
            for detalle in detalles:
                try:
                    stock = StockUnico.objects.get(id_producto=detalle.id_producto)
                    stock.stock_actual = F('stock_actual') + detalle.cantidad
                    stock.save()
                except StockUnico.DoesNotExist:
                    pass
            
            # Marcar venta como anulada
            venta.estado = 'ANULADA'
            venta.observaciones = f'ANULADA: {motivo}. Autorizado el {timezone.now().strftime("%d/%m/%Y %H:%M")}'
            venta.save()
            
            # Actualizar el log de autorización
            log.descripcion += f' | Anulada: Venta #{venta.id_venta} - {cliente_nombre} - Gs. {total:,.0f} ({fecha}). Motivo: {motivo}'
            log.id_registro_afectado = venta_id
            log.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Venta anulada correctamente. #{venta.id_venta} - Gs. {total:,.0f}'
            })
        
    except Exception as e:
        # Registrar error en el log si existe
        if log_id:
            try:
                log = LogAutorizacion.objects.get(id_log=log_id)
                log.resultado = 'ERROR'
                log.descripcion += f' | Error: {str(e)}'
                log.save()
            except:
                pass
        
        return JsonResponse({
            'success': False,
            'message': f'Error al anular venta: {str(e)}'
        })


@login_required
@require_http_methods(["POST"])
def anular_recarga(request, recarga_id):
    """
    Anula una recarga después de validar autorización
    """
    try:
        # Obtener el log de autorización
        log_id = request.POST.get('log_id')
        motivo = request.POST.get('motivo', 'Sin motivo especificado')
        
        if not log_id:
            return JsonResponse({
                'success': False,
                'message': 'Se requiere autorización para anular'
            })
        
        # Verificar que el log sea reciente (últimos 30 segundos)
        try:
            log = LogAutorizacion.objects.get(
                id_log=log_id,
                resultado='EXITOSO',
                fecha_hora__gte=timezone.now() - timezone.timedelta(seconds=30)
            )
        except LogAutorizacion.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Autorización expirada. Escanee nuevamente la tarjeta.'
            })
        
        # Obtener la recarga
        recarga = get_object_or_404(CargasSaldo, id_carga_saldo=recarga_id)
        
        # Verificar que no esté ya anulada
        if hasattr(recarga, 'anulada') and recarga.anulada:
            return JsonResponse({
                'success': False,
                'message': 'Esta recarga ya está anulada'
            })
        
        with transaction.atomic():
            # Guardar información para el log
            tarjeta = recarga.nro_tarjeta
            hijo_nombre = tarjeta.id_hijo.nombre_completo if tarjeta.id_hijo else 'N/A'
            monto = recarga.monto_cargado
            fecha = recarga.fecha_carga
            
            # Descontar el monto de la tarjeta
            tarjeta.saldo_actual = F('saldo_actual') - monto
            tarjeta.save()
            
            # Crear registro de devolución en ConsumoTarjeta
            tarjeta.refresh_from_db()
            ConsumoTarjeta.objects.create(
                nro_tarjeta=tarjeta,
                fecha_consumo=timezone.now(),
                monto_consumido=monto,  # Positivo porque se descuenta
                detalle=f'Anulación Recarga #{recarga.id_carga_saldo}. Motivo: {motivo}',
                saldo_anterior=tarjeta.saldo_actual + monto,
                saldo_posterior=tarjeta.saldo_actual
            )
            
            # Marcar como anulada (o eliminar según lógica de negocio)
            try:
                recarga.anulada = True
                recarga.fecha_anulacion = timezone.now()
                recarga.motivo_anulacion = motivo
                recarga.save()
            except AttributeError:
                # Si no existe el campo, eliminar
                recarga.delete()
            
            # Actualizar el log de autorización
            log.descripcion += f' | Anulada: Recarga #{recarga_id} - {hijo_nombre} - Gs. {monto:,.0f} ({fecha}). Motivo: {motivo}'
            log.id_registro_afectado = recarga_id
            log.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Recarga anulada correctamente. #{recarga_id} - Gs. {monto:,.0f}'
            })
        
    except Exception as e:
        # Registrar error en el log si existe
        if log_id:
            try:
                log = LogAutorizacion.objects.get(id_log=log_id)
                log.resultado = 'ERROR'
                log.descripcion += f' | Error: {str(e)}'
                log.save()
            except:
                pass
        
        return JsonResponse({
            'success': False,
            'message': f'Error al anular recarga: {str(e)}'
        })


# =============================================================================
# ADMINISTRACIÓN DE TARJETAS DE AUTORIZACIÓN
# =============================================================================

@login_required
@require_http_methods(["GET"])
def admin_tarjetas_autorizacion(request):
    """
    Vista principal de administración de tarjetas de autorización
    """
    # Obtener todas las tarjetas
    tarjetas = TarjetaAutorizacion.objects.select_related('id_empleado').order_by('-activo', 'tipo_autorizacion')
    
    # Estadísticas
    total_tarjetas = tarjetas.count()
    tarjetas_activas = tarjetas.filter(activo=True).count()
    tarjetas_admin = tarjetas.filter(tipo_autorizacion='ADMIN').count()
    tarjetas_supervisor = tarjetas.filter(tipo_autorizacion='SUPERVISOR').count()
    
    # Últimas autorizaciones (últimas 50)
    ultimos_logs = LogAutorizacion.objects.select_related(
        'id_tarjeta_autorizacion'
    ).order_by('-fecha_hora')[:50]
    
    context = {
        'tarjetas': tarjetas,
        'total_tarjetas': total_tarjetas,
        'tarjetas_activas': tarjetas_activas,
        'tarjetas_admin': tarjetas_admin,
        'tarjetas_supervisor': tarjetas_supervisor,
        'ultimos_logs': ultimos_logs,
    }
    
    return render(request, 'pos/admin_autorizaciones.html', context)


@login_required
@require_http_methods(["POST"])
def crear_tarjeta_autorizacion(request):
    """
    Crear una nueva tarjeta de autorización
    """
    try:
        codigo_barra = request.POST.get('codigo_barra', '').strip()
        tipo_autorizacion = request.POST.get('tipo_autorizacion', 'SUPERVISOR')
        id_empleado = request.POST.get('id_empleado')
        
        # Permisos
        puede_anular_almuerzos = request.POST.get('puede_anular_almuerzos') == 'on'
        puede_anular_ventas = request.POST.get('puede_anular_ventas') == 'on'
        puede_anular_recargas = request.POST.get('puede_anular_recargas') == 'on'
        puede_modificar_precios = request.POST.get('puede_modificar_precios') == 'on'
        
        observaciones = request.POST.get('observaciones', '')
        fecha_vencimiento = request.POST.get('fecha_vencimiento') or None
        
        if not codigo_barra:
            return JsonResponse({
                'success': False,
                'message': 'El código de barra es obligatorio'
            })
        
        # Verificar que no exista
        if TarjetaAutorizacion.objects.filter(codigo_barra=codigo_barra).exists():
            return JsonResponse({
                'success': False,
                'message': 'Ya existe una tarjeta con este código'
            })
        
        # Obtener empleado si se especificó
        empleado = None
        if id_empleado:
            try:
                empleado = Empleado.objects.get(id_empleado=id_empleado)
            except Empleado.DoesNotExist:
                pass
        
        # Crear tarjeta
        tarjeta = TarjetaAutorizacion.objects.create(
            codigo_barra=codigo_barra,
            id_empleado=empleado,
            tipo_autorizacion=tipo_autorizacion,
            puede_anular_almuerzos=puede_anular_almuerzos,
            puede_anular_ventas=puede_anular_ventas,
            puede_anular_recargas=puede_anular_recargas,
            puede_modificar_precios=puede_modificar_precios,
            observaciones=observaciones,
            fecha_vencimiento=fecha_vencimiento,
            activo=True
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Tarjeta {codigo_barra} creada exitosamente',
            'tarjeta_id': tarjeta.id_tarjeta_autorizacion
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al crear tarjeta: {str(e)}'
        })


@login_required
@require_http_methods(["POST"])
def editar_tarjeta_autorizacion(request, tarjeta_id):
    """
    Editar una tarjeta de autorización existente
    """
    try:
        tarjeta = get_object_or_404(TarjetaAutorizacion, id_tarjeta_autorizacion=tarjeta_id)
        
        # Actualizar campos
        tarjeta.tipo_autorizacion = request.POST.get('tipo_autorizacion', tarjeta.tipo_autorizacion)
        
        # Permisos
        tarjeta.puede_anular_almuerzos = request.POST.get('puede_anular_almuerzos') == 'on'
        tarjeta.puede_anular_ventas = request.POST.get('puede_anular_ventas') == 'on'
        tarjeta.puede_anular_recargas = request.POST.get('puede_anular_recargas') == 'on'
        tarjeta.puede_modificar_precios = request.POST.get('puede_modificar_precios') == 'on'
        
        tarjeta.observaciones = request.POST.get('observaciones', '')
        fecha_vencimiento = request.POST.get('fecha_vencimiento')
        tarjeta.fecha_vencimiento = fecha_vencimiento if fecha_vencimiento else None
        
        # Empleado
        id_empleado = request.POST.get('id_empleado')
        if id_empleado:
            try:
                tarjeta.id_empleado = Empleado.objects.get(id_empleado=id_empleado)
            except Empleado.DoesNotExist:
                tarjeta.id_empleado = None
        else:
            tarjeta.id_empleado = None
        
        tarjeta.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Tarjeta {tarjeta.codigo_barra} actualizada'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al editar tarjeta: {str(e)}'
        })


@login_required
@require_http_methods(["POST"])
def toggle_tarjeta_autorizacion(request, tarjeta_id):
    """
    Activar/desactivar una tarjeta de autorización
    """
    try:
        tarjeta = get_object_or_404(TarjetaAutorizacion, id_tarjeta_autorizacion=tarjeta_id)
        tarjeta.activo = not tarjeta.activo
        tarjeta.save()
        
        estado = "activada" if tarjeta.activo else "desactivada"
        
        return JsonResponse({
            'success': True,
            'message': f'Tarjeta {tarjeta.codigo_barra} {estado}',
            'activo': tarjeta.activo
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al cambiar estado: {str(e)}'
        })


@login_required
@require_http_methods(["GET"])
def ver_logs_autorizacion(request):
    """
    Vista de logs de autorización con filtros
    """
    # Filtros
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')
    tipo_operacion = request.GET.get('tipo_operacion')
    resultado = request.GET.get('resultado')
    tarjeta_id = request.GET.get('tarjeta_id')
    
    logs = LogAutorizacion.objects.select_related(
        'id_tarjeta_autorizacion'
    ).order_by('-fecha_hora')
    
    # Aplicar filtros
    if fecha_desde:
        logs = logs.filter(fecha_hora__gte=fecha_desde)
    if fecha_hasta:
        logs = logs.filter(fecha_hora__lte=fecha_hasta)
    if tipo_operacion:
        logs = logs.filter(tipo_operacion=tipo_operacion)
    if resultado:
        logs = logs.filter(resultado=resultado)
    if tarjeta_id:
        logs = logs.filter(id_tarjeta_autorizacion_id=tarjeta_id)
    
    # Paginación
    from django.core.paginator import Paginator
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page', 1)
    logs_page = paginator.get_page(page_number)
    
    # Estadísticas
    total_logs = logs.count()
    exitosos = logs.filter(resultado='EXITOSO').count()
    rechazados = logs.filter(resultado='RECHAZADO').count()
    
    # Obtener todas las tarjetas para el filtro
    tarjetas = TarjetaAutorizacion.objects.filter(activo=True).order_by('codigo_barra')
    
    context = {
        'logs': logs_page,
        'total_logs': total_logs,
        'exitosos': exitosos,
        'rechazados': rechazados,
        'tarjetas': tarjetas,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'tipo_operacion': tipo_operacion,
        'resultado': resultado,
        'tarjeta_id': tarjeta_id,
    }
    
    return render(request, 'pos/logs_autorizaciones.html', context)


# =============================================================================
# GESTIÓN DE FOTOS DE HIJOS PARA IDENTIFICACIÓN EN POS
# =============================================================================

@login_required
@require_http_methods(["GET"])
def gestionar_fotos_hijos(request):
    """
    Vista para gestionar fotos de identificación de hijos
    """
    # Filtros
    buscar = request.GET.get('buscar', '')
    sin_foto = request.GET.get('sin_foto', '')
    
    hijos = Hijo.objects.filter(activo=True).select_related(
        'id_cliente_responsable',
        'tarjeta'
    ).order_by('apellido', 'nombre')
    
    if buscar:
        hijos = hijos.filter(
            Q(nombre__icontains=buscar) | 
            Q(apellido__icontains=buscar) |
            Q(tarjeta__nro_tarjeta__icontains=buscar)
        )
    
    if sin_foto:
        hijos = hijos.filter(Q(foto_perfil__isnull=True) | Q(foto_perfil=''))
    
    # Estadísticas
    total_hijos = Hijo.objects.filter(activo=True).count()
    con_foto = Hijo.objects.filter(activo=True).exclude(
        Q(foto_perfil__isnull=True) | Q(foto_perfil='')
    ).count()
    sin_foto_count = total_hijos - con_foto
    
    context = {
        'hijos': hijos,
        'total_hijos': total_hijos,
        'con_foto': con_foto,
        'sin_foto_count': sin_foto_count,
        'buscar': buscar,
        'filtro_sin_foto': sin_foto,
    }
    
    return render(request, 'pos/gestionar_fotos.html', context)


@login_required
@require_http_methods(["POST"])
@csrf_exempt
def capturar_foto_hijo(request, hijo_id):
    """
    Captura y guarda la foto de un hijo desde webcam
    """
    try:
        hijo = get_object_or_404(Hijo, id_hijo=hijo_id)
        
        # Obtener imagen en base64 del POST
        imagen_data = request.POST.get('imagen')
        
        if not imagen_data:
            return JsonResponse({
                'success': False,
                'message': 'No se recibió la imagen'
            })
        
        # Remover el prefijo data:image/png;base64,
        if ',' in imagen_data:
            imagen_data = imagen_data.split(',')[1]
        
        # Decodificar base64
        try:
            imagen_bytes = base64.b64decode(imagen_data)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error al decodificar imagen: {str(e)}'
            })
        
        # Crear directorio si no existe
        fotos_dir = os.path.join(settings.MEDIA_ROOT, 'fotos_hijos')
        os.makedirs(fotos_dir, exist_ok=True)
        
        # Nombre de archivo único
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f'hijo_{hijo.id_hijo}_{timestamp}.png'
        ruta_completa = os.path.join(fotos_dir, nombre_archivo)
        
        # Guardar archivo
        with open(ruta_completa, 'wb') as f:
            f.write(imagen_bytes)
        
        # Eliminar foto anterior si existe
        if hijo.foto_perfil:
            ruta_anterior = os.path.join(settings.MEDIA_ROOT, hijo.foto_perfil)
            if os.path.exists(ruta_anterior):
                try:
                    os.remove(ruta_anterior)
                except:
                    pass
        
        # Actualizar modelo
        hijo.foto_perfil = f'fotos_hijos/{nombre_archivo}'
        hijo.fecha_foto = timezone.now()
        hijo.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Foto de {hijo.nombre_completo} guardada exitosamente',
            'foto_url': settings.MEDIA_URL + hijo.foto_perfil
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al guardar foto: {str(e)}'
        })


@login_required
@require_http_methods(["POST"])
def eliminar_foto_hijo(request, hijo_id):
    """
    Elimina la foto de un hijo
    """
    try:
        hijo = get_object_or_404(Hijo, id_hijo=hijo_id)
        
        if not hijo.foto_perfil:
            return JsonResponse({
                'success': False,
                'message': 'Este hijo no tiene foto'
            })
        
        # Eliminar archivo físico
        ruta_foto = os.path.join(settings.MEDIA_ROOT, hijo.foto_perfil)
        if os.path.exists(ruta_foto):
            try:
                os.remove(ruta_foto)
            except Exception as e:
                print(f"Error al eliminar archivo: {e}")
        
        # Limpiar campos en BD
        hijo.foto_perfil = None
        hijo.fecha_foto = None
        hijo.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Foto de {hijo.nombre_completo} eliminada'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al eliminar foto: {str(e)}'
        })


@login_required
@require_http_methods(["GET"])
def obtener_foto_hijo(request):
    """
    Obtiene la foto de un hijo por número de tarjeta (para POS)
    """
    nro_tarjeta = request.GET.get('nro_tarjeta', '').strip()
    
    if not nro_tarjeta:
        return JsonResponse({
            'success': False,
            'message': 'Número de tarjeta requerido'
        })
    
    try:
        tarjeta = Tarjeta.objects.select_related('id_hijo').get(nro_tarjeta=nro_tarjeta)
        hijo = tarjeta.id_hijo
        
        response_data = {
            'success': True,
            'hijo': {
                'id': hijo.id_hijo,
                'nombre': hijo.nombre_completo,
                'tiene_foto': hijo.tiene_foto,
                'foto_url': settings.MEDIA_URL + hijo.foto_perfil if hijo.foto_perfil else None,
                'fecha_foto': hijo.fecha_foto.strftime('%d/%m/%Y %H:%M') if hijo.fecha_foto else None
            }
        }
        
        return JsonResponse(response_data)
        
    except Tarjeta.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Tarjeta no encontrada'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })


# =============================================================================
# GESTIÓN DE GRADOS Y PROMOCIONES
# =============================================================================

@login_required
@require_http_methods(["GET"])
def gestionar_grados_view(request):
    """
    Vista para gestionar grados y promociones de estudiantes
    """
    from gestion.models import Grado, HistorialGradoHijo
    from django.db.models import Count
    
    # Filtros
    buscar = request.GET.get('buscar', '')
    grado_filtro = request.GET.get('grado', '')
    
    # Obtener todos los hijos activos
    hijos = Hijo.objects.filter(activo=True).select_related(
        'id_cliente_responsable',
        'tarjeta'
    ).order_by('grado', 'apellido', 'nombre')
    
    if buscar:
        hijos = hijos.filter(
            Q(nombre__icontains=buscar) | 
            Q(apellido__icontains=buscar) |
            Q(tarjeta__nro_tarjeta__icontains=buscar)
        )
    
    if grado_filtro:
        hijos = hijos.filter(grado=grado_filtro)
    
    # Obtener todos los grados activos
    grados = Grado.objects.filter(activo=True).order_by('orden_visualizacion')
    
    # Estadísticas por grado
    estadisticas = hijos.values('grado').annotate(
        total=Count('id_hijo')
    ).order_by('grado')
    
    # Total de estudiantes sin grado asignado
    sin_grado = hijos.filter(Q(grado__isnull=True) | Q(grado='')).count()
    
    context = {
        'hijos': hijos,
        'grados': grados,
        'estadisticas': estadisticas,
        'sin_grado': sin_grado,
        'buscar': buscar,
        'grado_filtro': grado_filtro,
        'anio_actual': timezone.now().year,
    }
    
    return render(request, 'pos/gestionar_grados.html', context)


@login_required
@require_http_methods(["POST"])
def asignar_grado_hijo(request, hijo_id):
    """
    Asignar o cambiar el grado de un hijo
    """
    from gestion.models import HistorialGradoHijo
    
    try:
        hijo = get_object_or_404(Hijo, id_hijo=hijo_id)
        nuevo_grado = request.POST.get('grado', '').strip()
        observaciones = request.POST.get('observaciones', '')
        
        if not nuevo_grado:
            return JsonResponse({
                'success': False,
                'message': 'Debe seleccionar un grado'
            })
        
        grado_anterior = hijo.grado
        hijo.grado = nuevo_grado
        hijo.save()
        
        # Registrar en historial
        HistorialGradoHijo.objects.create(
            id_hijo=hijo,
            grado_anterior=grado_anterior,
            grado_nuevo=nuevo_grado,
            anio_escolar=timezone.now().year,
            motivo='CAMBIO_MANUAL',
            usuario_registro=request.user.username,
            observaciones=observaciones
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Grado actualizado: {hijo.nombre_completo} → {nuevo_grado}'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })


@login_required
@require_http_methods(["POST"])
def promocionar_grado_masivo(request):
    """
    Promocionar masivamente a todos los estudiantes de un grado al siguiente
    """
    from gestion.models import Grado, HistorialGradoHijo
    
    try:
        grado_actual_nombre = request.POST.get('grado_actual', '')
        
        if not grado_actual_nombre:
            return JsonResponse({
                'success': False,
                'message': 'Debe seleccionar el grado a promocionar'
            })
        
        # Buscar el grado actual
        try:
            grado_actual = Grado.objects.get(nombre_grado=grado_actual_nombre, activo=True)
        except Grado.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': f'Grado {grado_actual_nombre} no encontrado'
            })
        
        # Verificar si hay un siguiente grado
        siguiente_grado = grado_actual.siguiente_grado()
        
        if not siguiente_grado:
            return JsonResponse({
                'success': False,
                'message': f'{grado_actual_nombre} es el último grado. No hay siguiente nivel.'
            })
        
        # Obtener todos los hijos en ese grado
        hijos_a_promocionar = Hijo.objects.filter(
            grado=grado_actual_nombre,
            activo=True
        )
        
        total = hijos_a_promocionar.count()
        
        if total == 0:
            return JsonResponse({
                'success': False,
                'message': f'No hay estudiantes en {grado_actual_nombre} para promocionar'
            })
        
        # Promocionar todos
        promocionados = 0
        anio_escolar = timezone.now().year
        
        with transaction.atomic():
            for hijo in hijos_a_promocionar:
                hijo.grado = siguiente_grado.nombre_grado
                hijo.save()
                
                # Registrar en historial
                HistorialGradoHijo.objects.create(
                    id_hijo=hijo,
                    grado_anterior=grado_actual_nombre,
                    grado_nuevo=siguiente_grado.nombre_grado,
                    anio_escolar=anio_escolar,
                    motivo='PROMOCION',
                    usuario_registro=request.user.username,
                    observaciones=f'Promoción masiva de {grado_actual_nombre} a {siguiente_grado.nombre_grado}'
                )
                
                promocionados += 1
        
        return JsonResponse({
            'success': True,
            'message': f'✅ {promocionados} estudiantes promocionados de {grado_actual_nombre} a {siguiente_grado.nombre_grado}',
            'total': promocionados
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        })


@login_required
@require_http_methods(["GET"])
def historial_grados_view(request):
    """
    Ver historial de cambios de grados
    """
    from gestion.models import HistorialGradoHijo
    from django.core.paginator import Paginator
    
    # Filtros
    hijo_id = request.GET.get('hijo')
    anio = request.GET.get('anio', '')
    motivo = request.GET.get('motivo', '')
    
    historial = HistorialGradoHijo.objects.select_related('id_hijo').all()
    
    if hijo_id:
        historial = historial.filter(id_hijo_id=hijo_id)
    
    if anio:
        historial = historial.filter(anio_escolar=anio)
    
    if motivo:
        historial = historial.filter(motivo=motivo)
    
    # Paginación
    paginator = Paginator(historial, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Años disponibles para filtro
    anios_disponibles = HistorialGradoHijo.objects.values_list(
        'anio_escolar', flat=True
    ).distinct().order_by('-anio_escolar')
    
    context = {
        'page_obj': page_obj,
        'anios_disponibles': anios_disponibles,
        'hijo_id': hijo_id,
        'anio': anio,
        'motivo': motivo,
    }
    
    return render(request, 'pos/historial_grados.html', context)


# =============================================================================
# ENDPOINTS PARA MATCHING DE RESTRICCIONES Y PROMOCIONES
# =============================================================================

@login_required
@require_http_methods(["POST"])
def analizar_restriccion_producto(request):
    """
    Endpoint para analizar si un producto tiene conflictos con restricciones alimentarias.
    
    POST params:
        - producto_id: ID del producto
        - restricciones: Texto de restricciones del estudiante
    
    Returns:
        JSON con análisis de conflictos
    """
    try:
        data = json.loads(request.body)
        producto_id = data.get('producto_id')
        restricciones = data.get('restricciones', '')
        
        if not producto_id:
            return JsonResponse({
                'success': False,
                'error': 'producto_id es requerido'
            }, status=400)
        
        # Analizar producto
        analisis = analizar_restricciones_producto(producto_id, restricciones)
        
        return JsonResponse({
            'success': True,
            'analisis': analisis
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def analizar_carrito_restricciones(request):
    """
    Endpoint para analizar todo el carrito contra restricciones.
    
    POST params:
        - items: Array de {producto_id, cantidad, descripcion}
        - restricciones: Texto de restricciones
    
    Returns:
        JSON con análisis completo del carrito
    """
    try:
        data = json.loads(request.body)
        items = data.get('items', [])
        restricciones = data.get('restricciones', '')
        
        if not items:
            return JsonResponse({
                'success': True,
                'analisis': {
                    'tiene_conflictos': False,
                    'productos_con_conflicto': [],
                    'productos_seguros': 0
                }
            })
        
        # Analizar carrito completo
        analisis = analizar_carrito_completo(items, restricciones)
        
        return JsonResponse({
            'success': True,
            'analisis': analisis
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def calcular_promociones_carrito(request):
    """
    Endpoint para calcular promociones disponibles para el carrito actual.
    
    POST params:
        - items: Array de {producto_id, cantidad, precio_unitario, categoria_id}
        - grado_estudiante: Grado del estudiante (opcional)
        - codigo_promocion: Código de promoción (opcional)
    
    Returns:
        JSON con promociones aplicables
    """
    try:
        data = json.loads(request.body)
        items = data.get('items', [])
        grado = data.get('grado_estudiante')
        codigo = data.get('codigo_promocion')
        
        if not items:
            return JsonResponse({
                'success': True,
                'resultado': {
                    'promociones_aplicables': [],
                    'mejor_promocion': None,
                    'descuento_maximo': 0,
                    'subtotal_original': 0,
                    'total_con_descuento': 0
                }
            })
        
        # Calcular promociones
        resultado = calcular_promociones_disponibles(items, grado, codigo)
        
        return JsonResponse({
            'success': True,
            'resultado': resultado
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def validar_supervisor(request):
    """
    Validar que una tarjeta corresponde a un supervisor para autorizar ventas a crédito.
    
    POST params:
        - nro_tarjeta: Número de tarjeta del supervisor
    
    Returns:
        - success: bool
        - nombre: Nombre completo del supervisor
        - id_empleado: ID del empleado supervisor
        - error: Mensaje de error si falla
    """
    try:
        data = json.loads(request.body)
        nro_tarjeta = data.get('nro_tarjeta', '').strip()
        
        if not nro_tarjeta:
            return JsonResponse({
                'success': False,
                'error': 'Debe escanear una tarjeta'
            })
        
        # Buscar tarjeta de supervisor
        try:
            tarjeta = Tarjeta.objects.select_related(
                'id_hijo',
                'id_hijo__id_cliente_responsable'
            ).get(
                nro_tarjeta=nro_tarjeta,
                tipo_autorizacion='SUPERVISOR',
                estado='ACTIVA'
            )
            
            # Obtener el empleado asociado a esta tarjeta de supervisor
            # La tarjeta de supervisor puede estar asociada a un hijo cuyo responsable es el empleado
            # O podemos buscar al empleado por otro criterio
            
            # Buscar empleado por el cliente responsable de la tarjeta
            if tarjeta.id_hijo and tarjeta.id_hijo.id_cliente_responsable:
                cliente = tarjeta.id_hijo.id_cliente_responsable
                
                # Buscar empleado que coincida con el RUC/CI del cliente
                try:
                    empleado = Empleado.objects.get(
                        ci=cliente.ruc_ci,
                        activo=True
                    )
                    
                    # Verificar que el empleado tenga rol de supervisor
                    if empleado.id_rol.nombre_rol not in ['SUPERVISOR', 'ADMINISTRADOR', 'GERENTE']:
                        return JsonResponse({
                            'success': False,
                            'error': 'Esta tarjeta no pertenece a un supervisor autorizado'
                        })
                    
                    return JsonResponse({
                        'success': True,
                        'nombre': f'{empleado.nombre} {empleado.apellido}',
                        'id_empleado': empleado.id_empleado,
                        'rol': empleado.id_rol.nombre_rol
                    })
                    
                except Empleado.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': 'No se encontró empleado asociado a esta tarjeta de supervisor'
                    })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Tarjeta de supervisor no tiene datos asociados'
                })
                
        except Tarjeta.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Tarjeta no encontrada o no es una tarjeta de supervisor activa'
            })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Datos inválidos'
        }, status=400)
    except Exception as e:
        print(f"❌ ERROR al validar supervisor: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Error al validar supervisor: {str(e)}'
        }, status=500)
